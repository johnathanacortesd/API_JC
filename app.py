# ======================================
# Importaciones
# ======================================
import streamlit as st
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, NamedStyle, Alignment
from collections import defaultdict, Counter
from difflib import SequenceMatcher
from copy import deepcopy
import datetime
import io
import openai
import re
import time
from unidecode import unidecode
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
import json
import asyncio
import hashlib
from typing import List, Dict, Tuple, Optional, Any, Callable, Sequence
from dataclasses import dataclass, field
import random
import joblib
import gc
import requests
import os
import zipfile
import xml.etree.ElementTree as ET
import html
from pathlib import Path

# ── Errores reintentables de OpenAI (openai==0.28) ───────────────────────────
# Distinguir un fallo de red de una respuesta vacía evita que un 429 se
# convierta silenciosamente en "Neutro".
try:  # openai==0.28
    from openai.error import (
        APIConnectionError,
        APIError,
        RateLimitError,
        ServiceUnavailableError,
        Timeout,
    )

    _ERRORES_REINTENTABLES: Tuple[type, ...] = (
        RateLimitError,
        APIError,
        Timeout,
        ServiceUnavailableError,
        APIConnectionError,
    )
    _ERROR_RATE_LIMIT: Tuple[type, ...] = (RateLimitError,)
except Exception:  # pragma: no cover
    _ERRORES_REINTENTABLES = (Exception,)
    _ERROR_RATE_LIMIT = ()

MODELO_CLASIFICACION_POR_DEFECTO = "gpt-4.1-nano-2025-04-14"
MOTOR_PRECISION_DISPONIBLE = True

# ======================================
# Configuración general
# ======================================
st.set_page_config(
    page_title="Análisis de Noticias · API - Realizado por Johnathan Cortés",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="collapsed"
)

OPENAI_MODEL_EMBEDDING     = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = "gpt-4.1-nano-2025-04-14"

CONCURRENT_REQUESTS          = 24   # Alineado con el motor de precisión: evita 429 y resultados truncados.
SIMILARITY_THRESHOLD_TONO    = 0.96  # El tono solo se hereda entre republicaciones casi idénticas.
SIMILARITY_THRESHOLD_TITULOS = 0.94

# ── Umbrales base (corpus grande ≥ 20 noticias) ──────────────────────────────
UMBRAL_SUBTEMA = 0.78
UMBRAL_TEMA    = 0.72
NUM_TEMAS_MAX  = 15

UMBRAL_DEDUP_LABEL           = 0.86
UMBRAL_FUSION_SUBTEMAS       = 0.88
UMBRAL_FUSION_INTERGRUPO     = 0.90
MAX_ITER_FUSION              = 3

UMBRAL_MIN_PERTENENCIA_SUBTEMA = 0.60
UMBRAL_MIN_PERTENENCIA_TEMA    = 0.52

UMBRAL_COHERENCIA_ETIQUETA   = 0.35

MAX_GRUPO_ETIQUETA           = 40

# ── Umbrales mínimos de similitud REAL para agrupar ──────────────────────────
SIM_MINIMA_AGRUPACION_SUBTEMA = 0.90
SIM_MINIMA_KEYWORDS_RARAS     = 0.86   
SIM_MINIMA_FUSION_INTER       = 0.90   

PRICE_INPUT_1M     = 0.10
PRICE_OUTPUT_1M    = 0.40
PRICE_EMBEDDING_1M = 0.02

# ── Motor de precisión ───────────────────────────────────────────────────────
# Cambiar a False para volver al pipeline clásico sin tocar nada más.
USAR_MOTOR_PRECISION = True

if 'tokens_input' not in st.session_state: st.session_state['tokens_input']     = 0
if 'tokens_output' not in st.session_state: st.session_state['tokens_output']    = 0
if 'tokens_embedding' not in st.session_state: st.session_state['tokens_embedding'] = 0

STOPWORDS_ES = set("""
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante
para por segun sin so sobre tras y o u e la el los las un una unos unas lo
al del se su sus le les mi mis tu tus nuestro nuestros vuestra vuestras este
esta estos estas ese esa esos esas aquel aquella aquellos aquellas que cual
cuales quien quienes cuyo cuya cuyos cuyas como cuando donde cual es son fue
fueron era eran sera seran seria serian he ha han habia han hay hubo habra
habria estoy esta estan estaba estaban estamos estan estar estare estaria
estuvieron estarian estuvo asi ya mas menos tan tanto cada muy todo toda todos
todas ser haber hacer tener poder deber ir dar ver saber querer llegar pasar
encontrar creer decir poner salir volver seguir llevar sentir cambiar
""".split())

_TRAILING_INCOMPLETE = {
    "de","del","la","el","los","las","un","una","unos","unas","al","su","sus",
    "en","con","sin","por","para","sobre","ante","bajo","contra","desde",
    "entre","hacia","hasta","mediante","tras","y","o","u","e","lo","que","se",
    "como","donde","cuando","cual","cuyo","cuya","cuyos","cuyas",
    "este","esta","estos","estas","ese","esa","esos","esas",
    "aquel","aquella","aquellos","aquellas","cada","todo","toda","todos","todas",
    "otro","otra","otros","otras","nuevo","nueva","nuevos","nuevas",
    "gran","grandes","mayor","mayores","menor","menores","mejor","mejores",
    "peor","peores","primer","primera","segundo","segunda","tercer","tercera",
    "más","mas","muy","tan","tanto","tanta","tantos","tantas",
    "mi","mis","tu","tus","nuestro","nuestra","nuestros","nuestras",
    "a","ha","he","ser","estar","haber","hacer","tener","poder","deber",
    "ir","dar","ver","saber","querer","llegar","pasar","decir","poner",
}

_PATRON_TITULAR = re.compile(
    r"^(nuevo|nueva|anuncia|lanza|presenta|inaugura|llega|abre|inicia|"
    r"logra|alcanza|supera|confirma|destaca|revela|señala|advierte|"
    r"lanzamiento|anuncio|apertura|inicio|presentacion|presentación)\b",
    re.IGNORECASE
)
_PATRON_ESTADO = re.compile(
    r"\b(calma|caos|urgente|hoy|ya|ahora|yesterday|mañana|nuevo|nueva|"
    r"gran|grande|importante|especial|exclusivo)\s*$",
    re.IGNORECASE
)

_TILDE_MAP = {
    "regulacion":"regulación","regulaciones":"regulaciones","innovacion":"innovación",
    "innovaciones":"innovaciones","tecnologia":"tecnología","tecnologias":"tecnologías",
    "tecnologica":"tecnológica","tecnologico":"tecnológico","educacion":"educación",
    "gestion":"gestión","administracion":"administración","informacion":"información",
    "comunicacion":"comunicación","comunicaciones":"comunicaciones","operacion":"operación",
    "operaciones":"operaciones","inversion":"inversión","inversiones":"inversiones",
    "expansion":"expansión","adquisicion":"adquisición","adquisiciones":"adquisiciones",
    "fusion":"fusión","fusiones":"fusiones","transicion":"transición",
    "transformacion":"transformación","digitalizacion":"digitalización",
    "automatizacion":"automatización","modernizacion":"modernización",
    "optimizacion":"optimización","implementacion":"implementación","evaluacion":"evaluación",
    "planificacion":"planificación","organizacion":"organización","atencion":"atención",
    "produccion":"producción","construccion":"construcción","distribucion":"distribución",
    "exportacion":"exportación","importacion":"importación","comercializacion":"comercialización",
    "negociacion":"negociación","negociaciones":"negociaciones","participacion":"participación",
    "colaboracion":"colaboración","asociacion":"asociación","integracion":"integración",
    "relacion":"relación","relaciones":"relaciones","situacion":"situación",
    "condicion":"condición","condiciones":"condiciones","solucion":"solución",
    "soluciones":"soluciones","prevencion":"prevención","proteccion":"protección",
    "fiscalizacion":"fiscalización","sancion":"sanción","sanciones":"sanciones",
    "investigacion":"investigación","investigaciones":"investigaciones","accion":"acción",
    "acciones":"acciones","direccion":"dirección","decision":"decisión",
    "decisiones":"decisiones","eleccion":"elección","elecciones":"elecciones",
    "votacion":"votación","aprobacion":"aprobación","legislacion":"legislación",
    "reclamacion":"reclamación","reclamaciones":"reclamaciones","obligacion":"obligación",
    "obligaciones":"obligaciones","inflacion":"inflación","tributacion":"tributación",
    "financiera":"financiera","financiero":"financiero","economica":"económica",
    "economico":"económico","economia":"economía","credito":"crédito",
    "creditos":"créditos","prestamo":"préstamo","prestamos":"préstamos",
    "interes":"interés","comision":"comisión","comisiones":"comisiones",
    "politica":"política","politicas":"políticas","politico":"político",
    "publica":"pública","publico":"público","estrategia":"estrategia",
    "estrategica":"estratégica","estrategico":"estratégico","logistica":"logística",
    "analisis":"análisis","diagnostico":"diagnóstico","indice":"índice",
    "vehiculo":"vehículo","vehiculos":"vehículos","electrico":"eléctrico",
    "electrica":"eléctrica","energia":"energía","energetica":"energética",
    "petroleo":"petróleo","mineria":"minería","agricola":"agrícola",
    "biologica":"biológica","ecologica":"ecológica","inclusion":"inclusión",
    "exclusion":"exclusión","pension":"pensión","pensiones":"pensiones",
    "jubilacion":"jubilación","compensacion":"compensación","remuneracion":"remuneración",
    "contratacion":"contratación","capacitacion":"capacitación","formacion":"formación",
    "certificacion":"certificación","habilitacion":"habilitación","autorizacion":"autorización",
    "concesion":"concesión","licitacion":"licitación","migracion":"migración",
    "poblacion":"población","recaudacion":"recaudación","asignacion":"asignación",
    "corporacion":"corporación","fundacion":"fundación","institucion":"institución",
    "instituciones":"instituciones","region":"región","unico":"único","unica":"única",
    "ultimo":"último","ultima":"última","proximo":"próximo","basico":"básico",
    "basica":"básica","historico":"histórico","historica":"histórica",
    "medico":"médico","medica":"médica","farmaceutica":"farmacéutica",
    "clinica":"clínica","numero":"número","telefono":"teléfono","telefonia":"telefonía",
    "movil":"móvil","moviles":"móviles","codigo":"código","informatica":"informática",
    "electronica":"electrónica","robotica":"robótica","ciberseguridad":"ciberseguridad",
    "trafico":"tráfico","transito":"tránsito","aereo":"aéreo","maritimo":"marítimo",
    "turistica":"turística","turistico":"turístico","gastronomia":"gastronomía",
    "academica":"académica","academico":"académico","pedagogica":"pedagógica",
    "cientifica":"científica","cientifico":"científico","juridica":"jurídica",
    "juridico":"jurídico","constitucion":"constitución","resolucion":"resolución",
    "notificacion":"notificación","programacion":"programación","actualizacion":"actualización",
    "verificacion":"verificación","validacion":"validación","liquidacion":"liquidación",
    "facturacion":"facturación","evasion":"evasión","corrupcion":"corrupción",
    "deforestacion":"deforestación","contaminacion":"contaminación","conservacion":"conservación",
    "restauracion":"restauración","rehabilitacion":"rehabilitación","renovacion":"renovación",
    "ampliacion":"ampliación","inauguracion":"inauguración","celebracion":"celebración",
    "clasificacion":"clasificación","eliminacion":"eliminación","motivacion":"motivación",
    "satisfaccion":"satisfacción","reputacion":"reputación","disposicion":"disposición",
}

_ENIE_MAP = {
    "desempeno":"desempeño","desempenos":"desempeños","empeno":"empeño","empenos":"empeños",
    "ensenanza":"enseñanza","ensenanzas":"enseñanzas","diseno":"diseño","disenos":"diseños",
    "disenador":"diseñador","disenadora":"diseñadora","disenadores":"diseñadores",
    "nino":"niño","nina":"niña","ninos":"niños","ninas":"niñas","ninez":"niñez",
    "ano":"año","anos":"años","danio":"daño","danios":"daños","dano":"daño","danos":"daños",
    "danino":"dañino","danina":"dañina","montana":"montaña","montanas":"montañas",
    "espana":"España","espanol":"español","espanola":"española","espanoles":"españoles",
    "companero":"compañero","companera":"compañera","companeros":"compañeros","companeras":"compañeras",
    "compania":"compañía","companias":"compañías","acompanamiento":"acompañamiento",
    "banio":"baño","banios":"baños","bano":"baño","banos":"baños",
    "penon":"peñón","senor":"señor","senora":"señora",
    "senores":"señores","senoras":"señoras","senal":"señal","senales":"señales",
    "senalizacion":"señalización","pequeno":"pequeño","pequena":"pequeña",
    "pequenos":"pequeños","pequenas":"pequeñas","sueno":"sueño","suenos":"sueños",
    "dueno":"dueño","duena":"dueña","duenos":"dueños","duenas":"dueñas",
    "otono":"otoño","punio":"puño","punios":"puños","puno":"puño",
    "canones":"cañones","manana":"mañana","mananas":"mañanas",
    "cabana":"cabaña","cabanas":"cabañas","banera":"bañera","vinedo":"viñedo",
    "vinedos":"viñedos","rebano":"rebaño","rebanos":"rebaños","extrano":"extraño",
    "extrana":"extraña","extranos":"extraños","extranas":"extrañas",
    "enganio":"engaño","engano":"engaño","enganos":"engaños","tamanio":"tamaño",
    "tamano":"tamaño","tamanos":"tamaños","muneca":"muñeca","munecas":"muñecas",
    "cunado":"cuñado","cunada":"cuñada","cunados":"cuñados","albanil":"albañil",
    "albaniles":"albañiles","narino":"Nariño","quindio":"Quindío",
    "ibanez":"Ibáñez","nunez":"Núñez","munoz":"Muñoz","ordonez":"Ordóñez",
    "yanez":"Yáñez","castaneda":"Castañeda","penalosa":"Peñalosa",
    "vineta":"viñeta","vinetas":"viñetas","banado":"bañado","banada":"bañada",
    "rinon":"riñón","rinones":"riñones","panial":"pañal","paniales":"pañales",
    "panal":"pañal","panales":"pañales","arana":"araña","aranas":"arañas",
    "pestana":"pestaña","pestanas":"pestañas","guino":"guiño","guinos":"guiños",
    "munequera":"muñequera","lenador":"leñador","lenadores":"leñadores",
    "resena":"reseña","resenas":"reseñas","panuelo":"pañuelo","panuelos":"pañuelos",
    "companerismo":"compañerismo","desengano":"desengaño","lenio":"leño","leno":"leño",
}

def corregir_tildes(texto: str) -> str:
    if not texto: return texto
    palabras = texto.split()
    resultado = []
    for p in palabras:
        low = p.lower()
        if low in _TILDE_MAP:
            c = _TILDE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        elif low in _ENIE_MAP:
            c = _ENIE_MAP[low]
            if p[0].isupper() and not c[0].isupper(): c = c[0].upper() + c[1:]
            resultado.append(c)
        else:
            resultado.append(p)
    return " ".join(resultado)


# ======================================
# CSS
# ======================================
def load_custom_css():
    st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Google+Sans:wght@400;500;700&family=Google+Sans+Text:wght@400;500;700&family=Roboto+Mono:wght@400;500&display=swap');
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
:root {
    --bg:#f8f9fa;--s1:#ffffff;--s2:#f1f3f4;--s3:#e8eaed;
    --border:#dadce0;--border2:#bdc1c6;--border-focus:#f97316;
    --text:#202124;--text2:#3c4043;--text3:#5f6368;--text4:#9aa0a6;
    --accent:#f97316;--accent2:#ea580c;--accent3:#c2410c;
    --accent-bg:#fff7ed;--accent-bg2:#ffedd5;--accent-bdr:#fed7aa;
    --green:#059669;--green2:#047857;--green-bg:#ecfdf5;--green-bdr:#a7f3d0;
    --red:#dc2626;--amber:#d97706;--blue:#1a73e8;
    --r:8px;--r2:12px;--r3:16px;--r4:20px;
    --shadow-sm:0 1px 2px rgba(60,64,67,0.1),0 1px 3px rgba(60,64,67,0.08);
    --shadow-md:0 1px 3px rgba(60,64,67,0.12),0 4px 8px rgba(60,64,67,0.08);
    --shadow-lg:0 2px 6px rgba(60,64,67,0.1),0 8px 24px rgba(60,64,67,0.1);
    --transition:all 0.2s cubic-bezier(0.4,0,0.2,1);
}
html,body,[data-testid="stApp"]{
    background:var(--bg)!important;color:var(--text)!important;
    font-family:'Google Sans Text','Inter',-apple-system,BlinkMacSystemFont,'Segoe UI',Roboto,sans-serif;
    font-size:14px;-webkit-font-smoothing:antialiased;letter-spacing:0.01em;
}
#MainMenu,footer,header{visibility:hidden}.stDeployButton{display:none}
.block-container{padding-top:1rem!important;padding-bottom:0!important}
[data-testid="stAppViewBlockContainer"]{padding-top:1rem!important}
.app-header{background:var(--s1);border:1px solid var(--border);border-radius:var(--r3);padding:1rem 1.5rem;margin-bottom:1rem;display:flex;align-items:center;gap:1rem;box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.app-header-icon{width:40px;height:40px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.2rem;color:white;flex-shrink:0;box-shadow:0 2px 8px rgba(249,115,22,0.3);}
.app-header-text{flex:1}
.app-header-title{font-family:'Google Sans',sans-serif;font-size:1.25rem;font-weight:700;color:var(--text);letter-spacing:-0.01em;line-height:1.3}
.app-header-version{font-family:'Roboto Mono',monospace;font-size:0.65rem;color:var(--text3);letter-spacing:0.03em;margin-top:0.15rem}
.app-header-badge{background:var(--accent-bg);border:1px solid var(--accent-bdr);color:var(--accent2);font-family:'Roboto Mono',monospace;font-size:0.6rem;font-weight:500;padding:0.25rem 0.75rem;border-radius:100px;letter-spacing:0.04em;text-transform:uppercase;white-space:nowrap;}
[data-testid="stTabs"] [data-testid="stTabsList"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;padding:4px!important;gap:4px!important;box-shadow:var(--shadow-sm)!important;margin-bottom:0.75rem!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]{font-family:'Google Sans',sans-serif!important;font-size:0.88rem!important;font-weight:500!important;color:var(--text2)!important;border-radius:var(--r)!important;padding:0.45rem 1.2rem!important;border:none!important;background:transparent!important;transition:var(--transition)!important;}
[data-testid="stTabs"] button[data-baseweb="tab"]:hover{background:var(--s2)!important;color:var(--text)!important}
[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"]{background:var(--accent-bg)!important;color:var(--accent2)!important;border:1px solid var(--accent-bdr)!important;font-weight:700!important;}
.metrics-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:0.6rem;margin:0.8rem 0}
.metric-card{background:var(--s1);border:1px solid var(--border);border-radius:var(--r2);padding:0.8rem 0.6rem;text-align:center;transition:var(--transition);box-shadow:var(--shadow-sm);position:relative;overflow:hidden;}
.metric-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:var(--r2) var(--r2) 0 0}
.metric-card.m-total::before{background:linear-gradient(90deg,#5f6368,#9aa0a6)}
.metric-card.m-unique::before{background:linear-gradient(90deg,#059669,#34d399)}
.metric-card.m-dup::before{background:linear-gradient(90deg,#f59e0b,#fbbf24)}
.metric-card.m-time::before{background:linear-gradient(90deg,#1a73e8,#4285f4)}
.metric-card.m-cost::before{background:linear-gradient(90deg,#f97316,#fb923c)}
.metric-card:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.metric-val{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;line-height:1;margin-bottom:0.3rem;letter-spacing:-0.01em}
.metric-lbl{font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);text-transform:uppercase;letter-spacing:0.08em;font-weight:500}
[data-testid="stForm"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r3)!important;padding:1.2rem 1.5rem!important;box-shadow:var(--shadow-md)!important;}
.sec-label{font-family:'Google Sans',sans-serif;font-size:0.72rem;font-weight:700;color:var(--text2);letter-spacing:0.08em;text-transform:uppercase;padding-bottom:0.3rem;border-bottom:2px solid var(--s3);margin:0.8rem 0 0.5rem;display:flex;align-items:center;gap:0.5rem;}
.sec-label::before{content:'';display:inline-block;width:3px;height:12px;background:linear-gradient(180deg,#f97316,#ea580c);border-radius:2px}
.upload-zone{display:grid;grid-template-columns:repeat(3,1fr);gap:0.6rem;margin:0.3rem 0}
.upload-zone-card{background:var(--s1);border:1.5px dashed var(--border);border-radius:var(--r2);padding:0.6rem 0.8rem;display:flex;align-items:center;gap:0.6rem;transition:var(--transition);}
.upload-zone-card:hover{border-color:var(--accent);border-style:solid;transform:translateY(-1px);box-shadow:var(--shadow-md)}
.upload-zone-icon{width:32px;height:32px;border-radius:8px;display:flex;align-items:center;justify-content:center;font-size:1rem;flex-shrink:0;}
.upload-zone-icon.uz-dossier{background:#fff7ed;color:#f97316}
.upload-zone-icon.uz-region{background:#ecfdf5;color:#059669}
.upload-zone-icon.uz-internet{background:#eff6ff;color:#1a73e8}
.upload-zone-text{flex:1;min-width:0}
.upload-zone-title{font-family:'Google Sans',sans-serif;font-size:0.82rem;font-weight:700;color:var(--text);line-height:1.2}
.upload-zone-desc{font-size:0.7rem;color:var(--text3);line-height:1.3}
[data-testid="stFileUploader"]{background:var(--s1)!important;border:1.5px dashed var(--border)!important;border-radius:var(--r)!important;padding:0.4rem 0.6rem!important;transition:var(--transition)!important;min-height:auto!important;}
[data-testid="stFileUploader"]:hover{border-color:var(--accent)!important;border-style:solid!important;background:var(--accent-bg)!important;}
[data-testid="stFileUploader"] section{padding:0.2rem!important}
[data-testid="stFileUploader"] section>div{font-size:0.78rem!important;color:var(--text2)!important}
[data-testid="stFileUploader"] section small{font-size:0.7rem!important;color:var(--text3)!important}
[data-testid="stFileUploader"] button{background:var(--accent-bg)!important;border:1px solid var(--accent-bdr)!important;color:var(--accent2)!important;font-weight:500!important;font-size:0.75rem!important;border-radius:100px!important;padding:0.25rem 0.8rem!important;font-family:'Google Sans',sans-serif!important;transition:var(--transition)!important;}
[data-testid="stFileUploader"] button:hover{background:var(--accent)!important;color:white!important;border-color:var(--accent)!important}
[data-testid="stTextInput"] input,[data-testid="stTextArea"] textarea{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:var(--r)!important;font-family:'Google Sans Text',sans-serif!important;font-size:0.9rem!important;padding:0.5rem 0.75rem!important;transition:var(--transition)!important;}
[data-testid="stTextInput"] input:focus,[data-testid="stTextArea"] textarea:focus{border-color:var(--accent)!important;box-shadow:0 0 0 3px rgba(249,115,22,0.12)!important;}
[data-testid="stTextInput"] input::placeholder,[data-testid="stTextArea"] textarea::placeholder{color:var(--text4)!important;font-size:0.85rem!important;}
label[data-testid="stWidgetLabel"] p{font-family:'Google Sans',sans-serif!important;color:var(--text2)!important;font-size:0.82rem!important;font-weight:500!important;margin-bottom:0.15rem!important;}
.stButton>button,[data-testid="stDownloadButton"]>button{background:var(--s1)!important;border:1.5px solid var(--border)!important;color:var(--text)!important;border-radius:100px!important;font-family:'Google Sans',sans-serif!important;font-weight:500!important;font-size:0.88rem!important;transition:var(--transition)!important;padding:0.5rem 1.2rem!important;box-shadow:none!important;}
.stButton>button:hover,[data-testid="stDownloadButton"]>button:hover{border-color:var(--accent)!important;color:var(--accent2)!important;background:var(--accent-bg)!important;box-shadow:var(--shadow-sm)!important;transform:translateY(-1px)!important;}
.stButton>button[kind="primary"],[data-testid="stDownloadButton"]>button[kind="primary"]{background:var(--accent)!important;border:none!important;color:#fff!important;font-weight:500!important;font-size:0.92rem!important;padding:0.6rem 1.5rem!important;box-shadow:0 1px 3px rgba(249,115,22,0.3),0 4px 12px rgba(249,115,22,0.15)!important;letter-spacing:0.01em!important;}
.stButton>button[kind="primary"]:hover,[data-testid="stDownloadButton"]>button[kind="primary"]:hover{background:var(--accent2)!important;box-shadow:0 2px 6px rgba(234,88,12,0.35),0 8px 24px rgba(234,88,12,0.18)!important;transform:translateY(-1px)!important;color:#fff!important;}
[data-testid="stRadio"] label{font-family:'Google Sans Text',sans-serif!important;color:var(--text)!important;font-size:0.88rem!important;font-weight:400!important;}
[data-testid="stRadio"]{margin-bottom:0!important}
[data-testid="stRadio"]>div{gap:0!important}
[data-testid="stStatus"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;font-family:'Roboto Mono',monospace!important;font-size:0.8rem!important;}
[data-testid="stAlert"]{background:var(--s1)!important;border:1px solid var(--border)!important;border-radius:var(--r2)!important;color:var(--text2)!important;font-size:0.85rem!important;padding:0.6rem 0.8rem!important;}
.success-banner{background:linear-gradient(135deg,#ecfdf5,#d1fae5);border:1px solid var(--green-bdr);border-left:4px solid var(--green);border-radius:var(--r2);padding:0.8rem 1.2rem;margin:0.5rem 0 0.8rem;display:flex;align-items:center;gap:0.8rem;}
.success-icon{width:34px;height:34px;background:linear-gradient(135deg,#059669,#047857);border-radius:50%;display:flex;align-items:center;justify-content:center;color:white;font-size:1rem;flex-shrink:0;}
.success-title{font-family:'Google Sans',sans-serif;font-size:1rem;font-weight:700;color:#047857;margin-bottom:0.1rem}
.success-sub{font-size:0.8rem;color:var(--text2)}
.auth-wrap{max-width:380px;margin:8vh auto 0;text-align:center}
.auth-icon{width:60px;height:60px;background:linear-gradient(135deg,#f97316,#ea580c);border-radius:16px;display:inline-flex;align-items:center;justify-content:center;font-size:1.6rem;color:white;margin-bottom:1rem;box-shadow:0 4px 16px rgba(249,115,22,0.3);}
.auth-title{font-family:'Google Sans',sans-serif;font-size:1.5rem;font-weight:700;color:var(--text);margin-bottom:0.3rem}
.auth-sub{font-size:0.85rem;color:var(--text3);margin-bottom:2rem}
.cluster-info{background:var(--accent-bg);border:1px solid var(--accent-bdr);border-radius:var(--r);padding:0.5rem 0.8rem;margin:0.4rem 0;font-family:'Roboto Mono',monospace;font-size:0.68rem;color:var(--text2);line-height:1.6;}
.cluster-info b{color:var(--accent2);font-size:0.72rem}
.config-badge{display:inline-flex;align-items:center;gap:0.4rem;background:var(--s2);border:1px solid var(--border);border-radius:100px;padding:0.2rem 0.7rem;font-family:'Roboto Mono',monospace;font-size:0.62rem;color:var(--text3);margin-bottom:0.6rem;}
[data-testid="stProgressBar"]>div>div{background:linear-gradient(90deg,#f97316,#fb923c,#fdba74)!important;border-radius:100px!important;height:5px!important;}
[data-testid="stDataFrame"]{border:1px solid var(--border)!important;border-radius:var(--r2)!important;box-shadow:var(--shadow-sm)!important;overflow:hidden!important;}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:var(--s2);border-radius:3px}
::-webkit-scrollbar-thumb{background:var(--border2);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--accent)}
.footer{font-family:'Roboto Mono',monospace;font-size:0.6rem;color:var(--text4);text-align:center;padding:0.8rem 0 0.5rem;letter-spacing:0.04em;border-top:1px solid var(--s3);margin-top:1rem;}
.stElementContainer{margin-bottom:0!important}
[data-testid="stVerticalBlock"]>div{gap:0.3rem!important}
[data-testid="stHorizontalBlock"]>div{gap:0.4rem!important}
hr{border-color:var(--s3)!important;margin:0.5rem 0!important}
[data-testid="stSelectbox"]>div>div{font-family:'Google Sans Text',sans-serif!important;font-size:0.88rem!important;color:var(--text)!important;}
@media(max-width:768px){
    .metrics-grid{grid-template-columns:repeat(2,1fr)}
    .upload-zone{grid-template-columns:1fr}
    .app-header{flex-direction:column;text-align:center;gap:0.5rem;padding:1rem}
}
</style>
""", unsafe_allow_html=True)


# ======================================
# Umbrales adaptativos según tamaño del corpus
# ======================================
def _umbrales_adaptativos(n: int) -> dict:
    if n <= 5:
        return dict(
            subtema=0.93,
            tema=0.85,
            dedup_label=0.90,
            fusion_subtemas=0.92,
            fusion_intergrupo=0.95,
            min_pertenencia_subtema=0.80,
            min_pertenencia_tema=0.75,
            coherencia_etiqueta=0.50,
            sim_minima_agrupacion=0.93,
            sim_minima_keywords=0.93,
            max_iter_fusion=1,
            num_temas_max=n,
            usar_paso2b=False,
            usar_fusion_iterativa=False,
        )
    elif n <= 10:
        return dict(
            subtema=0.90,
            tema=0.84,
            dedup_label=0.88,
            fusion_subtemas=0.90,
            fusion_intergrupo=0.93,
            min_pertenencia_subtema=0.72,
            min_pertenencia_tema=0.65,
            coherencia_etiqueta=0.42,
            sim_minima_agrupacion=0.90,
            sim_minima_keywords=0.90,
            max_iter_fusion=2,
            num_temas_max=min(n, 5),
            usar_paso2b=False,
            usar_fusion_iterativa=False,
        )
    elif n <= 20:
        return dict(
            subtema=0.87,
            tema=0.82,
            dedup_label=0.86,
            fusion_subtemas=0.88,
            fusion_intergrupo=0.91,
            min_pertenencia_subtema=0.66,
            min_pertenencia_tema=0.58,
            coherencia_etiqueta=0.38,
            sim_minima_agrupacion=0.87,
            sim_minima_keywords=0.87,
            max_iter_fusion=3,
            num_temas_max=min(n // 2, NUM_TEMAS_MAX),
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )
    else:
        return dict(
            subtema=UMBRAL_SUBTEMA,
            tema=UMBRAL_TEMA,
            dedup_label=UMBRAL_DEDUP_LABEL,
            fusion_subtemas=UMBRAL_FUSION_SUBTEMAS,
            fusion_intergrupo=UMBRAL_FUSION_INTERGRUPO,
            min_pertenencia_subtema=UMBRAL_MIN_PERTENENCIA_SUBTEMA,
            min_pertenencia_tema=UMBRAL_MIN_PERTENENCIA_TEMA,
            coherencia_etiqueta=UMBRAL_COHERENCIA_ETIQUETA,
            sim_minima_agrupacion=SIM_MINIMA_AGRUPACION_SUBTEMA,
            sim_minima_keywords=SIM_MINIMA_KEYWORDS_RARAS,
            max_iter_fusion=MAX_ITER_FUSION,
            num_temas_max=NUM_TEMAS_MAX,
            usar_paso2b=True,
            usar_fusion_iterativa=True,
        )


# ======================================
# Caché Global de Embeddings
# ======================================
class EmbeddingCache:
    def __init__(self):
        self._cache: Dict[str, List[float]] = {}
        self._hits = 0
        self._misses = 0

    def _key(self, text):
        return hashlib.md5(text[:2000].encode('utf-8', errors='ignore')).hexdigest()

    def get(self, text):
        k = self._key(text)
        if k in self._cache:
            self._hits += 1
            return self._cache[k]
        self._misses += 1
        return None

    def put(self, text, emb):
        self._cache[self._key(text)] = emb

    def get_many(self, textos):
        results = [None] * len(textos)
        missing = []
        for i, t in enumerate(textos):
            c = self.get(t)
            if c is not None:
                results[i] = c
            else:
                missing.append(i)
        return results, missing

    def stats(self):
        total = self._hits + self._misses
        rate = (self._hits / total * 100) if total > 0 else 0
        return f"Cache: {self._hits} hits, {self._misses} misses ({rate:.0f}%)"

    def clear(self):
        self._cache.clear()
        self._hits = 0
        self._misses = 0

if '_emb_cache' not in st.session_state:
    st.session_state['_emb_cache'] = EmbeddingCache()

def get_embedding_cache():
    return st.session_state['_emb_cache']

# ======================================
# Configuración vía Google Sheets (CSV público)
# ======================================
CONFIG_CACHE_TTL = 300  # segundos

@st.cache_data(ttl=CONFIG_CACHE_TTL, show_spinner=False)
def _fetch_map_from_csv(csv_url: str) -> dict:
    df = pd.read_csv(csv_url, header=None, dtype=str)
    df = df.dropna(how="all")
    mapping = pd.Series(
        df.iloc[:, 1].values,
        index=df.iloc[:, 0].astype(str).str.lower().str.strip()
    ).to_dict()
    mapping = {k: v for k, v in mapping.items() if k not in ("nan", "")}
    return mapping

def load_config_from_sheets():
    regiones_url = st.secrets.get("REGIONES_CSV_URL")
    internet_url = st.secrets.get("INTERNET_CSV_URL")

    if not regiones_url or not internet_url:
        st.error(
            "❌ Faltan las URLs de configuración. Agrega REGIONES_CSV_URL e "
            "INTERNET_CSV_URL en los Secrets de la app."
        )
        st.stop()

    try:
        region_map = _fetch_map_from_csv(regiones_url)
        internet_map = _fetch_map_from_csv(internet_url)
    except Exception as e:
        st.error(f"❌ No se pudo leer la configuración desde Google Sheets: {e}")
        st.stop()

    return region_map, internet_map

def refresh_config_cache():
    _fetch_map_from_csv.clear()


# ======================================
# Funciones Auxiliares de Limpieza, Enlaces y Conversión
# ======================================

def check_password():
    if st.session_state.get("password_correct", False):
        return True
    st.markdown("""
    <div class="auth-wrap">
        <div class="auth-icon">◈</div>
        <div class="auth-title">Sistema de Análisis</div>
        <div class="auth-sub">Ingresa tus credenciales para continuar</div>
    </div>""", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        with st.form("pw"):
            pw = st.text_input("Contraseña", type="password", placeholder="Ingresa tu contraseña")
            if st.form_submit_button("Ingresar", use_container_width=True, type="primary"):
                if pw == st.secrets.get("APP_PASSWORD", "INVALID"):
                    st.session_state["password_correct"] = True
                    st.rerun()
                else:
                    st.error("Contraseña incorrecta")
    return False

def call_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            time.sleep(d)
            d *= 2

async def acall_with_retries(fn, *a, **kw):
    d = 1
    for att in range(3):
        try:
            return await fn(*a, **kw)
        except Exception as e:
            if att == 2: raise e
            await asyncio.sleep(d)
            d *= 2

def norm_key(text):
    if text is None: return ""
    return re.sub(r"[^a-z0-9]+", "", unidecode(str(text).strip().lower()))

def capitalizar_etiqueta(tema):
    if not tema or not tema.strip(): return "Sin tema"
    tema = tema.strip().lower()
    tema = corregir_tildes(tema)
    return tema[0].upper() + tema[1:]

def _frase_esta_completa(texto):
    if not texto or not texto.strip(): return False
    palabras = texto.strip().split()
    if not palabras: return False
    ultima = palabras[-1].lower().rstrip(".,;:!?")
    return unidecode(ultima) not in _TRAILING_INCOMPLETE and len(ultima) > 1

def _recortar_frase_completa(texto, max_palabras=7):
    if not texto: return "Sin tema"
    palabras = texto.strip().split()
    if len(palabras) > max_palabras: palabras = palabras[:max_palabras]
    while palabras and unidecode(palabras[-1].lower().rstrip(".,;:!?")) in _TRAILING_INCOMPLETE:
        palabras.pop()
    if not palabras: return texto.strip().split()[0] if texto.strip() else "Sin tema"
    return " ".join(palabras)

def limpiar_tema(tema):
    if not tema: return "Sin tema"
    tema = tema.strip().strip('"\'')
    for px in ["subtema:", "tema:", "categoría:", "categoria:", "category:"]:
        if tema.lower().startswith(px): tema = tema[len(px):].strip()
    tema = _recortar_frase_completa(tema, max_palabras=7)
    return capitalizar_etiqueta(tema) if tema else "Sin tema"

def limpiar_tema_geografico(tema, marca, aliases):
    if not tema: return "Sin tema"
    tl = unidecode(tema.lower())
    for n in [marca] + [a for a in aliases if a]:
        patron = r'\b' + re.escape(unidecode(n.strip().lower())) + r'\b'
        tl = re.sub(patron, '', tl)
    frases_eliminar = [
        "en colombia", "de colombia", "del pais", "en el pais",
        "territorio nacional", "a nivel nacional", "en todo el pais",
    ]
    for frase in frases_eliminar:
        tl = re.sub(r'\b' + re.escape(frase) + r'\b', '', tl)
    tl = re.sub(r'\s+', ' ', tl).strip()
    if not tl: return "Sin tema"
    tokens_orig = tema.split()
    tokens_norm = unidecode(tema.lower()).split()
    norm_disponibles = tl.split()
    resultado_tokens = []
    for orig, norm in zip(tokens_orig, tokens_norm):
        if norm_disponibles and norm == norm_disponibles[0]:
            resultado_tokens.append(orig)
            norm_disponibles.pop(0)
    resultado = " ".join(resultado_tokens).strip()
    resultado = corregir_tildes(resultado) if resultado else ""
    return limpiar_tema(resultado) if resultado.strip() else "Sin tema"

def string_norm_label(s):
    if not s: return ""
    s = unidecode(s.lower())
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(t for t in s.split() if t not in STOPWORDS_ES)

_ACCIONES_OPUESTAS = [
    ({"aprobacion", "aprueba", "apoyo", "acuerdo", "aval", "respaldo"}, {"rechazo", "rechaza", "desacuerdo", "oposicion", "critica"}),
    ({"aumento", "crecimiento", "alza", "subida", "incremento"}, {"caida", "reduccion", "baja", "disminucion", "descenso"}),
    ({"apertura", "inauguracion", "inicio", "lanzamiento", "estreno"}, {"cierre", "suspension", "fin", "clausura", "cancelacion"}),
    ({"exito", "logro", "triunfo", "premio", "reconocimiento"}, {"fracaso", "derrota", "problema", "crisis", "sancion"}),
    ({"demanda", "denuncia", "investigacion", "sancion", "multa"}, {"absolucion", "archivo", "exoneracion", "acuerdo"}),
]

_TOKENS_DEBILES_AGRUPACION = STOPWORDS_ES | {
    "noticia", "noticias", "informe", "informacion", "comunicado", "anuncio",
    "colombia", "pais", "nacional", "regional", "local", "sector", "sectores",
    "empresa", "empresas", "entidad", "entidades", "autoridad", "autoridades",
    "gobierno", "alcaldia", "gobernacion", "ministerio", "nuevo", "nueva",
    "nuevos", "nuevas", "plan", "programa", "proyecto", "iniciativa",
    "actividad", "actividades", "gestion", "tema", "caso", "casos",
}

def _tokens_distintivos(texto: str, min_len: int = 4) -> set:
    norm = string_norm_label(texto)
    return {
        t for t in norm.split()
        if len(t) >= min_len and t not in _TOKENS_DEBILES_AGRUPACION and not t.isdigit()
    }

def _overlap_distintivo(a: str, b: str) -> float:
    ta, tb = _tokens_distintivos(a), _tokens_distintivos(b)
    if not ta or not tb: return 0.0
    return len(ta & tb) / max(1, min(len(ta), len(tb)))

def _hay_conflicto_accion(a: str, b: str) -> bool:
    ta, tb = _tokens_distintivos(a, min_len=3), _tokens_distintivos(b, min_len=3)
    for grupo_a, grupo_b in _ACCIONES_OPUESTAS:
        if (ta & grupo_a and tb & grupo_b) or (ta & grupo_b and tb & grupo_a):
            return True
    return False

def _etiquetas_compatibles(a: str, b: str, min_overlap: float = 0.45) -> bool:
    na, nb = string_norm_label(a), string_norm_label(b)
    if not na or not nb: return False
    if _hay_conflicto_accion(na, nb): return False
    if SequenceMatcher(None, na, nb).ratio() >= 0.90: return True
    return _overlap_distintivo(na, nb) >= min_overlap

def _grupos_contenido_compatibles(
    textos_a: list,
    textos_b: list,
    etiqueta_a: str = "",
    etiqueta_b: str = "",
    min_sim: float = 0.88,
    min_overlap: float = 0.20,
) -> bool:
    muestra_a = [str(t) for t in textos_a[:20] if str(t).strip()]
    muestra_b = [str(t) for t in textos_b[:20] if str(t).strip()]
    if not muestra_a or not muestra_b: return False
    texto_a = " ".join(muestra_a)[:2500]
    texto_b = " ".join(muestra_b)[:2500]
    if _hay_conflicto_accion(f"{etiqueta_a} {texto_a}", f"{etiqueta_b} {texto_b}"):
        return False
    overlap = _overlap_distintivo(f"{etiqueta_a} {texto_a}", f"{etiqueta_b} {texto_b}")
    labels_muy_cercanas = _etiquetas_compatibles(etiqueta_a, etiqueta_b, min_overlap=0.55)
    if overlap < min_overlap and not labels_muy_cercanas:
        return False
    embs = get_embeddings_batch([texto_a, texto_b])
    if len(embs) < 2 or embs[0] is None or embs[1] is None:
        return labels_muy_cercanas and overlap >= min_overlap
    sim = cosine_similarity(
        np.array(embs[0]).reshape(1, -1),
        np.array(embs[1]).reshape(1, -1)
    )[0][0]
    return sim >= min_sim

def _validar_estructura_subtema(etiqueta: str) -> bool:
    if not etiqueta or len(etiqueta.split()) < 2: return False
    if len(etiqueta.split()) > 7: return False
    if _PATRON_TITULAR.match(etiqueta): return False
    if _PATRON_ESTADO.search(etiqueta): return False
    palabras = etiqueta.split()
    if len(palabras) <= 4:
        nexos = {
            "de","del","para","sobre","en","con","por","ante","hacia",
            "entre","sin","al","las","los","una","uno","que","como",
            "y","o","a","e","u",
        }
        tiene_nexo = any(unidecode(p.lower().rstrip(".,;:!?")) in nexos for p in palabras[1:])
        if not tiene_nexo: return False
    return True

def extract_link(cell):
    if hasattr(cell, "hyperlink") and cell.hyperlink:
        return {"value": "Link", "url": cell.hyperlink.target}
    if isinstance(cell.value, str) and "=HYPERLINK" in cell.value:
        m = re.search(r'=HYPERLINK\("([^"]+)"', cell.value)
        if m: return {"value": "Link", "url": m.group(1)}
    return {"value": cell.value, "url": None}

def extract_link_from_cell(cell):
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target
    return None

def convert_html_entities(text):
    if not isinstance(text, str):
        return text
    text = html.unescape(text)
    html_entities = {
        '&#xF3;': 'ó', '&#xE1;': 'á', '&#xE9;': 'é', '&#xED;': 'í',
        '&#xFA;': 'ú', '&#xF1;': 'ñ', '&#xDC;': 'Ü', '&#xFC;': 'ü',
        '&#xC1;': 'Á', '&#xC9;': 'É', '&#xCD;': 'Í', '&#xD3;': 'Ó',
        '&#xDA;': 'Ú', '&#xD1;': 'Ñ', '&#xC7;': 'Ç', '&#xE7;': 'ç',
    }
    for entity, char in html_entities.items():
        text = text.replace(entity, char)

    def replace_hex_entity(match):
        try:
            return chr(int(match.group(1), 16))
        except Exception:
            return match.group(0)

    def replace_decimal_entity(match):
        try:
            return chr(int(match.group(1)))
        except Exception:
            return match.group(0)

    text = re.sub(r'&#x([0-9A-Fa-f]+);', replace_hex_entity, text)
    text = re.sub(r'&#(\d+);', replace_decimal_entity, text)

    for bad, good in {'\u201c': '"', '\u201d': '"', '\u2018': "'", '\u2019': "'",
                      'Â': '', 'â': '', '€': '', '™': ''}.items():
        text = text.replace(bad, good)
    return text

def clean_text(text):
    if not isinstance(text, str):
        return text
    return convert_html_entities(text).strip()

def clean_cuerpo(text):
    if not isinstance(text, str) or text.strip() == '':
        return text
    text = convert_html_entities(text)
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()


# ======================================
# FUNCIÓN DE NORMALIZACIÓN DE TÍTULOS (MEJORADA)
# ======================================
def normalize_title_for_comparison(title):
    if not isinstance(title, str): 
        return ""
    
    cleaned = re.sub(r"\s+[\|–—-]\s+[^\|–—-]+$", "", title).strip()
    
    if ":" in cleaned:
        parts = cleaned.split(":", 1)
        suffix = parts[1].strip()
        if len(suffix) >= 10:
            cleaned = suffix
            
    return re.sub(r"\W+", " ", cleaned).lower().strip()


def clean_title_for_output(title):
    return re.sub(r"\s*\|\s*[\w\s]+$", "", str(title)).strip()

def corregir_texto(text):
    if not isinstance(text, str): return text
    text = re.sub(r"(<br>|\[\.\.\.\]|\s+)", " ", text).strip()
    m = re.search(r"[A-ZÁÉÍÓÚÑ]", text)
    if m: text = text[m.start():]
    if text and not text.endswith("..."): text = text.rstrip(".") + "..."
    return text

def normalizar_tipo_medio(tipo_raw):
    if not isinstance(tipo_raw, str): return str(tipo_raw)
    t = unidecode(tipo_raw.strip().lower())
    return {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio', 'radio': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión', 'tv': 'Televisión',
        'television': 'Televisión', 'televisión': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }.get(t, str(tipo_raw).strip().title() or "Otro")

def parse_numeric(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    s = str(val).strip()
    if not s:
        return None
    if 'e' in s.lower():
        s = s.replace(',', '.')
    else:
        if ',' in s and '.' in s:
            if s.rfind('.') < s.rfind(','):
                s = s.replace('.', '').replace(',', '.')
            else:
                s = s.replace(',', '')
        elif ',' in s:
            parts = s.split(',')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0,')):
                s = s.replace(',', '')
            else:
                s = s.replace(',', '.')
        elif '.' in s:
            parts = s.split('.')
            if len(parts) > 2 or (len(parts) == 2 and len(parts[1]) == 3 and not s.lower().startswith('0.')):
                s = s.replace('.', '')
    try:
        f_val = float(s)
        if f_val.is_integer():
            return int(f_val)
        return f_val
    except ValueError:
        return None

def texto_para_embedding(titulo, resumen, max_len=1800):
    t = str(titulo or "").strip()
    r = str(resumen or "").strip()
    return f"{t}. {t}. {t}. {r}"[:max_len]

def _normalizar_mencion(texto: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", unidecode(str(texto).lower()))).strip()

def _coincide_nombre_completo(texto: str, nombre: str) -> bool:
    nombre = _normalizar_mencion(nombre)
    if len(nombre) < 3:
        return False
    return bool(re.search(rf"(?<![a-z0-9]){re.escape(nombre)}(?![a-z0-9])", texto))

def _validar_etiqueta_completa(etiqueta, titulos_grp=None, resumenes_grp=None, marca="", aliases=None, fallback_fn=None):
    if not etiqueta or etiqueta.strip().lower() in ("sin tema", "varios", "n/a"):
        if fallback_fn: return fallback_fn(titulos_grp or [])
        return "Cobertura informativa general"
    if _frase_esta_completa(etiqueta): return etiqueta
    recortada = _recortar_frase_completa(etiqueta, max_palabras=7)
    if _frase_esta_completa(recortada) and len(recortada.split()) >= 2:
        return capitalizar_etiqueta(recortada)
    if titulos_grp and len(titulos_grp) > 0:
        try:
            prompt = (
                f"La frase '{etiqueta}' está incompleta o es genérica. "
                f"Genera una frase temática COMPLETA en español de 4-6 palabras "
                f"con preposición (de/del/para/sobre/en):\n\n"
                + "\n".join(f"  · {t[:120]}" for t in titulos_grp[:4])
                + "\n\nREGLAS: frase nominal con preposición, terminar en sustantivo/adjetivo, "
                "tildes y ñ correctas, sin marcas ni ciudades.\n"
                "CORRECTO: 'Proyecto de terminal de transportes', 'Operación del Canal del Dique'\n"
                "INCORRECTO: 'Terminal transportes', 'Operación canal'\n"
                'JSON: {"subtema":"..."}'
            )
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=80,
                temperature=0.1,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
            raw = json.loads(resp.choices[0].message.content).get("subtema", "")
            if raw:
                cleaned = limpiar_tema_geografico(limpiar_tema(raw), marca, aliases or [])
                if _frase_esta_completa(cleaned) and len(cleaned.split()) >= 2:
                    return capitalizar_etiqueta(cleaned)
        except:
            pass
    if fallback_fn: return fallback_fn(titulos_grp or [])
    return capitalizar_etiqueta(recortada) if recortada and len(recortada.split()) >= 2 else "Cobertura informativa general"

def dedup_labels(etiquetas, umbral=UMBRAL_DEDUP_LABEL):
    unique = list(dict.fromkeys(etiquetas))
    if len(unique) <= 1:
        return etiquetas
    normed = [string_norm_label(u) for u in unique]
    n = len(unique)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    def _es_fusion_segura(s1, s2):
        return _etiquetas_compatibles(s1, s2, min_overlap=0.45)

    for i in range(n):
        if not normed[i]: continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j): continue
            if SequenceMatcher(None, normed[i], normed[j]).ratio() >= max(umbral, 0.88):
                if _es_fusion_segura(normed[i], normed[j]):
                    union(i, j)
                    
    for i in range(n):
        if not normed[i]: continue
        tokens_i = set(normed[i].split())
        if len(tokens_i) < 2: continue
        for j in range(i + 1, n):
            if not normed[j] or find(i) == find(j): continue
            tokens_j = set(normed[j].split())
            if len(tokens_j) < 2: continue
            interseccion = tokens_i & tokens_j
            menor = min(len(tokens_i), len(tokens_j))
            if menor > 0 and len(interseccion) / menor >= 0.78:
                if _es_fusion_segura(normed[i], normed[j]):
                    union(i, j)
                    
    le = get_embeddings_batch(unique)
    vp = [(i, le[i]) for i in range(n) if le[i] is not None]
    if len(vp) >= 2:
        vi, vv = zip(*vp)
        sm = cosine_similarity(np.array(vv))
        for pi in range(len(vi)):
            for pj in range(pi + 1, len(vi)):
                if sm[pi][pj] >= max(umbral, 0.90):
                    if find(vi[pi]) != find(vi[pj]):
                        if _es_fusion_segura(normed[vi[pi]], normed[vi[pj]]):
                            union(vi[pi], vi[pj])

    freq = Counter(etiquetas)
    grupos = defaultdict(list)
    for i in range(n):
        grupos[find(i)].append(i)
    canon = {}
    for root, members in grupos.items():
        cands = [unique[m] for m in members]
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        va = [c for c in cands if c.lower() not in ("sin tema", "varios")]
        if vc:
            canon[root] = max(vc, key=lambda c: (freq[c], len(c)))
        elif va:
            best = max(va, key=lambda c: (freq[c], len(c)))
            r = _recortar_frase_completa(best)
            canon[root] = r if _frase_esta_completa(r) else best
        else:
            canon[root] = cands[0]
    lm = {unique[i]: canon[find(i)] for i in range(n)}
    return [capitalizar_etiqueta(lm.get(e, e)) for e in etiquetas]

def _fusionar_subtemas_semanticos(subtemas, textos_por_subtema, marca, aliases, umbral=UMBRAL_FUSION_SUBTEMAS):
    unique_subs = list(dict.fromkeys(subtemas))
    if len(unique_subs) <= 1: return subtemas
    repr_texts = []
    for sub in unique_subs:
        txts = textos_por_subtema.get(sub, [])
        palabras = []
        for t in txts[:20]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3: palabras.append(w)
        top_kw = " ".join(w for w, _ in Counter(palabras).most_common(10))
        repr_texts.append(f"{sub}. {sub}. {sub}. {top_kw}"[:600])
    emb_repr = get_embeddings_batch(repr_texts)
    valid = [(i, emb_repr[i]) for i in range(len(unique_subs)) if emb_repr[i] is not None]
    if len(valid) < 2: return subtemas
    v_idx, v_emb = zip(*valid)
    sim = cosine_similarity(np.array(v_emb))
    n = len(v_idx)
    parent = list(range(n))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[rb] = ra

    for i in range(n):
        for j in range(i + 1, n):
            if find(i) == find(j): continue
            sub_i, sub_j = unique_subs[v_idx[i]], unique_subs[v_idx[j]]
            if sim[i][j] >= max(umbral, 0.88) and _grupos_contenido_compatibles(
                textos_por_subtema.get(sub_i, []),
                textos_por_subtema.get(sub_j, []),
                sub_i,
                sub_j,
                min_sim=max(umbral, 0.88),
                min_overlap=0.22,
            ):
                union(i, j)
            
    grupos = defaultdict(list)
    for i in range(n): grupos[find(i)].append(v_idx[i])
    freq = Counter(subtemas)
    lm = {}
    for root, members in grupos.items():
        cands = [unique_subs[m] for m in members]
        if len(cands) == 1:
            lm[cands[0]] = cands[0]
            continue
        vc = [c for c in cands if c.lower() not in ("sin tema", "varios") and _frase_esta_completa(c)]
        best = max(vc, key=lambda c: (freq.get(c, 0), len(c))) if vc else max(cands, key=lambda c: (freq.get(c, 0), len(c)))
        if len(cands) <= 3:
            unified = _unificar_subtemas_llm(cands, textos_por_subtema, marca, aliases)
            if unified and _frase_esta_completa(unified): best = unified
        for c in cands: lm[c] = capitalizar_etiqueta(best)
    return [lm.get(s, s) for s in subtemas]

def _unificar_subtemas_llm(subtemas_a_unificar, textos_por_subtema, marca, aliases):
    subs_str = "\n".join(f"  · {s}" for s in subtemas_a_unificar)
    all_kw = []
    for sub in subtemas_a_unificar:
        for t in textos_por_subtema.get(sub, [])[:5]:
            for w in string_norm_label(str(t)).split():
                if len(w) > 3: all_kw.append(w)
    kw_str = " · ".join(w for w, _ in Counter(all_kw).most_common(8))
    prompt = (
        f"Estos subtemas son variaciones del MISMO tema. "
        f"Genera UN subtema unificado (4-6 palabras) como frase nominal completa:\n\n"
        f"{subs_str}\n\nKeywords: {kw_str}\n\n"
        "REGLAS: frase coherente con preposición (de/del/para/sobre/en), "
        "sin marcas ni ciudades, tildes y ñ correctas, terminar en sustantivo/adjetivo.\n"
        "CORRECTO: 'Regulación de tarifas eléctricas', 'Apertura de nuevas sucursales'\n"
        "INCORRECTO: 'Tarifas energía', 'Apertura sucursales', 'Actividad corporativa'\n"
        'JSON: {"subtema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=60,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
        if u:
            st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
            st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
        raw = json.loads(resp.choices[0].message.content).get("subtema", "")
        if raw: return limpiar_tema_geografico(limpiar_tema(raw), marca, aliases)
    except:
        pass
    return None

def get_embeddings_batch(textos, batch_size=100):
    if not textos: return []
    cache = get_embedding_cache()
    resultados, missing = cache.get_many(textos)
    if not missing: return resultados
    mt = [textos[i][:2000] if textos[i] else "" for i in missing]
    for i in range(0, len(mt), batch_size):
        batch = mt[i:i + batch_size]
        bidx = missing[i:i + batch_size]
        try:
            resp = call_with_retries(openai.Embedding.create, input=batch, model=OPENAI_MODEL_EMBEDDING)
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_embedding'] += (u.get('total_tokens') if isinstance(u, dict) else getattr(u, 'total_tokens', 0)) or 0
            for j, d in enumerate(resp["data"]):
                oi = bidx[j]
                emb = d["embedding"]
                resultados[oi] = emb
                cache.put(textos[oi], emb)
        except Exception:
            # El lote falló: se reintenta ítem por ítem para no perder todo el bloque.
            st.session_state['embeddings_lotes_fallidos'] = st.session_state.get('embeddings_lotes_fallidos', 0) + 1
            for j, t in enumerate(batch):
                oi = bidx[j]
                try:
                    r = call_with_retries(openai.Embedding.create, input=[t], model=OPENAI_MODEL_EMBEDDING)
                    emb = r["data"][0]["embedding"]
                    resultados[oi] = emb
                    cache.put(textos[oi], emb)
                except Exception as e_item:
                    # Sin embedding esta nota no puede agruparse: se registra en lugar de silenciarse.
                    st.session_state['embeddings_fallidos'] = st.session_state.get('embeddings_fallidos', 0) + 1
                    st.session_state.setdefault('embeddings_errores', [])
                    if len(st.session_state['embeddings_errores']) < 5:
                        st.session_state['embeddings_errores'].append(f"{type(e_item).__name__}: {e_item}")
    return resultados

class DSU:
    def __init__(self, n):
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i):
        path = []
        while self.p[i] != i:
            path.append(i)
            i = self.p[i]
        for node in path: self.p[node] = i
        return i

    def union(self, i, j):
        ri, rj = self.find(i), self.find(j)
        if ri == rj: return
        if self.rank[ri] < self.rank[rj]: ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]: self.rank[ri] += 1

    def grupos(self, n):
        c = defaultdict(list)
        for i in range(n): c[self.find(i)].append(i)
        return dict(c)

def agrupar_textos_similares(textos, umbral):
    if not textos: return {}
    embs = get_embeddings_batch(textos)
    valid = [(i, e) for i, e in enumerate(embs) if e is not None]
    if len(valid) < 2: return {}
    idxs, M = zip(*valid)
    labels = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral, metric="cosine", linkage="average"
    ).fit(np.array(M)).labels_
    g = defaultdict(list)
    for k, lbl in enumerate(labels): g[lbl].append(idxs[k])
    return dict(enumerate(g.values()))

def agrupar_por_titulo_similar(titulos):
    gid, grupos, used = 0, {}, set()
    norm = [normalize_title_for_comparison(t) for t in titulos]
    for i in range(len(norm)):
        if i in used or not norm[i]: continue
        grp = [i]
        used.add(i)
        for j in range(i + 1, len(norm)):
            if j in used or not norm[j]: continue
            if SequenceMatcher(None, norm[i], norm[j]).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                grp.append(j)
                used.add(j)
        if len(grp) >= 2:
            grupos[gid] = list(set(grp))
            gid += 1
    return grupos

def seleccionar_representante(indices, textos):
    embs = get_embeddings_batch([textos[i] for i in indices])
    validos = [(indices[k], e) for k, e in enumerate(embs) if e is not None]
    if not validos: return indices[0], textos[indices[0]]
    idxs, M = zip(*validos)
    centro = np.mean(M, axis=0, keepdims=True)
    best = int(np.argmax(cosine_similarity(np.array(M), centro)))
    return idxs[best], textos[idxs[best]]


# ======================================
# TONO (Sistema Reputacional por IA)
# ======================================
class ClasificadorTono:
    def __init__(self, marca, aliases):
        self.marca = marca.strip()
        self.aliases = [a.strip() for a in (aliases or []) if a.strip()]
        self._all_names = [self.marca] + self.aliases

    def _menciona_marca(self, texto):
        t = _normalizar_mencion(texto)
        return any(_coincide_nombre_completo(t, nombre) for nombre in self._all_names)

    async def _clasificar_llm(self, texto, sem):
        async with sem:
            if not self._menciona_marca(texto):
                return {"tono": "Neutro"}

            aliases_str = f" (también conocida como: {', '.join(self.aliases)})" if self.aliases else ""
            prompt = (
                f"Eres un experto analista en Relaciones Públicas y Gestión de Reputación. "
                f"Tu tarea es evaluar el impacto reputacional DIRECTO de la siguiente noticia sobre la marca '{self.marca}'{aliases_str}.\n\n"
                f"TEXTO A EVALUAR:\n{texto[:1600]}\n\n"
                f"REGLAS DE CLASIFICACIÓN ESTRICTAS:\n"
                f"🔴 NEGATIVO: un hecho perjudica, cuestiona o expone directamente a '{self.marca}' "
                f"(demandas, multas, fraudes, fallas propias, quejas, investigaciones, pérdidas o retiro de productos).\n"
                f"🟢 POSITIVO: el hecho acredita directamente un logro, mejora o aporte verificable de '{self.marca}' "
                f"(premio, crecimiento, lanzamiento exitoso, inversión realizada, innovación, expansión o reconocimiento).\n"
                f"⚪ NEUTRO: La marca se menciona SIN impacto a su imagen. Ejemplos:\n"
                f"  - La noticia habla de una crisis del sector/país, pero la marca solo es mencionada informando o adaptándose.\n"
                f"  - Se menciona a la marca como patrocinador menor o en una lista de empresas.\n"
                f"  - Una persona, autoridad, proveedor o tercero es quien recibe el efecto positivo o negativo.\n"
                f"  - Emite un comunicado regular sin evidencia de crisis ni logro relevante.\n\n"
                f"⚠️ ATENCIÓN: Ignora si la noticia es trágica a nivel general (ej. una pandemia o accidente de terceros). "
                f"No infieras tono por palabras emocionales ni por el tono del sector. Evalúa ÚNICAMENTE cómo el hecho afecta "
                f"la reputación corporativa de '{self.marca}': mejora (Positivo), empeora (Negativo) o no cambia (Neutro).\n\n"
                f'Responde ÚNICAMENTE con JSON en este formato: {{"tono": "Positivo|Negativo|Neutro"}}'
            )

            try:
                resp = await acall_with_retries(
                    openai.ChatCompletion.acreate,
                    model=OPENAI_MODEL_CLASIFICACION,
                    messages=[{"role": "user", "content": prompt}],
                    max_tokens=40,
                    temperature=0.0,
                    response_format={"type": "json_object"}
                )
                
                u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
                if u:
                    st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                    st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0
                
                resultado = json.loads(resp.choices[0].message.content)
                tono = str(resultado.get("tono", "Neutro")).strip().title()
                
                return {"tono": tono if tono in ("Positivo", "Negativo", "Neutro") else "Neutro"}
            except Exception as e:
                return {"tono": "Neutro"}

    async def procesar_lote_async(self, textos, pbar, resumenes, titulos):
        n = len(textos)
        txts = textos.tolist()
        pbar.progress(0.05, "Agrupando noticias para análisis de tono...")
        
        txts_emb = [texto_para_embedding(str(titulos.iloc[i]), str(resumenes.iloc[i])) for i in range(n)]
        dsu = DSU(n)
        
        embs = get_embeddings_batch(txts_emb)
        candidatos = agrupar_textos_similares(txts_emb, SIMILARITY_THRESHOLD_TONO)
        candidatos.update({len(candidatos) + k: v for k, v in agrupar_por_titulo_similar(titulos.tolist()).items()})
        for idxs in candidatos.values():
            for pos, i in enumerate(idxs):
                for j in idxs[pos + 1:]:
                    ti, tj = normalize_title_for_comparison(titulos.iloc[i]), normalize_title_for_comparison(titulos.iloc[j])
                    titulo_casi_igual = SequenceMatcher(None, ti, tj).ratio() >= 0.96
                    contenido_casi_igual = (
                        embs[i] is not None and embs[j] is not None
                        and cosine_similarity(np.array(embs[i]).reshape(1, -1), np.array(embs[j]).reshape(1, -1))[0][0] >= SIMILARITY_THRESHOLD_TONO
                        and _overlap_distintivo(txts_emb[i], txts_emb[j]) >= 0.45
                    )
                    if (titulo_casi_igual or contenido_casi_igual) and not _hay_conflicto_accion(txts_emb[i], txts_emb[j]):
                        dsu.union(i, j)
                
        grupos = dsu.grupos(n)
        reps = {cid: seleccionar_representante(idxs, txts)[1] for cid, idxs in grupos.items()}
        
        sem = asyncio.Semaphore(CONCURRENT_REQUESTS)
        cids = list(reps.keys())
        
        async def _clasificar_con_cid(cid):
            return cid, await self._clasificar_llm(reps[cid], sem)

        tasks = [_clasificar_con_cid(c) for c in cids]
        rpg = {}
        
        for i, f in enumerate(asyncio.as_completed(tasks)):
            cid, r = await f
            rpg[cid] = r
            pbar.progress(0.1 + 0.85 * (i + 1) / len(tasks), f"Evaluando Reputación {i + 1}/{len(tasks)}")
            
        final = [None] * n
        
        for cid, idxs in grupos.items():
            r = rpg.get(cid, {"tono": "Neutro"})
            for i in idxs: final[i] = r
            
        pbar.progress(1.0, "Análisis de Tono completado")
        return final

def analizar_tono_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        TM = {1: "Positivo", "1": "Positivo", 0: "Neutro", "0": "Neutro", -1: "Negativo", "-1": "Negativo"}
        return [{"tono": TM.get(p, str(p).title())} for p in pipeline.predict(textos)]
    except Exception as e:
        st.error(f"Error pkl tono: {e}")
        return None

def analizar_temas_con_pkl(textos, pkl_file):
    try:
        pipeline = joblib.load(pkl_file)
        predicciones = pipeline.predict(textos)
        return [str(p).strip() for p in predicciones]
    except Exception as e:
        st.error(f"Error pkl temas: {e}")
        return None

# ======================================
# SUBTEMAS
# ======================================
class ClasificadorSubtema:
    def __init__(self, marca, aliases):
        self.marca = marca
        self.aliases = aliases or []
        self._cache = {}
        self._umbrales: dict = {}

    def _paso1(self, titulos, resumenes, dsu):
        def nt(t, n):
            return ' '.join(re.sub(r'[^a-z0-9\s]', '', unidecode(str(t).lower())).split()[:n])
        bt, br = defaultdict(list), defaultdict(list)
        for i, (ti, re_) in enumerate(zip(titulos, resumenes)):
            a, b = nt(ti, 40), nt(re_, 15)
            if a: bt[hashlib.md5(a.encode()).hexdigest()].append(i)
            b = nt(re_, 120)
            if len(b.split()) >= 25: br[hashlib.md5(b.encode()).hexdigest()].append(i)
        for bk in (bt, br):
            for idxs in bk.values():
                for j in idxs[1:]: dsu.union(idxs[0], j)

    def _paso2(self, titulos, dsu):
        norm = [normalize_title_for_comparison(t) for t in titulos]
        n = len(norm)
        for i in range(n):
            if not norm[i]: continue
            for j in range(i + 1, n):
                if not norm[j] or dsu.find(i) == dsu.find(j): continue
                ratio = SequenceMatcher(None, norm[i], norm[j]).ratio()
                comparte_asunto = _overlap_distintivo(norm[i], norm[j]) >= 0.40
                if ratio >= SIMILARITY_THRESHOLD_TITULOS and comparte_asunto and not _hay_conflicto_accion(norm[i], norm[j]):
                    dsu.union(i, j)

    def _paso2b_keywords(self, titulos, dsu, ae):
        sim_min = self._umbrales.get('sim_minima_keywords', SIM_MINIMA_KEYWORDS_RARAS)
        stop = {
            'el','la','los','las','un','una','unos','unas','de','del','al',
            'en','con','por','para','que','se','su','sus','es','son','fue',
            'como','mas','pero','sin','sobre','entre','tras','esta','este',
            'esto','hay','ser','han','ha','ya','muy','otro','otra','otros',
            'otras','todo','toda','todos','todas','puede','desde','hasta',
            'donde','cuando','quien','cual','cada','nos','les','ante','bajo',
            'nueva','nuevo','nuevos','nuevas','forma','hace','asi','sera',
            'segun','tiene','fueron','sido','hacer','dice','dijo','tambien',
        }
        titulo_words = []
        for t in titulos:
            ws = set()
            for w in re.findall(r'[a-z]+', unidecode(str(t).lower())):
                if len(w) >= 5 and w not in stop: ws.add(w)
            titulo_words.append(ws)
        word_freq = Counter()
        for ws in titulo_words:
            for w in ws: word_freq[w] += 1
        n = len(titulos)
        max_freq = max(2, int(n * 0.03))
        rare_index = defaultdict(list)
        for i, ws in enumerate(titulo_words):
            for w in ws:
                if 2 <= word_freq[w] <= max_freq: rare_index[w].append(i)
        for idxs in rare_index.values():
            for a in range(len(idxs)):
                for b in range(a + 1, len(idxs)):
                    ia, ib = idxs[a], idxs[b]
                    if dsu.find(ia) == dsu.find(ib): continue
                    ea, eb = ae[ia], ae[ib]
                    if ea is None or eb is None: continue
                    sim = cosine_similarity(
                        np.array(ea).reshape(1, -1),
                        np.array(eb).reshape(1, -1)
                    )[0][0]
                    if sim >= sim_min and not _hay_conflicto_accion(str(titulos[ia]), str(titulos[ib])):
                        dsu.union(ia, ib)

    def _paso3(self, et, ae, dsu, pbar, ps):
        umbral_cluster = max(self._umbrales.get('subtema', UMBRAL_SUBTEMA), 0.82)
        sim_min = max(self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA), 0.90)
        n = len(et)
        if n < 2: return

        def _puede_unir(i, j):
            if _hay_conflicto_accion(et[i], et[j]):
                return False
            if _overlap_distintivo(et[i], et[j]) >= 0.30:
                return True
            return SequenceMatcher(
                None,
                normalize_title_for_comparison(et[i]),
                normalize_title_for_comparison(et[j])
            ).ratio() >= 0.96

        B = 500
        if n <= B:
            pbar.progress(ps, "Clustering semántico...")
            ok = [(k, e) for k, e in enumerate(ae) if e is not None]
            if len(ok) < 2: return
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))
            linkage = 'complete' if n <= 10 else 'average'
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage=linkage
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels): g[lbl].append(io_[k])
            for cl in g.values():
                if len(cl) < 2: continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2: continue
                centroid = np.mean(vecs, axis=0)
                sims_al_centroid = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                todos_ok = all(s >= sim_min for s in sims_al_centroid)
                if todos_ok:
                    for j in cl[1:]:
                        if _puede_unir(cl[0], j):
                            dsu.union(cl[0], j)
                else:
                    mejor_idx = int(np.argmax(sims_al_centroid))
                    repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                    for k_local, i_global in enumerate(cl):
                        if ae[i_global] is None: continue
                        sim_vs_repr = cosine_similarity(
                            np.array(ae[i_global]).reshape(1, -1), repr_vec
                        )[0][0]
                        if sim_vs_repr >= sim_min and _puede_unir(cl[mejor_idx], i_global):
                            dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.18, "Clustering completado")
            return

        tb = max(1, (n + B - 1) // B)
        for bn_, bs in enumerate(range(0, n, B)):
            bi = list(range(bs, min(bs + B, n)))
            ok = [(idx, ae[idx]) for idx in bi if ae[idx] is not None]
            if len(ok) < 2: continue
            io_, M = zip(*ok)
            sim_matrix = cosine_similarity(np.array(M))
            labels = AgglomerativeClustering(
                n_clusters=None, distance_threshold=1 - umbral_cluster,
                metric='precomputed', linkage='average'
            ).fit(1 - sim_matrix).labels_
            g = defaultdict(list)
            for k, lbl in enumerate(labels): g[lbl].append(io_[k])
            for cl in g.values():
                if len(cl) < 2: continue
                vecs = np.array([ae[i] for i in cl if ae[i] is not None])
                if len(vecs) < 2: continue
                centroid = np.mean(vecs, axis=0)
                sims = cosine_similarity(vecs, centroid.reshape(1, -1)).flatten()
                mejor_idx = int(np.argmax(sims))
                repr_vec = np.array(ae[cl[mejor_idx]]).reshape(1, -1)
                for k_local, i_global in enumerate(cl):
                    if ae[i_global] is None: continue
                    s = cosine_similarity(np.array(ae[i_global]).reshape(1, -1), repr_vec)[0][0]
                    if s >= sim_min and _puede_unir(cl[mejor_idx], i_global):
                        dsu.union(cl[mejor_idx], i_global)
            pbar.progress(ps + 0.15 * (bn_ + 1) / tb, f"Clustering {bn_ + 1}/{tb}...")

        pbar.progress(ps + 0.16, "Unificando...")
        usar_fusion = self._umbrales.get('usar_fusion_iterativa', True)
        if usar_fusion: self._fusion(et, ae, dsu, pbar, ps + 0.16)

    def _fusion(self, textos, ae, dsu, pbar, ps):
        n = len(textos)
        umbral_inter = self._umbrales.get('fusion_intergrupo', UMBRAL_FUSION_INTERGRUPO)
        max_iter = self._umbrales.get('max_iter_fusion', MAX_ITER_FUSION)
        sim_min = self._umbrales.get('sim_minima_agrupacion', SIM_MINIMA_AGRUPACION_SUBTEMA)
        for it in range(max_iter):
            grupos = dsu.grupos(n)
            if len(grupos) < 2: break
            centroids, vg = [], []
            for gid, idxs in grupos.items():
                vecs = [ae[i] for i in idxs[:50] if ae[i] is not None]
                if vecs:
                    centroids.append(np.mean(vecs, axis=0))
                    vg.append(gid)
            if len(vg) < 2: break
            sim = cosine_similarity(np.array(centroids))
            umbral_efectivo = max(umbral_inter, sim_min)
            pairs = sorted(
                [(sim[i][j], i, j) for i in range(len(vg)) for j in range(i + 1, len(vg))
                 if sim[i][j] >= umbral_efectivo], reverse=True
            )
            fus = 0
            for _, i, j in pairs:
                ri, rj = grupos[vg[i]][0], grupos[vg[j]][0]
                if dsu.find(ri) != dsu.find(rj):
                    textos_i = [textos[k] for k in grupos[vg[i]][:20]]
                    textos_j = [textos[k] for k in grupos[vg[j]][:20]]
                    if _grupos_contenido_compatibles(
                        textos_i,
                        textos_j,
                        "",
                        "",
                        min_sim=umbral_efectivo,
                        min_overlap=0.16,
                    ):
                        dsu.union(ri, rj)
                        fus += 1
            pbar.progress(min(ps + 0.04 * (it + 1), 0.52), f"Fusión {it + 1}: {fus}")
            if fus == 0: break

    def _extraer_keywords_titulos(self, titulos_grp: list, top_n: int = 6) -> list:
        palabras = []
        for t in titulos_grp[:10]:
            for w in string_norm_label(t).split():
                if len(w) > 3: palabras.append(w)
        return [w for w, _ in Counter(palabras).most_common(top_n)]

    def _generar_etiqueta(self, textos_grp, titulos_grp, resumenes_grp, subtemas_existentes=None):
        tn = sorted(set(normalize_title_for_comparison(t) for t in titulos_grp if t))
        existentes_key = "|".join(sorted(string_norm_label(s) for s in (subtemas_existentes or []))[:20])
        ck = hashlib.md5(("|".join(tn[:12]) + f"#{len(titulos_grp)}#{existentes_key}").encode()).hexdigest()
        if ck in self._cache: return self._cache[ck]

        tm = list(dict.fromkeys(str(t)[:130] for t in titulos_grp if pd.notna(t) and str(t).strip() and str(t).strip().lower() != 'nan'))[:6]
        rm = [str(r)[:200] for r in resumenes_grp[:3] if r and len(str(r)) > 20]

        kw_list = self._extraer_keywords_titulos(titulos_grp, top_n=8)
        palabras_res = []
        for r in resumenes_grp[:5]:
            for w in string_norm_label(str(r)).split():
                if len(w) > 4: palabras_res.append(w)
        kw_res = [w for w, _ in Counter(palabras_res).most_common(4)
                  if w not in {unidecode(k.lower()) for k in kw_list}]
        kw_todos = kw_list + kw_res
        kw = ", ".join(kw_todos[:10])

        ctx_resumenes = (
            "\nRESÚMENES (para contexto):\n"
            + "\n".join(f"  · {r}" for r in rm)
        ) if rm else ""

        if len(kw_list) >= 3:
            ejemplo_dinamico = (
                f"'{kw_list[0].title()} de {kw_list[1].title()}' o "
                f"'{kw_list[0].title()} del {kw_list[2].title()}'"
            )
        elif len(kw_list) >= 2:
            ejemplo_dinamico = f"'{kw_list[0].title()} de {kw_list[1].title()}'"
        elif len(kw_list) == 1:
            ejemplo_dinamico = f"'{kw_list[0].title()} en la región'"
        else:
            ejemplo_dinamico = "'Proyecto de terminal de transportes'"

        lista_existentes = ""
        if subtemas_existentes and len(subtemas_existentes) > 0:
            lista_existentes = (
                "\n\nSUBTEMAS YA CREADOS (ÚSALOS SI APLICAN EXACTAMENTE):\n" +
                ", ".join(f"'{s}'" for s in subtemas_existentes[:15]) +
                "\nREGLA: Si este grupo de noticias trata EXACTAMENTE del mismo tema que uno de los subtemas ya creados, responde con ese subtema. Si es un tema distinto, crea uno nuevo."
            )

        prompt = (
            "Eres editor jefe de un periódico. "
            "Genera UN subtema periodístico (4-7 palabras) que sea una FRASE NOMINAL "
            "— sin sujeto ni verbo conjugado — para este grupo de noticias.\n\n"
            "TÍTULOS:\n" + "\n".join(f"  · {t}" for t in tm)
            + ctx_resumenes
            + f"\n\nPALABRAS CLAVE: {kw}"
            + lista_existentes
            + "\n\nREGLAS OBLIGATORIAS:\n"
            "  1. FRASE NOMINAL PURA: empieza con sustantivo, usa preposición para unir conceptos.\n"
            "     NUNCA empieces con cargo/persona ('Alcalde', 'Gobernador', 'Ministro').\n"
            "     NUNCA incluyas verbo conjugado ('presenta', 'anuncia', 'lanza', 'inaugura').\n"
            f"     CORRECTO: {ejemplo_dinamico}\n"
            "     INCORRECTO: 'Alcalde presenta proyecto terminal', "
            "'Gobernador anuncia inversión', 'Alcaldía lanza plan'\n"
            "  2. USA preposiciones (de, del, para, sobre, en, por) para conectar concepts.\n"
            "  3. SÉ ESPECÍFICO: describe el asunto real, no el actor.\n"
            "  4. Ciudades y regiones SÍ pueden aparecer si son relevantes al tema.\n"
            "  5. Sin nombre de marcas privadas. Tildes y ñ correctas.\n\n"
            "EJEMPLOS CORRECTOS: 'Proyecto de terminal de transportes', "
            "'Operación del Canal del Dique', 'Plan de infraestructura vial', "
            "'Regulación de tarifas eléctricas', 'Inversión en salud pública'\n"
            "EJEMPLOS INCORRECTOS: 'Alcalde presenta proyecto', 'Gobernador lanza plan', "
            "'Tarifas energía', 'Gestión corporativa', 'Actividad legislativa'\n\n"
            'JSON: {"subtema":"..."}'
        )

        _VERBOS_FRASES = re.compile(
            r'\b(presenta|presentan|anuncia|anuncian|lanza|lanzan|inaugura|inauguran|'
            r'realiza|realizan|desarrolla|desarrollan|ejecuta|ejecutan|gestiona|gestionan|'
            r'impulsa|impulsan|promueve|promueven|lidera|lideran|encabeza|encabezan|'
            r'aprueba|aprueban|firma|firman|suscribe|suscriben|invierte|invierten|'
            r'construye|construyen|instala|instalan|entrega|entregan|recibe|reciben|'
            r'solicita|solicitan|visita|visitan|atiende|atienden|destaca|destacan|'
            r'señala|señalan|indica|indican|expresa|expresan|afirma|afirman|'
            r'propone|proponen|pide|piden|exige|exigen|apoya|apoyan|'
            r'informa|informan|reporta|reportan|advierte|advierten)\b',
            re.IGNORECASE
        )

        def _tiene_verbo_conjugado(s): return bool(_VERBOS_FRASES.search(s))

        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            u = resp.get('usage', {}) if isinstance(resp, dict) else getattr(resp, 'usage', {})
            if u:
                st.session_state['tokens_input'] += (u.get('prompt_tokens') if isinstance(u, dict) else getattr(u, 'prompt_tokens', 0)) or 0
                st.session_state['tokens_output'] += (u.get('completion_tokens') if isinstance(u, dict) else getattr(u, 'completion_tokens', 0)) or 0

            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema_geografico(limpiar_tema(raw), self.marca, self.aliases)

            if not et or et.strip().lower() == "sin tema":
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)
            if _tiene_verbo_conjugado(et):
                et = self._refinar(tm, kw, rm, forzar_preposicion=True, prohibir_verbos=True)

            def _es_robotico(s):
                palabras = s.split()
                if len(palabras) <= 3:
                    nexos = {"de", "del", "para", "sobre", "en", "con", "por",
                             "ante", "hacia", "entre", "sin", "al", "las", "los",
                             "una", "uno", "que", "como", "y", "o", "a", "e", "u"}
                    tiene_nexo = any(unidecode(p.lower()) in nexos for p in palabras[1:])
                    if not tiene_nexo: return True
                return False

            genericas = {"gestión", "gestion", "actividades", "acciones", "noticias",
                         "información", "informacion", "eventos", "varios", "sin tema",
                         "actividad corporativa", "gestion corporativa"}
            es_gen = string_norm_label(et) in {string_norm_label(g) for g in genericas}
            es_rob = _es_robotico(et)

            if es_gen or es_rob or len(et.split()) < 3:
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)

            if not _validar_estructura_subtema(et):
                et = self._refinar(tm, kw, rm, forzar_preposicion=True)
                if not _validar_estructura_subtema(et):
                    et = self._fallback(titulos_grp)

            et = _validar_etiqueta_completa(
                et, titulos_grp=titulos_grp, resumenes_grp=resumenes_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=self._fallback
            )
        except:
            et = self._fallback(titulos_grp)

        et = capitalizar_etiqueta(et)
        self._cache[ck] = et
        return et

    def _refinar(self, titulos, kw, resumenes=None, forzar_preposicion=False, prohibir_verbos=False):
        ctx = ("\nContexto de resúmenes: " + " | ".join(r[:100] for r in resumenes[:3])) if resumenes else ""
        kw_parts = [w.strip() for w in kw.split(",") if w.strip()]

        if len(kw_parts) >= 3:
            ej_bueno = f"'{kw_parts[0].title()} de {kw_parts[1].title()}', '{kw_parts[0].title()} en {kw_parts[2].title()}'"
        elif len(kw_parts) >= 2:
            ej_bueno = f"'{kw_parts[0].title()} de {kw_parts[1].title()}'"
        elif len(kw_parts) == 1:
            ej_bueno = f"'{kw_parts[0].title()} en la región'"
        else:
            ej_bueno = "'Proyecto de terminal de transportes'"

        ej_malo = f"'{kw_parts[0].title()} {kw_parts[1].title()}' (sin preposición)" if len(kw_parts) >= 2 else "'Actividad corporativa', 'Gestión institucional'"

        instruccion_prep = (
            "  OBLIGATORIO: usa una preposición (de, del, para, sobre, en, por) "
            "entre los conceptos. NUNCA dos sustantivos pegados sin nexo.\n"
        ) if forzar_preposicion else ""

        instruccion_verbo = (
            "  PROHIBIDO: verbos conjugados ('presenta', 'anuncia', 'lanza', 'inaugura', etc.). "
            "Solo frases nominales (sustantivos + preposiciones).\n"
            "  NUNCA empieces con cargo ('Alcalde', 'Gobernador', 'Ministro', 'Director').\n"
        ) if prohibir_verbos else ""

        prompt = (
            "Eres editor jefe. Genera UN subtema periodístico (4-7 palabras) "
            "como frase nominal sin verbo conjugado.\n\n"
            f"Títulos: {' | '.join(titulos[:5])}{ctx}\n"
            f"Keywords: {kw}\n\n"
            f"{instruccion_prep}{instruccion_verbo}"
            f"CORRECTO: {ej_bueno}, 'Tarifas de energía eléctrica'\n"
            f"INCORRECTO: {ej_malo}, 'Alcalde presenta plan'\n"
            "Tildes y ñ correctas. Sin marcas privadas.\n"
            'JSON: {"subtema":"..."}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=60,
                temperature=0.2,
                response_format={"type": "json_object"}
            )
            raw = json.loads(resp.choices[0].message.content).get("subtema", "Varios")
            et = limpiar_tema_geografico(limpiar_tema(raw), self.marca, self.aliases)
            if not _frase_esta_completa(et):
                et = _recortar_frase_completa(et)
                if not _frase_esta_completa(et): return self._fallback(titulos)
            return et
        except:
            return self._fallback([])

    def _fallback(self, titulos):
        if not titulos: return "Cobertura de información relevante"
        palabras = []
        for t in titulos[:5]:
            for w in string_norm_label(t).split():
                if len(w) > 4: palabras.append(w)
        if palabras:
            top = [w for w, _ in Counter(palabras).most_common(3)]
            if len(top) >= 2:
                frase = f"{top[0]} de {top[1]}"
                if _frase_esta_completa(frase): return capitalizar_etiqueta(frase)
                return capitalizar_etiqueta(f"Asuntos de {top[0]} y {top[1]}")
            return capitalizar_etiqueta(f"Asuntos relacionados con {top[0]}")
        return "Cobertura de información relevante"

    def _consolidar_sinonimos_llm(self, subtemas_unicos):
        if len(subtemas_unicos) <= 1:
            return {s: s for s in subtemas_unicos}
            
        prompt = (
            "Eres un analista de datos. Tienes la siguiente lista de subtemas periodísticos:\n"
            f"{', '.join(subtemas_unicos)}\n\n"
            "Tu tarea es encontrar SUBTEMAS SINÓNIMOS que signifiquen exactamente lo mismo "
            "(aunque usen palabras ligeramente distintas) y unificarlos bajo el nombre más claro y representativo.\n"
            "REGLAS:\n"
            "1. NO fusiones temas que sean distintos (ej. 'Inversión en vías' y 'Mantenimiento de vías' son distintos).\n"
            "2. SÍ fusiona sinónimos (ej. 'Lanzamiento de plataforma web' y 'Estreno de portal digital').\n"
            "3. Devuelve un objeto JSON donde las claves sean los subtemas originales y el valor sea el subtema unificado.\n\n"
            'Ejemplo de salida:\n'
            '{"Tendencias de consumo de pollo": "Tendencias de consumo de pollo", "Hábitos de compra de aves": "Tendencias de consumo de pollo"}'
        )
        try:
            resp = call_with_retries(
                openai.ChatCompletion.create,
                model=OPENAI_MODEL_CLASIFICACION,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=1000,
                temperature=0.0,
                response_format={"type": "json_object"}
            )
            return json.loads(resp.choices[0].message.content)
        except:
            return {s: s for s in subtemas_unicos}

    def procesar_lote(self, col, pbar, res_puros, tit_puros):
        textos   = col.tolist()
        titulos  = tit_puros.tolist()
        resumenes = res_puros.tolist()
        n = len(textos)

        self._umbrales = _umbrales_adaptativos(n)
        u = self._umbrales
        st.caption(
            f"📐 Corpus: **{n}** noticias · Umbral subtema: **{u['subtema']}** · "
            f"Sim mínima: **{u['sim_minima_agrupacion']}**"
        )

        et = [texto_para_embedding(titulos[i], resumenes[i]) for i in range(n)]

        pbar.progress(0.05, "Fase 1 · Idénticas...")
        dsu = DSU(n)
        self._paso1(titulos, resumenes, dsu)
        
        pbar.progress(0.12, "Fase 2 · Títulos...")
        self._paso2(titulos, dsu)

        pbar.progress(0.18, "Embeddings...")
        ae = get_embeddings_batch(et)

        if u['usar_paso2b']:
            pbar.progress(0.15, "Fase 2b · Keywords raras (con validación semántica)...")
            self._paso2b_keywords(titulos, dsu, ae)

        pbar.progress(0.20, "Fase 3 · Clustering...")
        self._paso3(et, ae, dsu, pbar, 0.20)

        gf = dsu.grupos(n)
        ng = len(gf)
        pbar.progress(0.55, f"Fase 4 · Etiquetando {ng} grupos...")
        mapa = {}
        sg = sorted(gf.items(), key=lambda x: -len(x[1]))
        subtemas_aprobados = [] 
        textos_por_subtema_aprobado = defaultdict(list)

        def _generar_etiqueta_segura(idxs):
            textos_grp = [textos[i] for i in idxs]
            titulos_grp = [titulos[i] for i in idxs]
            resumenes_grp = [resumenes[i] for i in idxs]
            etiqueta = self._generar_etiqueta(
                textos_grp,
                titulos_grp,
                resumenes_grp,
                subtemas_existentes=subtemas_aprobados
            )
            if etiqueta in textos_por_subtema_aprobado:
                previos = textos_por_subtema_aprobado.get(etiqueta, [])
                if not _grupos_contenido_compatibles(
                    textos_grp,
                    previos,
                    etiqueta,
                    etiqueta,
                    min_sim=max(u['sim_minima_agrupacion'], 0.88),
                    min_overlap=0.24,
                ):
                    etiqueta = self._generar_etiqueta(
                        textos_grp,
                        titulos_grp,
                        resumenes_grp,
                        subtemas_existentes=None
                    )
            if etiqueta not in subtemas_aprobados:
                subtemas_aprobados.append(etiqueta)
            textos_por_subtema_aprobado[etiqueta].extend(textos_grp[:MAX_GRUPO_ETIQUETA])
            return etiqueta

        for k, (lid, idxs) in enumerate(sg):
            if k % 10 == 0: pbar.progress(0.55 + 0.25 * (k / max(ng, 1)), f"Etiquetando {k + 1}/{ng}...")
            
            if len(idxs) > MAX_GRUPO_ETIQUETA:
                subgrupos = [idxs[i:i + MAX_GRUPO_ETIQUETA] for i in range(0, len(idxs), MAX_GRUPO_ETIQUETA)]
                for sg_ in subgrupos:
                    e = _generar_etiqueta_segura(sg_)
                    for i in sg_: mapa[i] = e
            else:
                e = _generar_etiqueta_segura(idxs)
                for i in idxs: mapa[i] = e

        subtemas = [mapa.get(i, "Varios") for i in range(n)]

        pbar.progress(0.80, "Fase 4b · Coherencia etiqueta↔texto...")
        umbral_coherencia = u['coherencia_etiqueta']
        subtemas_unicos = list(set(subtemas))
        embs_sub_lista = get_embeddings_batch(subtemas_unicos)
        emb_subtemas = {sub: emb for sub, emb in zip(subtemas_unicos, embs_sub_lista) if emb is not None}

        incoherentes = 0
        for i in range(n):
            sub = subtemas[i]
            emb_txt = ae[i]
            emb_sub = emb_subtemas.get(sub)
            if emb_txt is None or emb_sub is None: continue
            sim = cosine_similarity(np.array(emb_txt).reshape(1, -1), np.array(emb_sub).reshape(1, -1))[0][0]
            if sim < umbral_coherencia:
                mejor_sub, mejor_sim = sub, sim
                for otro_sub, emb_otro in emb_subtemas.items():
                    if otro_sub == sub: continue
                    sim_otro = cosine_similarity(np.array(emb_txt).reshape(1, -1), np.array(emb_otro).reshape(1, -1))[0][0]
                    if sim_otro > mejor_sim: mejor_sim = sim_otro; mejor_sub = otro_sub
                if mejor_sub != sub and mejor_sim > umbral_coherencia:
                    subtemas[i] = mejor_sub
                else:
                    nueva = self._generar_etiqueta([textos[i]], [titulos[i]], [resumenes[i]], subtemas_existentes=subtemas_aprobados)
                    subtemas[i] = capitalizar_etiqueta(nueva)
                    if nueva not in subtemas_aprobados: subtemas_aprobados.append(nueva)
                incoherentes += 1

        pbar.progress(0.82, "Fase 5 · Dedup...")
        subtemas = dedup_labels(subtemas, u['dedup_label'])

        pbar.progress(0.86, "Fase 5b · Fusión semántica...")
        textos_por_sub = defaultdict(list)
        for i, s in enumerate(subtemas): textos_por_sub[s].append(textos[i])
        subtemas = _fusionar_subtemas_semanticos(subtemas, textos_por_sub, self.marca, self.aliases, u['fusion_subtemas'])

        pbar.progress(0.90, "Fase 6 · Consistencia...")
        subtemas = self._consistencia(subtemas, ae, pbar, u)

        indices_reclass = [i for i, s in enumerate(subtemas) if s == "_RECLASSIFICAR"]
        if indices_reclass:
            pbar.progress(0.93, f"Fase 6b · Reclasificando...")
            for i in indices_reclass:
                et_ind = self._generar_etiqueta([textos[i]], [titulos[i]], [resumenes[i]], subtemas_existentes=subtemas_aprobados)
                subtemas[i] = capitalizar_etiqueta(et_ind)
                if et_ind not in subtemas_aprobados: subtemas_aprobados.append(et_ind)

        pbar.progress(0.93, "Fase 7 · Completitud...")
        subtemas = self._validar_completitud_final(subtemas, textos, titulos, resumenes)

        pbar.progress(0.97, "Fase 8 · Dedup final...")
        subtemas = dedup_labels(subtemas, u['dedup_label'])
        
        pbar.progress(0.99, "Consolidación final IA de sinónimos...")
        unicos_finales = list(dict.fromkeys(subtemas))
        if 1 < len(unicos_finales) <= 50:
            mapa_sinonimos = self._consolidar_sinonimos_llm(unicos_finales)
            textos_por_sub_final = defaultdict(list)
            for i, s in enumerate(subtemas):
                textos_por_sub_final[s].append(textos[i])
            mapa_seguro = {}
            for origen, destino in mapa_sinonimos.items():
                if origen == destino:
                    mapa_seguro[origen] = destino
                    continue
                if destino not in textos_por_sub_final:
                    continue
                if _grupos_contenido_compatibles(
                    textos_por_sub_final.get(origen, []),
                    textos_por_sub_final.get(destino, []),
                    origen,
                    destino,
                    min_sim=max(u['fusion_subtemas'], 0.88),
                    min_overlap=0.24,
                ):
                    mapa_seguro[origen] = destino
            subtemas = [mapa_seguro.get(s, s) for s in subtemas]

        subtemas = [capitalizar_etiqueta(s) for s in subtemas]
        nf = len(set(subtemas))
        pbar.progress(1.0, f"{nf} subtemas")
        st.info(f"Subtemas: **{nf}** · Grupos originales: **{ng}**")
        return subtemas

    def _validar_completitud_final(self, subtemas, textos, titulos, resumenes):
        por_subtema = defaultdict(list)
        for i, s in enumerate(subtemas): por_subtema[s].append(i)
        resultado = list(subtemas)
        for sub, idxs in por_subtema.items():
            if _frase_esta_completa(sub): continue
            recortada = _recortar_frase_completa(sub)
            if _frase_esta_completa(recortada) and len(recortada.split()) >= 2:
                for i in idxs: resultado[i] = capitalizar_etiqueta(recortada)
                continue
            tit_grp = [titulos[i] for i in idxs[:6]]
            res_grp = [resumenes[i] for i in idxs[:3]]
            nueva = _validar_etiqueta_completa(
                sub, titulos_grp=tit_grp, resumenes_grp=res_grp,
                marca=self.marca, aliases=self.aliases, fallback_fn=self._fallback
            )
            for i in idxs: resultado[i] = capitalizar_etiqueta(nueva)
        return resultado

    def _consistencia(self, subtemas, ae, pbar, umbrales=None):
        min_sub = umbrales.get('min_pertenencia_subtema', UMBRAL_MIN_PERTENENCIA_SUBTEMA)
        ps = defaultdict(list)
        for i, s in enumerate(subtemas): ps[s].append(i)
        r = list(subtemas)
        centroids = {}
        for sub, idxs in ps.items():
            vecs = [ae[i] for i in idxs if ae[i] is not None]
            if vecs: centroids[sub] = np.mean(vecs, axis=0)
        for sub in [s for s in centroids if len(ps[s]) >= 3]:
            idxs = ps[sub]
            if sub.lower() in ("sin tema", "varios") or len(idxs) < 3: continue
            vi = [(i, ae[i]) for i in idxs if ae[i] is not None]
            if len(vi) < 3: continue
            v_i, v_v = zip(*vi)
            M = np.array(v_v)
            sims = cosine_similarity(M, centroids[sub].reshape(1, -1)).flatten()
            thr = max(0.60, np.mean(sims) - 2 * np.std(sims))
            for k, (oi, sv) in enumerate(zip(v_i, sims)):
                if sv >= thr: continue
                bs, bsim = sub, sv
                emb = ae[oi]
                for os_, oc in centroids.items():
                    if os_ == sub: continue
                    s2 = cosine_similarity(np.array(emb).reshape(1, -1), oc.reshape(1, -1))[0][0]
                    if s2 > bsim and s2 > 0.75: bsim = s2; bs = os_
                if bs != sub: r[oi] = bs
                elif sv < min_sub: r[oi] = "_RECLASSIFICAR"
        return r

# ======================================
# TEMAS  
# ======================================
def _construir_representacion_grupo(subtema, textos_grupo, max_textos=30):
    palabras = []
    for t in textos_grupo[:max_textos]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3: palabras.append(w)
    kw_str = " ".join(w for w, _ in Counter(palabras).most_common(12))
    return f"{subtema}. {subtema}. {kw_str}"[:500]

def _validar_estructura_tema(tema: str) -> bool:
    if not tema or len(tema.split()) < 2: return False
    if len(tema.split()) > 5: return False
    if re.match(r'^[0-9]', tema): return False
    num_palabras = re.compile(
        r'^(uno|dos|tres|cuatro|cinco|seis|siete|ocho|nueve|diez|'
        r'once|doce|veinte|cien|varios|cada)', re.IGNORECASE
    )
    if num_palabras.match(tema): return False
    if _PATRON_TITULAR.match(tema): return False
    if _PATRON_ESTADO.search(tema): return False
    genericos = {
        "economia", "politica", "tecnologia", "seguridad", "justicia",
        "actualidad", "nacional", "internacional", "empresas", "sociedad",
        "negocios", "informacion", "noticias", "varios", "general",
    }
    if string_norm_label(tema) in genericos: return False
    return True

def _tema_es_igual_a_subtema(tema: str, subtemas_grupo: list) -> bool:
    if not tema or not subtemas_grupo: return False
    tn = string_norm_label(tema)
    for sub in subtemas_grupo:
        sn = string_norm_label(sub)
        if not tn or not sn: continue
        if SequenceMatcher(None, tn, sn).ratio() >= 0.80: return True
        if tn in sn or sn in tn: return True
    return False

def _generar_nombre_tema_llm(subtemas_grupo, textos_muestra, titulos_muestra):
    subs_list = "\n".join(f"  · {s}" for s in subtemas_grupo[:8])
    palabras = []
    for t in titulos_muestra[:15]:
        for w in string_norm_label(str(t)).split():
            if len(w) > 3: palabras.append(w)
    kw = ", ".join(w for w, _ in Counter(palabras).most_common(6))
    tit_muestra = "\n".join(f"  · {t[:100]}" for t in list(dict.fromkeys(titulos_muestra))[:5])
    prompt = (
        "Eres editor jefe. Crea UN tema editorial preciso (2-5 palabras) que agrupe estos subtemas.\n\n"
        "SUBTEMAS:\n" + subs_list + "\n\nTÍTULOS DE REFERENCIA:\n" + tit_muestra +
        f"\n\nKEYWORDS: {kw}\n\n"
        "REGLAS ESTRICTAS:\n"
        "  1. Conserva el asunto común que diferencia este grupo; NO uses secciones vagas de una palabra.\n"
        "  2. Debe ser más general que los subtemas, pero no abstracto: nunca copies un titular ni repitas un subtema.\n"
        "  3. NUNCA incluyas números, cantidades ni nombres propios.\n"
        "  4. 2-5 palabras, sustantivo + complemento/adjetivo.\n"
        "  5. Tildes y ñ correctas.\n\n"
        "CORRECTO: 'Regulación financiera', 'Movilidad urbana', 'Infraestructura vial', 'Salud pública territorial'\n"
        "INCORRECTO: 'Economía', 'Política', 'Actualidad', 'Cinco congresistas con líos', 'Nuevo acuerdo'\n\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=40,
            temperature=0.05,
            response_format={"type": "json_object"}
        )
        raw = json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', '')
        nombre = limpiar_tema(raw)
        if not _validar_estructura_tema(nombre): return None
        return nombre
    except:
        return None

def _regenerar_tema_diferente(subtemas_grupo, titulos_muestra, intento=0):
    subs_list = ", ".join(subtemas_grupo[:8])
    prompt = (
        f"Subtemas: {subs_list}\n\n"
        "Genera UNA categoría precisa (2-5 palabras), diferente a los subtemas. "
        "Conserva el asunto común; no respondas una sección vaga de una palabra como Economía, Política o Actualidad. "
        "Tildes y ñ correctas, terminar en sustantivo/adjetivo.\n"
        'JSON: {"tema":"..."}'
    )
    try:
        resp = call_with_retries(
            openai.ChatCompletion.create,
            model=OPENAI_MODEL_CLASIFICACION,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=50,
            temperature=0.2 + intento * 0.1,
            response_format={"type": "json_object"}
        )
        nombre = limpiar_tema(json.loads(resp.choices[0].message.content).get("tema", "").strip().replace('"', '').replace('.', ''))
        return nombre if _validar_estructura_tema(nombre) else None
    except:
        return None

def consolidar_temas(subtemas, textos, pbar):
    n = len(textos)
    u = _umbrales_adaptativos(n)
    pbar.progress(0.05, "Preparando temas...")
    df = pd.DataFrame({'subtema': subtemas, 'texto': textos})
    us = list(df['subtema'].unique())
    if len(us) <= 1:
        pbar.progress(1.0, "Un tema")
        return [capitalizar_etiqueta(s) for s in subtemas]

    if n <= 5 and len(us) == n:
        pbar.progress(1.0, "Corpus pequeño: temas = subtemas")
        st.info(f"Temas: **{n}** (corpus pequeño — cada noticia tiene tema propio)")
        return [capitalizar_etiqueta(s) for s in subtemas]

    pbar.progress(0.10, "Representaciones...")
    textos_por_subtema = defaultdict(list)
    for i, sub in enumerate(subtemas): textos_por_subtema[sub].append(textos[i])
    repr_enriquecidas = [_construir_representacion_grupo(sub, textos_por_subtema[sub]) for sub in us]
    pbar.progress(0.20, "Embeddings contenido...")
    emb_repr = get_embeddings_batch(repr_enriquecidas)
    emb_labels = get_embeddings_batch(us)
    ae = get_embeddings_batch(textos)
    centroids_contenido = {}
    for sub in us:
        idxs = df.index[df['subtema'] == sub].tolist()[:50]
        vecs = [ae[i] for i in idxs if ae[i] is not None]
        if vecs: centroids_contenido[sub] = np.mean(vecs, axis=0)
    pbar.progress(0.35, "Similitudes...")
    vs = [s for s in us if s in centroids_contenido]
    if len(vs) < 2:
        pbar.progress(1.0, "Sin agrupación")
        return [capitalizar_etiqueta(s) for s in subtemas]
    idx_map = {s: i for i, s in enumerate(us)}
    M_content = np.array([centroids_contenido[s] for s in vs])
    sim_content = cosine_similarity(M_content)
    has_repr = all(emb_repr[idx_map[s]] is not None for s in vs)
    has_label = all(emb_labels[idx_map[s]] is not None for s in vs)
    if has_repr and has_label:
        sim_combined = (0.50 * sim_content + 0.35 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs])) + 0.15 * cosine_similarity(np.array([emb_labels[idx_map[s]] for s in vs])))
    elif has_repr:
        sim_combined = (0.60 * sim_content + 0.40 * cosine_similarity(np.array([emb_repr[idx_map[s]] for s in vs])))
    else:
        sim_combined = sim_content

    pbar.progress(0.45, "Clustering temas...")
    dist_matrix = np.clip(1 - sim_combined, 0, 2)
    np.fill_diagonal(dist_matrix, 0)
    umbral_tema = u['tema']
    num_temas_max = u['num_temas_max']
    linkage_temas = 'complete' if len(vs) <= 6 else 'average'
    cl = AgglomerativeClustering(
        n_clusters=None, distance_threshold=1 - umbral_tema,
        metric='precomputed', linkage=linkage_temas
    ).fit(dist_matrix)

    clusters = defaultdict(list)
    for i, lbl in enumerate(cl.labels_): clusters[lbl].append(vs[i])
    clusters_validados = {}
    next_cluster_id = 0
    for _, subs_cluster in clusters.items():
        if len(subs_cluster) <= 1:
            clusters_validados[next_cluster_id] = subs_cluster
            next_cluster_id += 1
            continue
        dsu_tema = DSU(len(subs_cluster))
        for i in range(len(subs_cluster)):
            for j in range(i + 1, len(subs_cluster)):
                sa, sb = subs_cluster[i], subs_cluster[j]
                if _grupos_contenido_compatibles(
                    textos_por_subtema.get(sa, []),
                    textos_por_subtema.get(sb, []),
                    sa,
                    sb,
                    min_sim=max(umbral_tema, 0.82),
                    min_overlap=0.16,
                ):
                    dsu_tema.union(i, j)
        for miembros in dsu_tema.grupos(len(subs_cluster)).values():
            clusters_validados[next_cluster_id] = [subs_cluster[i] for i in miembros]
            next_cluster_id += 1
    clusters = clusters_validados
    uc = [s for s in us if s not in vs]
    mt = {}
    tc = len(clusters)
    pbar.progress(0.50, f"Nombres {tc} temas...")
    for k, (cid, subtemas_cluster) in enumerate(clusters.items()):
        pbar.progress(0.50 + 0.35 * (k / max(tc, 1)), f"Tema {k + 1}/{tc}...")
        titulos_cluster = []
        textos_cluster = []
        for sub in subtemas_cluster:
            for idx in df.index[df['subtema'] == sub].tolist()[:10]:
                txt = str(textos[idx])
                partes = txt.split('. ')
                if partes: titulos_cluster.append(partes[0][:120])
                textos_cluster.append(txt[:200])
        if len(subtemas_cluster) == 1:
            sub_unico = subtemas_cluster[0]
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                p = sub_unico.split()
                nombre = _recortar_frase_completa(" ".join(p), max_palabras=3) if len(p) > 3 else sub_unico
                if _tema_es_igual_a_subtema(nombre, subtemas_cluster): nombre = sub_unico
        else:
            nombre = _generar_nombre_tema_llm(subtemas_cluster, textos_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                nombre = _regenerar_tema_diferente(subtemas_cluster, titulos_cluster, intento=1)
            if not nombre or _tema_es_igual_a_subtema(nombre, subtemas_cluster):
                all_words = []
                for sub in subtemas_cluster:
                    for w in string_norm_label(sub).split():
                        if len(w) > 3: all_words.append(w)
                nombre = capitalizar_etiqueta(" ".join(w for w, _ in Counter(all_words).most_common(2))) if all_words else subtemas_cluster[0]
        if not _frase_esta_completa(nombre):
            nombre = _recortar_frase_completa(nombre, max_palabras=4)
            if not _frase_esta_completa(nombre):
                freq = Counter(subtemas)
                nombre = _recortar_frase_completa(max(subtemas_cluster, key=lambda s: freq.get(s, 0)), max_palabras=4)
        nombre = capitalizar_etiqueta(nombre)
        for sub in subtemas_cluster: mt[sub] = nombre
    for sub in uc: mt[sub] = capitalizar_etiqueta(sub)

    pbar.progress(0.87, "Validando pertenencia mínima a temas...")
    min_tema = u['min_pertenencia_tema']
    tf_inicial = [mt.get(sub, sub) for sub in subtemas]
    tema_agrupacion: Dict[str, list] = defaultdict(list)
    for i, tema in enumerate(tf_inicial):
        if ae[i] is not None: tema_agrupacion[tema].append(ae[i])
    tema_centroids: Dict[str, np.ndarray] = {
        t: np.mean(vecs, axis=0) for t, vecs in tema_agrupacion.items() if vecs
    }
    tf_validado: List[str] = []
    n_forzadas = 0
    for i, (sub, tema_asignado) in enumerate(zip(subtemas, tf_inicial)):
        emb = ae[i]
        if emb is not None and tema_asignado in tema_centroids:
            sim = cosine_similarity(np.array(emb).reshape(1, -1), tema_centroids[tema_asignado].reshape(1, -1))[0][0]
            if sim < min_tema:
                tf_validado.append(capitalizar_etiqueta(_recortar_frase_completa(sub, max_palabras=4)))
                n_forzadas += 1
                continue
        tf_validado.append(capitalizar_etiqueta(tema_asignado))
    if n_forzadas: st.caption(f"ℹ️ {n_forzadas} noticias con baja pertenencia al tema agrupado → tema propio asignado.")

    pbar.progress(0.88, "Dedup temas...")
    tf_validado = dedup_labels(tf_validado, u['dedup_label'])

    pbar.progress(0.90, "Fusionando temas solapados...")
    mapa_fusion_temas = _fusionar_temas_contenidos(tf_validado)
    if mapa_fusion_temas:
        tf_validado = [mapa_fusion_temas.get(t, t) for t in tf_validado]

    pbar.progress(0.92, "Validando tema ≠ subtema...")
    tf_validado = _post_validar_tema_vs_subtema(tf_validado, subtemas)
    pbar.progress(0.95, "Completitud...")
    tf_validado = [capitalizar_etiqueta(_recortar_frase_completa(t) if not _frase_esta_completa(t) else t) for t in tf_validado]
    tf_validado = _unificar_tema_por_subtema(tf_validado, subtemas)
    st.info(f"Temas: **{len(set(tf_validado))}** (de {len(set(subtemas))} subtemas) · Máx: {num_temas_max}")
    pbar.progress(1.0, "Temas listos")
    return tf_validado

def _fusionar_temas_contenidos(temas: List[str]) -> Dict[str, str]:
    unique = list(dict.fromkeys(temas))
    if len(unique) < 2: return {}
    normed = {t: string_norm_label(t) for t in unique}
    mapa: Dict[str, str] = {}
    for i, ta in enumerate(unique):
        for tb in unique[i + 1:]:
            na, nb = normed[ta], normed[tb]
            if not na or not nb: continue
            if na == nb or SequenceMatcher(None, na, nb).ratio() >= 0.92:
                canon = tb if len(tb) >= len(ta) else ta
                reemplazar = ta if canon == tb else tb
                mapa[reemplazar] = canon
    umbral_relajado = 0.88
    candidatos = [(t, normed[t]) for t in unique if len(t.split()) <= 3 and t not in mapa]
    if len(candidatos) >= 2:
        textos_c = [t for t, _ in candidatos]
        embs = get_embeddings_batch(textos_c)
        validos = [(textos_c[i], embs[i]) for i in range(len(textos_c)) if embs[i] is not None]
        if len(validos) >= 2:
            etqs, vecs = zip(*validos)
            sim = cosine_similarity(np.array(vecs))
            for i in range(len(etqs)):
                for j in range(i + 1, len(etqs)):
                    if sim[i][j] >= umbral_relajado:
                        ta, tb = etqs[i], etqs[j]
                        if ta in mapa or tb in mapa: continue
                        if _etiquetas_compatibles(ta, tb, min_overlap=0.60):
                            freq = Counter(temas)
                            canon = ta if freq.get(ta, 0) >= freq.get(tb, 0) else tb
                            reemplazar = tb if canon == ta else ta
                            mapa[reemplazar] = canon
    return mapa

def _post_validar_tema_vs_subtema(temas, subtemas):
    tema_a_subtemas = defaultdict(set)
    for t, s in zip(temas, subtemas): tema_a_subtemas[t].add(s)
    reemplazos = {}
    for tema, subs in tema_a_subtemas.items():
        if len(subs) == 1:
            sub_unico = list(subs)[0]
            tn = string_norm_label(tema)
            sn = string_norm_label(sub_unico)
            if tn and sn and SequenceMatcher(None, tn, sn).ratio() >= 0.80:
                nuevo = _regenerar_tema_diferente([sub_unico], [])
                if nuevo and not _tema_es_igual_a_subtema(nuevo, [sub_unico]) and _frase_esta_completa(nuevo):
                    reemplazos[tema] = capitalizar_etiqueta(nuevo)
    return [reemplazos.get(t, t) for t in temas] if reemplazos else temas

def _unificar_tema_por_subtema(temas, subtemas):
    sub_to_temas = defaultdict(list)
    for t, s in zip(temas, subtemas): sub_to_temas[s].append(t)
    sub_to_best = {}
    for sub, tema_list in sub_to_temas.items():
        sub_to_best[sub] = Counter(tema_list).most_common(1)[0][0]
    return [sub_to_best[s] for s in subtemas]

# ======================================
# Duplicados y Excel (Reglas Nuevas)
# ======================================
def _normalizar_url(url: str) -> str:
    if not url: return ""
    url = url.strip().lower()
    url = re.sub(r'^https?://', '', url)
    url = re.sub(r'^www\.', '', url)
    url = url.rstrip('/')
    return url

def detectar_duplicados_avanzado(rows, km):
    processed = deepcopy(rows)
    seen_url, seen_bcast = {}, {}
    seen_streaming: Dict[tuple, int] = {}
    tb = defaultdict(list)

    for i, row in enumerate(processed):
        if row.get("is_duplicate"): continue

        tipo    = normalizar_tipo_medio(str(row.get(km["tipodemedio"], "")))
        mencion = norm_key(row.get(km["menciones"], ""))
        medio   = norm_key(row.get(km["medio"], ""))

        streaming_url_raw = row.get(km["link_streaming"])
        if isinstance(streaming_url_raw, dict):
            streaming_url_raw = streaming_url_raw.get("url")
            
        if streaming_url_raw and mencion:
            streaming_url_norm = _normalizar_url(str(streaming_url_raw))
            if streaming_url_norm:
                sk = (streaming_url_norm, mencion)
                if sk in seen_streaming:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_streaming[sk]].get(km["idnoticia"], "")
                    continue
                seen_streaming[sk] = i

        if tipo == "Internet":
            li = row.get(km["link_nota"])
            url = li.get("url") if isinstance(li, dict) else li
            if url and mencion:
                url_norm = _normalizar_url(str(url))
                k = (url_norm, mencion)
                if k in seen_url:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_url[k]].get(km["idnoticia"], "")
                    continue
                seen_url[k] = i
            if medio and mencion:
                tb[(medio, mencion)].append(i)

        elif tipo in ("Radio", "Televisión"):
            hora = str(row.get(km["hora"], "")).strip()
            if mencion and medio and hora:
                k = (mencion, medio, hora)
                if k in seen_bcast:
                    row["is_duplicate"] = True
                    row[km["idduplicada"]] = processed[seen_bcast[k]].get(km["idnoticia"], "")
                else:
                    seen_bcast[k] = i

    for idxs in tb.values():
        if len(idxs) < 2: continue
        for i in range(len(idxs)):
            for j in range(i + 1, len(idxs)):
                a, b = idxs[i], idxs[j]
                if processed[a].get("is_duplicate") or processed[b].get("is_duplicate"): continue
                ta  = normalize_title_for_comparison(processed[a].get(km["titulo"]))
                tb_ = normalize_title_for_comparison(processed[b].get(km["titulo"]))
                if ta and tb_ and SequenceMatcher(None, ta, tb_).ratio() >= SIMILARITY_THRESHOLD_TITULOS:
                    if len(ta) < len(tb_):
                        processed[a]["is_duplicate"] = True
                        processed[a][km["idduplicada"]]  = processed[b].get(km["idnoticia"], "")
                    else:
                        processed[b]["is_duplicate"] = True
                        processed[b][km["idduplicada"]]  = processed[a].get(km["idnoticia"], "")

    return processed

def read_and_normalize_dossier(sheet, region_map, internet_map):
    headers = [cell.value for cell in sheet[1] if cell.value is not None]
    rows = []
    for row in sheet.iter_rows(min_row=2):
        if all(c.value is None for c in row):
            continue
        row_data = {}
        for i, h in enumerate(headers):
            if i < len(row):
                cell = row[i]
                val = cell.value
                url = cell.hyperlink.target if (cell.hyperlink and cell.hyperlink.target) else None
                if url:
                    row_data[h] = {"value": val or "Link", "url": url}
                else:
                    row_data[h] = val
        rows.append(row_data)

    df = pd.DataFrame(rows)

    tipo_medio_map = {
        'online': 'Internet', 'internet': 'Internet',
        'diario': 'Prensa',
        'am': 'Radio', 'fm': 'Radio',
        'aire': 'Televisión', 'cable': 'Televisión',
        'revista': 'Revistas', 'revistas': 'Revistas',
    }
    
    if 'Tipo de Medio' in df.columns:
        df['Tipo de Medio'] = (
            df['Tipo de Medio'].astype(str).str.lower().str.strip()
            .map(tipo_medio_map)
            .fillna(df['Tipo de Medio'].astype(str).str.strip())
        )
    else:
        df['Tipo de Medio'] = 'Otro'

    is_av = df['Tipo de Medio'].isin(['Radio', 'Televisión'])
    is_grafica = df['Tipo de Medio'].isin(['Prensa', 'Internet', 'Revistas'])
    is_internet = df['Tipo de Medio'] == 'Internet'

    if 'Medio' in df.columns:
        raw_medios_clean = df['Medio'].astype(str).str.lower().str.strip()
        df['Región'] = raw_medios_clean.map(region_map).fillna("N/A")
    else:
        df['Medio'] = 'N/A'
        df['Región'] = 'N/A'

    if 'Medio' in df.columns:
        df.loc[is_internet, 'Medio'] = (
            df.loc[is_internet, 'Medio']
            .astype(str).str.lower().str.strip()
            .map(internet_map)
            .fillna(df.loc[is_internet, 'Medio'])
        )

    df['ID Noticia'] = df.get('NoticiaId', df.get('ID Noticia', pd.Series(dtype=str)))
    df['Fecha'] = pd.to_datetime(df.get('Fecha', pd.Series(dtype=str)), dayfirst=True, errors='coerce').dt.normalize()
    df['Hora'] = df.get('Hora', pd.Series(dtype=str))
    df['Sección - Programa'] = df.get('Sección - Programa', pd.Series(dtype=str)).astype(str).apply(clean_text)
    
    titulo_col = 'Título' if 'Título' in df.columns else 'Titulo'
    df['Título'] = df.get(titulo_col, pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Autor - Conductor'] = df.get('Autor - Conductor', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Nro. Pagina'] = df.get('Nro. Pagina', pd.Series(dtype=str))
    
    dim_col = 'Dimensioncm2' if 'Dimensioncm2' in df.columns else 'Dimensión'
    df['Dimensión'] = df.get(dim_col, pd.Series(dtype=str))
    df['Duración - Nro. Caracteres'] = df.get('Duración - Nro. Caracteres', pd.Series(dtype=str))

    df.loc[is_av, 'Dimensión'] = df.loc[is_av, 'Duración - Nro. Caracteres']
    df.loc[is_av, 'Duración - Nro. Caracteres'] = 0

    cpe_av = df.get('CPE', pd.Series([np.nan] * len(df)))
    cpe_grafica = df.get('Valor de Nota', pd.Series([np.nan] * len(df)))
    df['CPE'] = np.where(is_av, cpe_av, np.where(is_grafica, cpe_grafica, np.nan))

    df['Tier'] = df.get('Tier', pd.Series(dtype=str))
    df['Audiencia'] = df.get('Audiencia', pd.Series(dtype=str))
    df['Tono'] = df.get('Tono', pd.Series(dtype=str)).astype(str).apply(clean_text)
    df['Tema'] = df.get('Tematica', df.get('Tema', pd.Series(dtype=str))).astype(str).apply(clean_text)
    df['Temas Generales - Tema'] = df.get('Temas Generales - Tema', pd.Series(dtype=str)).astype(str).apply(clean_text)

    cuerpo_col = 'CuerpoEs' if 'CuerpoEs' in df.columns else 'Resumen - Aclaracion'
    cuerpo_cleaned = df.get(cuerpo_col, pd.Series([''] * len(df))).astype(str).apply(clean_cuerpo)

    def fmt_grafica(text):
        if not isinstance(text, str) or not text.strip():
            return text
        parrafos = [p.strip() for p in text.split('\n') if p.strip()]
        return '\n\n'.join(parrafos) if len(parrafos) > 1 else text

    df['Resumen - Aclaracion'] = np.where(is_av, cuerpo_cleaned, cuerpo_cleaned.apply(fmt_grafica))

    # ── ADICIÓN: columna con el CuerpoEs COMPLETO, sin truncar ──────────────
    # Se guarda tal cual queda cuerpo_cleaned (HTML limpio, <br> -> saltos de línea),
    # SIN pasar por corregir_texto() (que es lo que recorta/añade "..." al final).
    df['Cuerpo Completo'] = cuerpo_cleaned

    url_nota_av = df.get('URL Nota AV', df.get('Link Nota AV', pd.Series([''] * len(df))))
    url_streaming = df.get('URL (Streaming - Imagen)', pd.Series([''] * len(df)))
    
    link_nota_final = []
    for val_av, val_str, is_av_row in zip(url_nota_av, url_streaming, is_av):
        if is_av_row:
            if isinstance(val_av, dict):
                url_t = val_av.get("url", "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
            else:
                url_t = str(val_av or "")
                link_nota_final.append({"value": "Link", "url": url_t.replace(".com.ar", ".com.co") if url_t else None})
        else:
            if isinstance(val_str, dict):
                link_nota_final.append(val_str)
            else:
                link_nota_final.append({"value": "Link", "url": val_str if val_str else None})
                
    df['Link Nota'] = link_nota_final

    url_nota_raw = df.get('URL Nota', pd.Series([''] * len(df)))
    link_stream_final = []
    for val_url, is_int in zip(url_nota_raw, is_internet):
        if is_int:
            if isinstance(val_url, dict):
                link_stream_final.append(val_url)
            else:
                link_stream_final.append({"value": "Link", "url": val_url if val_url else None})
        else:
            link_stream_final.append(None)
            
    df['Link (Streaming - Imagen)'] = link_stream_final

    menciones_av = df.get('Menciones - Empresa', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    menciones_grafica = df.get('Empresa rel.', pd.Series([''] * len(df))).fillna('').astype(str).apply(clean_text)
    df['Menciones - Empresa'] = np.where(is_av, menciones_av, np.where(is_grafica, menciones_grafica, menciones_av))

    return df

def generate_output_excel(rows, km):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ORDER = [
        "ID Noticia", "Fecha", "Hora", "Medio", "Tipo de Medio",
        "Sección - Programa", "Región", "Título", "Autor - Conductor",
        "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres",
        "CPE", "Tier", "Audiencia", "Tono", "Tono IA", "Tema", "Subtema",
        "Link Nota", "Resumen - Aclaracion", "Link (Streaming - Imagen)", "Menciones - Empresa",
        "ID duplicada",
        "Cuerpo Completo"   # ── ADICIÓN: columna final con el CuerpoEs completo, sin truncar ──
    ]
    NUM = {"ID Noticia", "Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "CPE", "Tier", "Audiencia"}
    ws.append(ORDER)
    
    font_hyperlink = Font(color="000000", underline=None)
    align_left = Alignment(horizontal='left')
    font_header = Font(bold=True)
    
    for i, col_name in enumerate(ORDER, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = font_header

    col_idx_map = {name: ORDER.index(name) + 1 for name in ORDER}
        
    for row in rows:
        tk = km.get("titulo")
        if tk and tk in row: row[tk] = clean_title_for_output(row.get(tk))
        rk = km.get("resumen")
        if rk and rk in row: row[rk] = corregir_texto(row.get(rk))
        
        out, links = [], {}
        for ci, h in enumerate(ORDER, start=1):
            dk = km.get(norm_key(h), norm_key(h))
            val = row.get(h)
            cv = None
            
            if h == 'Fecha' and pd.notna(val):
                if isinstance(val, pd.Timestamp):
                    cv = val.to_pydatetime()
                elif isinstance(val, (datetime.datetime, datetime.date)):
                    cv = val
                else:
                    cv = str(val) if val is not None else None
            elif h in NUM:
                cv = parse_numeric(val)
            elif isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"): links[ci] = val["url"]
            elif val is not None:
                if isinstance(val, str) and val.startswith("http"):
                    cv = "Link"
                    links[ci] = val
                else:
                    cv = str(val)
            out.append(cv)
        ws.append(out)
        
        current_row = ws.max_row
        for ci, url in links.items():
            cell = ws.cell(row=current_row, column=ci)
            cell.hyperlink = url
            cell.font = font_hyperlink
            cell.alignment = align_left
            
        date_col_idx = ORDER.index("Fecha") + 1
        date_cell = ws.cell(row=current_row, column=date_col_idx)
        if isinstance(date_cell.value, (datetime.datetime, datetime.date)):
            date_cell.number_format = 'DD/MM/YYYY'
            
        cols_millares = ["Nro. Pagina", "Dimensión", "Duración - Nro. Caracteres", "Tier", "Audiencia"]
        for col_name in cols_millares:
            col_idx = col_idx_map[col_name]
            cell = ws.cell(row=current_row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

        cpe_idx = col_idx_map["CPE"]
        cpe_cell = ws.cell(row=current_row, column=cpe_idx)
        if isinstance(cpe_cell.value, (int, float)):
            cpe_cell.number_format = '$#,##0'
            
    for i, col_name in enumerate(ORDER, start=1):
        letter = ws.cell(row=1, column=i).column_letter
        if col_name in ['Título', 'Resumen - Aclaracion', 'Cuerpo Completo']:
            ws.column_dimensions[letter].width = 50
        elif col_name in ['Link Nota', 'Link (Streaming - Imagen)']:
            ws.column_dimensions[letter].width = 15
        else:
            ws.column_dimensions[letter].width = 20
            
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==============================================================================
# ==============================================================================
#   MOTOR DE PRECISIÓN  ·  jerarquía identidad -> asunto -> tema
# ==============================================================================
#  Reemplaza el nucleo analitico (tono, subtema y tema) por una arquitectura de
#  TRES NIVELES calculada UNA sola vez y compartida por las tres salidas:
#
#     NIVEL 0 · IDENTIDAD  "es la misma noticia" (republicacion / teletipo)
#                          -> comparte OBLIGATORIAMENTE tono + tema + subtema
#     NIVEL 1 · ASUNTO     "distintas noticias sobre el mismo hecho"
#                          -> comparte tema + subtema
#     NIVEL 2 · TEMA       "asuntos afines bajo una categoria editorial"
#                          -> comparte tema
#
#  Antes cada salida construia su propia agrupacion (el tono un DSU, el subtema
#  otro, el tema un clustering aparte), de modo que dos noticias podian quedar
#  unidas para el tono y separadas para el subtema. Aqui el grafo se calcula una
#  vez y se propaga.
#
#  Principio rector: PRECISION SOBRE COBERTURA. Ante la duda NO se agrupa y NO
#  se asigna tono distinto de Neutro. Toda union exige coincidencia en dos
#  familias de senales independientes (semantica + lexica/entidades).
#
#  Modelo de clasificacion: gpt-4.1-nano-2025-04-14 (sin cambios).
#  Sin dependencias nuevas: numpy, scikit-learn, unidecode y openai ya estaban.
# ==============================================================================

MODELO_CLASIFICACION_POR_DEFECTO = "gpt-4.1-nano-2025-04-14"


# ==============================================================================
# 1 · CONFIGURACIÓN
# ==============================================================================

@dataclass
class ConfigPrecision:
    """Umbrales del motor. Todos calibrados hacia la precisión.

    Los tres umbrales de similitud coseno (`identidad`, `asunto`, `tema`) operan
    sobre `text-embedding-3-small`, donde dos textos sin relación real rara vez
    bajan de 0.60. Por eso los pisos absolutos son altos: 0.80 en ese espacio NO
    significa "parecido", significa "vagamente del mismo campo semántico".
    """

    # ── Nivel 0 · identidad (misma noticia) ────────────────────────────────
    sim_identidad: float = 0.945
    sim_republicacion: float = 0.975      # copia casi literal: basta con esto
    ratio_titulo_identico: float = 0.93

    # ── Nivel 1 · asunto (subtema) ─────────────────────────────────────────
    sim_asunto: float = 0.880
    piso_absoluto_asunto: float = 0.855   # por debajo NUNCA se agrupa

    # ── Nivel 2 · tema ─────────────────────────────────────────────────────
    sim_tema: float = 0.800
    piso_absoluto_tema: float = 0.760

    # ── Señales léxicas exigidas además del embedding ──────────────────────
    jaccard_entidades_min: float = 0.30   # entidades fuertes compartidas
    overlap_tokens_min: float = 0.42      # tokens distintivos compartidos
    overlap_tokens_min_tema: float = 0.22

    # ── Entidades: filtro IDF ──────────────────────────────────────────────
    # Una entidad presente en más del X% del corpus no discrimina nada
    # ("Colombia", "Gobierno", "Bogotá") y se descarta como señal.
    df_max_entidad_fuerte: float = 0.15
    min_entidades_para_exigir: int = 2

    # ── Tono ───────────────────────────────────────────────────────────────
    confianza_minima_tono: float = 0.70   # por debajo -> segunda pasada
    revisar_grupos_grandes: int = 4       # grupos >= N notas -> siempre revisar

    # ── Etiquetas ──────────────────────────────────────────────────────────
    min_palabras_subtema: int = 3
    max_palabras_subtema: int = 7
    min_palabras_tema: int = 2
    max_palabras_tema: int = 4
    max_notas_por_etiqueta: int = 40
    vecinos_vocabulario: int = 8          # etiquetas previas que ve el modelo
    sim_unificar_etiquetas: float = 0.90  # dos etiquetas se estudian como sinónimas
    # Prueba de especificidad: la etiqueta debe describir SU grupo mejor de lo que
    # describe al corpus entero. Si no gana por este margen, es una generalidad.
    margen_especificidad: float = 0.03

    # ── Recursos ───────────────────────────────────────────────────────────
    max_pares_por_item: int = 30
    n_max_matriz_completa: int = 2500

    def escalar_por_corpus(self, n: int) -> "ConfigPrecision":
        """Endurece los umbrales en corpus pequeños.

        Con 5 noticias no hay evidencia estadística para agrupar nada: cualquier
        fusión errónea se lleva el 20% del informe. Con 500, el clustering tiene
        contexto suficiente y puede relajarse ligeramente.
        """
        c = ConfigPrecision(**self.__dict__)
        if n <= 5:
            c.sim_asunto, c.piso_absoluto_asunto = 0.930, 0.910
            c.sim_tema, c.piso_absoluto_tema = 0.880, 0.850
            c.overlap_tokens_min = 0.55
            c.jaccard_entidades_min = 0.40
        elif n <= 12:
            c.sim_asunto, c.piso_absoluto_asunto = 0.905, 0.880
            c.sim_tema, c.piso_absoluto_tema = 0.845, 0.810
            c.overlap_tokens_min = 0.50
            c.jaccard_entidades_min = 0.35
        elif n <= 30:
            c.sim_asunto, c.piso_absoluto_asunto = 0.892, 0.868
            c.sim_tema, c.piso_absoluto_tema = 0.820, 0.785
            c.overlap_tokens_min = 0.46
        return c


@dataclass
class Telemetria:
    """Contadores de calidad.

    Sin esto no se puede afirmar que el sistema "mejoró": los `except:` mudos
    del código original convertían cualquier fallo de API en un "Neutro" o en un
    subtema genérico, indistinguibles de una clasificación real.
    """

    llamadas_ok: int = 0
    llamadas_json_invalido: int = 0
    llamadas_error: int = 0
    llamadas_rate_limit: int = 0
    tokens_in: int = 0
    tokens_out: int = 0

    tono_sin_mencion: int = 0
    tono_revisado: int = 0
    tono_cambiado_en_revision: int = 0
    tono_forzado_neutro_por_afectado: int = 0
    tono_baja_confianza_final: int = 0
    tono_armonizado_por_asunto: int = 0

    uniones_identidad: int = 0
    uniones_asunto: int = 0
    uniones_rechazadas_entidades: int = 0
    uniones_rechazadas_overlap: int = 0
    uniones_rechazadas_conflicto: int = 0

    etiquetas_llm: int = 0
    etiquetas_fallback: int = 0
    etiquetas_unificadas: int = 0
    etiquetas_reformuladas_genericas: int = 0
    subtemas_reasignados_coherencia: int = 0

    avisos: List[str] = field(default_factory=list)

    def avisar(self, msg: str) -> None:
        if msg not in self.avisos:
            self.avisos.append(msg)

    @property
    def tasa_fallo_llm(self) -> float:
        total = self.llamadas_ok + self.llamadas_error + self.llamadas_json_invalido
        return 0.0 if total == 0 else (self.llamadas_error + self.llamadas_json_invalido) / total

    def resumen(self) -> Dict[str, Any]:
        return {
            "LLM · llamadas correctas": self.llamadas_ok,
            "LLM · JSON inválido": self.llamadas_json_invalido,
            "LLM · errores de API": self.llamadas_error,
            "LLM · rate limits": self.llamadas_rate_limit,
            "LLM · tasa de fallo": f"{self.tasa_fallo_llm * 100:.1f}%",
            "Tono · sin mención de marca": self.tono_sin_mencion,
            "Tono · revisados 2ª pasada": self.tono_revisado,
            "Tono · corregidos en revisión": self.tono_cambiado_en_revision,
            "Tono · forzados a Neutro": self.tono_forzado_neutro_por_afectado,
            "Tono · baja confianza final": self.tono_baja_confianza_final,
            "Tono · unificados por asunto": self.tono_armonizado_por_asunto,
            "Grupos · uniones identidad": self.uniones_identidad,
            "Grupos · uniones asunto": self.uniones_asunto,
            "Grupos · rechazos por entidades": self.uniones_rechazadas_entidades,
            "Grupos · rechazos por overlap": self.uniones_rechazadas_overlap,
            "Grupos · rechazos por acción opuesta": self.uniones_rechazadas_conflicto,
            "Etiquetas · generadas por LLM": self.etiquetas_llm,
            "Etiquetas · por fallback": self.etiquetas_fallback,
            "Etiquetas · unificadas": self.etiquetas_unificadas,
            "Etiquetas · reformuladas por genéricas": self.etiquetas_reformuladas_genericas,
            "Subtemas · reasignados por coherencia": self.subtemas_reasignados_coherencia,
        }


@dataclass
class EngineContext:
    """Dependencias externas inyectadas (evita acoplar el motor a Streamlit)."""

    embed: Callable[[List[str]], List[Optional[List[float]]]]
    modelo: str = MODELO_CLASIFICACION_POR_DEFECTO
    on_tokens: Optional[Callable[[int, int], None]] = None
    # 50 concurrentes contra un modelo nano provoca 429s; cada 429 agotado se
    # convertía en un "Neutro" silencioso. 24 sostiene el throughput sin ese sesgo.
    max_concurrencia: int = 24
    seed: int = 7
    tel: Telemetria = field(default_factory=Telemetria)


@dataclass
class ResultadoAnalisis:
    tonos: List[str]
    confianza_tono: List[float]
    subtemas: List[str]
    temas: List[str]
    id_identidad: List[int]
    id_asunto: List[int]
    telemetria: Telemetria


# ==============================================================================
# 2 · NORMALIZACIÓN LÉXICA
# ==============================================================================

_PE_STOPWORDS = set(
    """
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante para
por segun sin so sobre tras y o u e la el los las un una unos unas lo al del se
su sus le les mi mis tu tus nuestro nuestros vuestra vuestras este esta estos
estas ese esa esos esas aquel aquella aquellos aquellas que cual cuales quien
quienes cuyo cuya cuyos cuyas como cuando donde es son fue fueron era eran sera
seran seria serian he ha han habia hay hubo habra habria estoy estan estaba
estaban estamos estar estare estaria estuvieron estarian estuvo asi ya mas menos
tan tanto cada muy todo toda todos todas ser haber hacer tener poder deber ir
dar ver saber querer llegar pasar encontrar creer decir poner salir volver
seguir llevar sentir cambiar tambien pero aunque mientras segun ademas
""".split()
)

# Términos que aparecen en cualquier noticia y por tanto no distinguen una de
# otra. Sin este filtro, dos notas sin nada en común comparten "empresa",
# "gobierno" y "sector" y superan el test de overlap.
TOKENS_DEBILES = _PE_STOPWORDS | {
    "noticia", "noticias", "informe", "informacion", "comunicado", "anuncio",
    "colombia", "pais", "nacional", "regional", "local", "sector", "sectores",
    "empresa", "empresas", "entidad", "entidades", "autoridad", "autoridades",
    "gobierno", "alcaldia", "gobernacion", "ministerio", "nuevo", "nueva",
    "nuevos", "nuevas", "plan", "programa", "proyecto", "iniciativa",
    "actividad", "actividades", "gestion", "tema", "caso", "casos", "millones",
    "personas", "durante", "medio", "medios", "mercado", "grupo", "cifras",
    "aumento", "informo", "explico", "agrego", "senalo", "afirmo", "dijo",
}

PALABRAS_CORTE_ETIQUETA = {
    "de", "del", "la", "el", "los", "las", "un", "una", "unos", "unas", "al",
    "su", "sus", "en", "con", "sin", "por", "para", "sobre", "ante", "bajo",
    "contra", "desde", "entre", "hacia", "hasta", "mediante", "tras", "y", "o",
    "u", "e", "lo", "que", "se", "como", "donde", "cuando", "cual", "cuyo",
    "cuya", "cuyos", "cuyas", "este", "esta", "estos", "estas", "ese", "esa",
    "esos", "esas", "aquel", "aquella", "aquellos", "aquellas", "cada", "todo",
    "toda", "todos", "todas", "otro", "otra", "otros", "otras", "nuevo",
    "nueva", "nuevos", "nuevas", "gran", "grandes", "mayor", "mayores",
    "menor", "menores", "mejor", "mejores", "peor", "peores", "primer",
    "primera", "segundo", "segunda", "tercer", "tercera", "mas", "muy", "tan",
    "tanto", "tanta", "tantos", "tantas", "mi", "mis", "tu", "tus", "nuestro",
    "nuestra", "nuestros", "nuestras", "a", "ha", "he", "ser", "estar",
    "haber", "hacer", "tener", "poder", "deber", "ir", "dar", "ver", "saber",
}

NEXOS_VALIDOS = {
    "de", "del", "para", "sobre", "en", "con", "por", "ante", "hacia", "entre",
    "sin", "al", "las", "los", "una", "uno", "la", "el", "y", "o", "a", "e", "u",
    "contra", "desde", "hasta", "tras", "mediante",
}

RE_VERBO_CONJUGADO = re.compile(
    r"\b(presenta|presentan|anuncia|anuncian|lanza|lanzan|inaugura|inauguran|"
    r"realiza|realizan|desarrolla|desarrollan|ejecuta|ejecutan|gestiona|gestionan|"
    r"impulsa|impulsan|promueve|promueven|lidera|lideran|encabeza|encabezan|"
    r"aprueba|aprueban|firma|firman|suscribe|suscriben|invierte|invierten|"
    r"construye|construyen|instala|instalan|entrega|entregan|recibe|reciben|"
    r"solicita|solicitan|visita|visitan|atiende|atienden|destaca|destacan|"
    r"senala|senalan|indica|indican|expresa|expresan|afirma|afirman|"
    r"propone|proponen|pide|piden|exige|exigen|apoya|apoyan|abre|abren|"
    r"informa|informan|reporta|reportan|advierte|advierten|confirma|confirman)\b",
    re.IGNORECASE,
)

ETIQUETAS_GENERICAS = {
    "gestion", "actividades", "acciones", "noticias", "informacion", "eventos",
    "varios", "sin tema", "actividad corporativa", "gestion corporativa",
    "economia", "politica", "tecnologia", "seguridad", "justicia", "actualidad",
    "nacional", "internacional", "empresas", "sociedad", "negocios", "general",
    "cobertura informativa general", "otros", "otras noticias",
    # Frases nominales bien formadas pero que no dicen nada del hecho concreto:
    # pasaban el validador de estructura y acababan agrupando lo incomparable.
    "gestion institucional", "gestion administrativa", "gestion publica",
    "gestion empresarial", "actividad institucional", "agenda institucional",
    "temas de actualidad", "temas de interes", "asuntos de interes general",
    "informacion de interes", "informacion general", "cobertura de medios",
    "noticias del sector", "novedades del sector", "panorama del sector",
    "situacion actual del pais", "coyuntura nacional", "contexto economico",
    "desarrollo economico y social", "desarrollo del pais", "avances del sector",
    "impacto en la comunidad", "opinion publica", "declaraciones de autoridades",
    "anuncios oficiales", "comunicados de prensa", "resumen de noticias",
    "hechos destacados", "sucesos de la region", "eventos del sector",
}

# Pares de acciones que impiden fusionar dos grupos aunque el embedding los
# considere casi idénticos: "aprobación de la reforma" y "rechazo de la reforma"
# viven en el mismo campo semántico pero son noticias opuestas.
ACCIONES_OPUESTAS: List[Tuple[set, set]] = [
    ({"aprobacion", "aprueba", "aprobado", "apoyo", "acuerdo", "aval", "respaldo", "avala"},
     {"rechazo", "rechaza", "rechazado", "desacuerdo", "oposicion", "critica", "niega", "veto"}),
    ({"aumento", "crecimiento", "alza", "subida", "incremento", "sube", "crece", "record"},
     {"caida", "reduccion", "baja", "disminucion", "descenso", "cae", "desploma", "recorte"}),
    ({"apertura", "inauguracion", "inicio", "lanzamiento", "estreno", "abre", "reabre"},
     {"cierre", "suspension", "clausura", "cancelacion", "cierra", "suspende", "liquidacion"}),
    ({"exito", "logro", "triunfo", "premio", "reconocimiento", "gana", "ganador"},
     {"fracaso", "derrota", "problema", "crisis", "sancion", "pierde", "escandalo"}),
    ({"demanda", "denuncia", "investigacion", "sancion", "multa", "condena", "imputacion"},
     {"absolucion", "archivo", "exoneracion", "sobreseimiento", "inocente"}),
    ({"contratacion", "empleo", "vinculacion", "contrata"},
     {"despido", "despidos", "desvinculacion", "recorte", "liquidacion"}),
    ({"ganancias", "utilidades", "superavit", "rentabilidad"},
     {"perdidas", "deficit", "quiebra", "insolvencia"}),
]


def normalizar_texto(s: Any) -> str:
    """Minúsculas sin tildes ni puntuación, con stopwords eliminadas."""
    if not s:
        return ""
    s = unidecode(str(s).lower())
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    return " ".join(t for t in s.split() if t not in _PE_STOPWORDS)


def normalizar_titular(t: Any) -> str:
    """Quita la firma del medio y el antetítulo para comparar titulares.

    Los medios publican el mismo teletipo como "Titular real | El Tiempo" o
    "ÚLTIMA HORA: titular real"; sin esta limpieza, dos copias exactas de la
    misma nota puntúan bajo en similitud de cadena.
    """
    if not isinstance(t, str):
        return ""
    limpio = re.sub(r"\s+[\|–—-]\s+[^\|–—-]+$", "", t).strip()
    if ":" in limpio:
        partes = limpio.split(":", 1)
        sufijo = partes[1].strip()
        if len(sufijo) >= 10:
            limpio = sufijo
    return re.sub(r"\W+", " ", limpio).lower().strip()


def tokens_distintivos(texto: str, min_len: int = 4) -> set:
    return {
        t
        for t in normalizar_texto(texto).split()
        if len(t) >= min_len and t not in TOKENS_DEBILES and not t.isdigit()
    }


def overlap_distintivo(a: str, b: str) -> float:
    """Solapamiento sobre el conjunto más pequeño (contención, no Jaccard).

    Un titular breve contenido en un cuerpo largo debe puntuar alto; Jaccard lo
    penalizaría por la diferencia de tamaño.
    """
    ta, tb = tokens_distintivos(a), tokens_distintivos(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / max(1, min(len(ta), len(tb)))


def hay_conflicto_accion(a: str, b: str) -> bool:
    ta = tokens_distintivos(a, min_len=3)
    tb = tokens_distintivos(b, min_len=3)
    for grupo_a, grupo_b in ACCIONES_OPUESTAS:
        if (ta & grupo_a and tb & grupo_b) or (ta & grupo_b and tb & grupo_a):
            return True
    return False


def jaccard(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


# ==============================================================================
# 3 · ENTIDADES  (la señal que faltaba)
# ==============================================================================

RE_SIGLA = re.compile(r"\b[A-ZÁÉÍÓÚÑ]{2,7}\b")
RE_PROPIO = re.compile(
    r"\b[A-ZÁÉÍÓÚÑ][a-záéíóúñ]{2,}"
    r"(?:\s+(?:de|del|la|los|las|y)\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]{2,})*"
)
RE_CIFRA = re.compile(
    r"\b\d[\d.,]{1,}\s*(?:%|millones|millón|millon|mil|billones|"
    r"pesos|dólares|dolares|usd|cop|km|kw|mw|toneladas)?",
    re.IGNORECASE,
)
RE_FECHA_ANO = re.compile(r"\b(?:19|20)\d{2}\b")

# Palabras que abren frase o son cargos y aparecen capitalizadas por gramática,
# no por ser entidades.
NO_ENTIDADES = {
    "el", "la", "los", "las", "un", "una", "este", "esta", "esto", "ese", "esa",
    "por", "para", "con", "sin", "sobre", "entre", "desde", "hasta", "durante",
    "segun", "ademas", "tras", "ante", "pero", "aunque", "mientras", "cuando",
    "sin", "asi", "tambien", "aun", "todo", "toda", "cada", "otro", "otra",
    "alcalde", "alcaldesa", "gobernador", "gobernadora", "ministro", "ministra",
    "presidente", "presidenta", "director", "directora", "gerente", "senador",
    "senadora", "representante", "concejal", "secretario", "secretaria",
    "lunes", "martes", "miercoles", "jueves", "viernes", "sabado", "domingo",
    "enero", "febrero", "marzo", "abril", "mayo", "junio", "julio", "agosto",
    "septiembre", "octubre", "noviembre", "diciembre",
    "nuevo", "nueva", "gran", "primer", "primera", "segundo", "ultimo",
}

SIGLAS_VACIAS = {"EL", "LA", "LOS", "DE", "DEL", "EN", "UN", "SE", "NO", "SI", "ES", "AL", "POR"}


def _es_title_case(texto: str) -> bool:
    """Detecta titulares escritos con Todas Las Palabras En Mayúscula.

    En ese formato la mayúscula deja de señalar nombre propio y la extracción
    debe ignorar el titular para no inventar entidades en cada palabra.
    """
    palabras = [p for p in texto.split() if len(p) > 2 and p[0].isalpha()]
    if len(palabras) < 4:
        return False
    con_mayus = sum(1 for p in palabras if p[0].isupper())
    return con_mayus / len(palabras) >= 0.65


def extraer_entidades(titulo: str, cuerpo: str = "") -> set:
    """Nombres propios, siglas y cifras significativas de una noticia.

    Sin dependencias de NLP pesadas (spaCy no está en requirements y añadiría
    ~500 MB al contenedor de Streamlit). La precisión de esta heurística es
    suficiente porque no se usa para etiquetar, solo para *vetar* fusiones.
    """
    titulo = str(titulo or "")
    cuerpo = str(cuerpo or "")
    ents: set = set()

    fuentes = []
    if titulo and not _es_title_case(titulo):
        fuentes.append(titulo)
    fuentes.append(cuerpo[:2500])

    for fuente in fuentes:
        if not fuente:
            continue
        for m in RE_PROPIO.finditer(fuente):
            bruto = m.group(0).strip()
            norm = unidecode(bruto.lower())
            cabeza = norm.split()[0] if norm.split() else ""
            if cabeza in NO_ENTIDADES or len(cabeza) < 4:
                continue
            ents.add(norm)
        for m in RE_SIGLA.finditer(fuente):
            s = m.group(0)
            if s in SIGLAS_VACIAS or len(s) < 3:
                continue
            ents.add(unidecode(s.lower()))

    texto_cifras = f"{titulo} {cuerpo[:1500]}"
    for m in RE_CIFRA.finditer(texto_cifras):
        bruto = m.group(0).strip().lower()
        digitos = re.sub(r"\D", "", bruto)
        # Una cifra es identificadora si es grande o lleva unidad; "3 personas"
        # no distingue una noticia de otra.
        if len(digitos) >= 4 or re.search(r"[a-z%]", bruto):
            if not RE_FECHA_ANO.fullmatch(bruto):
                ents.add(re.sub(r"\s+", "", unidecode(bruto)))

    return ents


class IndiceEntidades:
    """Índice de entidades con filtro IDF.

    Una entidad que aparece en la mayoría del corpus ("Colombia" en un dossier
    colombiano) tiene poder discriminante nulo. Solo las entidades por debajo del
    umbral de frecuencia documental se consideran "fuertes" y pueden vetar o
    avalar una fusión.
    """

    def __init__(self, entidades_por_doc: List[set], cfg: ConfigPrecision):
        self.n = len(entidades_por_doc)
        self.crudas = entidades_por_doc
        df: Counter = Counter()
        for ents in entidades_por_doc:
            for e in ents:
                df[e] += 1
        tope = max(2, int(self.n * cfg.df_max_entidad_fuerte))
        self.df = df
        self.fuertes: List[set] = [{e for e in ents if df[e] <= tope} for ents in entidades_por_doc]
        self.tope_df = tope

    def indice_invertido(self) -> Dict[str, List[int]]:
        """Entidad -> documentos. Genera pares candidatos sin comparar n×n."""
        inv: Dict[str, List[int]] = defaultdict(list)
        for i, ents in enumerate(self.fuertes):
            for e in ents:
                inv[e].append(i)
        return inv


# ==============================================================================
# 4 · DECISIÓN DE FUSIÓN  (el corazón de "no agrupar generalizadamente")
# ==============================================================================

class DSUPrecision:
    def __init__(self, n: int):
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i: int) -> int:
        raiz = i
        while self.p[raiz] != raiz:
            raiz = self.p[raiz]
        while self.p[i] != raiz:
            self.p[i], i = raiz, self.p[i]
        return raiz

    def union(self, i: int, j: int) -> bool:
        ri, rj = self.find(i), self.find(j)
        if ri == rj:
            return False
        if self.rank[ri] < self.rank[rj]:
            ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]:
            self.rank[ri] += 1
        return True

    def grupos(self) -> Dict[int, List[int]]:
        c: Dict[int, List[int]] = defaultdict(list)
        for i in range(len(self.p)):
            c[self.find(i)].append(i)
        return dict(c)

    def etiquetas(self) -> List[int]:
        raices = {}
        salida = []
        for i in range(len(self.p)):
            r = self.find(i)
            if r not in raices:
                raices[r] = len(raices)
            salida.append(raices[r])
        return salida


def decidir_union(
    sim: float,
    i: int,
    j: int,
    textos: Sequence[str],
    idx_ents: IndiceEntidades,
    cfg: ConfigPrecision,
    tel: Telemetria,
    *,
    nivel: str = "asunto",
) -> bool:
    """¿Son `i` y `j` la misma noticia / el mismo asunto?

    Exige DOS familias de señales independientes. El embedding por sí solo agrupa
    "subida del precio del café" con "caída del precio del cacao": misma forma,
    hecho distinto. La regla de veto por entidades es la que evita esas fusiones.
    """
    if nivel == "identidad":
        piso, umbral = cfg.sim_identidad, cfg.sim_identidad
    elif nivel == "tema":
        piso, umbral = cfg.piso_absoluto_tema, cfg.sim_tema
    else:
        piso, umbral = cfg.piso_absoluto_asunto, cfg.sim_asunto

    if sim < piso:
        return False

    ta, tb = textos[i], textos[j]

    if hay_conflicto_accion(ta, tb):
        tel.uniones_rechazadas_conflicto += 1
        return False

    # Republicación literal: la similitud es tan alta que no cabe otra lectura.
    if sim >= cfg.sim_republicacion:
        return True

    ea, eb = idx_ents.fuertes[i], idx_ents.fuertes[j]
    jac_ent = jaccard(ea, eb)
    ov = overlap_distintivo(ta, tb)

    # VETO POR ENTIDADES: ambas noticias identifican actores/cifras concretos y
    # no comparten ninguno -> hablan de hechos distintos, por alto que sea `sim`.
    if nivel != "tema":
        if (
            len(ea) >= cfg.min_entidades_para_exigir
            and len(eb) >= cfg.min_entidades_para_exigir
            and jac_ent == 0.0
            and ov < 0.60
        ):
            tel.uniones_rechazadas_entidades += 1
            return False

    if sim < umbral:
        return False

    min_ov = cfg.overlap_tokens_min_tema if nivel == "tema" else cfg.overlap_tokens_min
    if jac_ent >= cfg.jaccard_entidades_min or ov >= min_ov:
        return True

    tel.uniones_rechazadas_overlap += 1
    return False


def _pares_candidatos(
    embs: np.ndarray,
    validos: List[int],
    cfg: ConfigPrecision,
    piso: float,
) -> List[Tuple[float, int, int]]:
    """Pares por encima del piso, limitados a los k vecinos más próximos.

    Sin el tope por ítem, un corpus temáticamente homogéneo genera O(n²) pares y
    el coste de validarlos domina el tiempo total.
    """
    if len(validos) < 2:
        return []

    sim = cosine_similarity(embs)
    np.fill_diagonal(sim, -1.0)
    k = min(cfg.max_pares_por_item, len(validos) - 1)
    pares: Dict[Tuple[int, int], float] = {}

    vecinos = np.argpartition(-sim, kth=k - 1, axis=1)[:, :k] if k >= 1 else np.empty((len(validos), 0), int)
    for a in range(len(validos)):
        for b in vecinos[a]:
            s = float(sim[a][b])
            if s < piso:
                continue
            ia, ib = validos[a], validos[int(b)]
            clave = (ia, ib) if ia < ib else (ib, ia)
            pares[clave] = s

    return sorted(((s, i, j) for (i, j), s in pares.items()), reverse=True)


# ==============================================================================
# 5 · NIVEL 0 y 1 · AGRUPACIÓN
# ==============================================================================

def construir_identidad(
    titulos: Sequence[str],
    resumenes: Sequence[str],
    textos: Sequence[str],
    embs: List[Optional[List[float]]],
    idx_ents: IndiceEntidades,
    cfg: ConfigPrecision,
    tel: Telemetria,
) -> List[int]:
    """NIVEL 0 · agrupa republicaciones de la misma noticia."""
    n = len(titulos)
    dsu = DSUPrecision(n)

    # a) Titulares idénticos tras normalizar (mismo teletipo en varios medios).
    por_hash: Dict[str, List[int]] = defaultdict(list)
    for i, t in enumerate(titulos):
        norm = normalizar_titular(t)
        if len(norm) >= 15:
            por_hash[hashlib.md5(norm.encode()).hexdigest()].append(i)
    for idxs in por_hash.values():
        for j in idxs[1:]:
            if dsu.union(idxs[0], j):
                tel.uniones_identidad += 1

    # b) Cuerpos idénticos.
    por_cuerpo: Dict[str, List[int]] = defaultdict(list)
    for i, r in enumerate(resumenes):
        norm = " ".join(normalizar_texto(r).split()[:120])
        if len(norm.split()) >= 25:
            por_cuerpo[hashlib.md5(norm.encode()).hexdigest()].append(i)
    for idxs in por_cuerpo.values():
        for j in idxs[1:]:
            if dsu.union(idxs[0], j):
                tel.uniones_identidad += 1

    # c) Titulares casi idénticos + validación semántica.
    norms = [normalizar_titular(t) for t in titulos]
    validos = [i for i in range(n) if embs[i] is not None]
    if len(validos) >= 2:
        M = np.array([embs[i] for i in validos])
        for sim, i, j in _pares_candidatos(M, validos, cfg, cfg.sim_identidad):
            if dsu.find(i) == dsu.find(j):
                continue
            ratio = SequenceMatcher(None, norms[i], norms[j]).ratio() if norms[i] and norms[j] else 0.0
            if ratio >= cfg.ratio_titulo_identico or sim >= cfg.sim_republicacion:
                if decidir_union(sim, i, j, textos, idx_ents, cfg, tel, nivel="identidad"):
                    if dsu.union(i, j):
                        tel.uniones_identidad += 1

    return dsu.etiquetas()


def construir_asunto(
    textos: Sequence[str],
    embs: List[Optional[List[float]]],
    id_identidad: List[int],
    idx_ents: IndiceEntidades,
    cfg: ConfigPrecision,
    tel: Telemetria,
) -> List[int]:
    """NIVEL 1 · agrupa noticias distintas sobre el mismo hecho.

    Opera sobre los centroides de nivel 0, no sobre las notas sueltas: promediar
    las republicaciones cancela el ruido de cada redacción y deja la señal del
    hecho, además de reducir el número de comparaciones.
    """
    n = len(textos)
    clusters_id: Dict[int, List[int]] = defaultdict(list)
    for i, c in enumerate(id_identidad):
        clusters_id[c].append(i)

    ids = sorted(clusters_id.keys())
    centroides: Dict[int, np.ndarray] = {}
    representantes: Dict[int, int] = {}
    for c in ids:
        vecs = [embs[i] for i in clusters_id[c] if embs[i] is not None]
        if vecs:
            centroides[c] = np.mean(vecs, axis=0)
            representantes[c] = clusters_id[c][0]

    validos = [c for c in ids if c in centroides]
    dsu = DSUPrecision(len(ids))
    pos = {c: k for k, c in enumerate(ids)}

    if len(validos) >= 2:
        M = np.array([centroides[c] for c in validos])
        sim = cosine_similarity(M)
        np.fill_diagonal(sim, 0.0)

        # `complete` linkage: TODOS los pares del cluster superan el umbral. Con
        # `average` (el original) basta con que la media lo supere, lo que
        # permite encadenamientos A-B-C donde A y C no se parecen en nada.
        dist = np.clip(1.0 - sim, 0.0, 2.0)
        np.fill_diagonal(dist, 0.0)
        etiquetas = AgglomerativeClustering(
            n_clusters=None,
            distance_threshold=1.0 - cfg.sim_asunto,
            metric="precomputed",
            linkage="complete",
        ).fit(dist).labels_

        provisional: Dict[int, List[int]] = defaultdict(list)
        for k, lbl in enumerate(etiquetas):
            provisional[lbl].append(k)

        # El clustering solo propone; la unión se confirma par a par con las
        # señales léxicas y de entidades.
        for miembros in provisional.values():
            if len(miembros) < 2:
                continue
            for a in range(len(miembros)):
                for b in range(a + 1, len(miembros)):
                    ka, kb = miembros[a], miembros[b]
                    ca, cb = validos[ka], validos[kb]
                    ia, ib = representantes[ca], representantes[cb]
                    if decidir_union(
                        float(sim[ka][kb]), ia, ib, textos, idx_ents, cfg, tel, nivel="asunto"
                    ):
                        if dsu.union(pos[ca], pos[cb]):
                            tel.uniones_asunto += 1

    mapa_cluster = dsu.etiquetas()
    salida = [0] * n
    for c in ids:
        for i in clusters_id[c]:
            salida[i] = mapa_cluster[pos[c]]
    return salida


# ==============================================================================
# 6 · LLAMADAS AL MODELO
# ==============================================================================

def _extraer_json(bruto: str) -> Optional[dict]:
    if not bruto:
        return None
    bruto = bruto.strip()
    if bruto.startswith("```"):
        bruto = re.sub(r"^```(?:json)?\s*|\s*```$", "", bruto, flags=re.IGNORECASE)
    try:
        obj = json.loads(bruto)
        return obj if isinstance(obj, dict) else None
    except Exception:
        pass
    m = re.search(r"\{.*\}", bruto, re.DOTALL)
    if m:
        try:
            obj = json.loads(m.group(0))
            return obj if isinstance(obj, dict) else None
        except Exception:
            return None
    return None


def _registrar_uso(resp: Any, ctx: EngineContext) -> None:
    u = resp.get("usage", {}) if isinstance(resp, dict) else getattr(resp, "usage", None)
    if not u:
        return
    pin = (u.get("prompt_tokens") if isinstance(u, dict) else getattr(u, "prompt_tokens", 0)) or 0
    pout = (u.get("completion_tokens") if isinstance(u, dict) else getattr(u, "completion_tokens", 0)) or 0
    ctx.tel.tokens_in += pin
    ctx.tel.tokens_out += pout
    if ctx.on_tokens:
        ctx.on_tokens(pin, pout)


async def chat_json_async(
    sistema: str,
    usuario: str,
    ctx: EngineContext,
    *,
    max_tokens: int = 160,
    temperature: float = 0.0,
    intentos: int = 4,
) -> Optional[dict]:
    """Llamada JSON con backoff exponencial + jitter y telemetría por tipo de fallo.

    A diferencia del `except: pass` original, distingue error de API de respuesta
    mal formada: solo así se puede saber si un informe lleno de "Neutro" refleja
    la realidad o una tanda de 429s.
    """
    espera = 1.0
    for intento in range(intentos):
        try:
            resp = await openai.ChatCompletion.acreate(
                model=ctx.modelo,
                messages=[
                    {"role": "system", "content": sistema},
                    {"role": "user", "content": usuario},
                ],
                max_tokens=max_tokens,
                temperature=temperature,
                seed=ctx.seed,
                response_format={"type": "json_object"},
            )
            _registrar_uso(resp, ctx)
            datos = _extraer_json(resp.choices[0].message.content)
            if datos is None:
                ctx.tel.llamadas_json_invalido += 1
                return None
            ctx.tel.llamadas_ok += 1
            return datos
        except _ERROR_RATE_LIMIT:
            ctx.tel.llamadas_rate_limit += 1
            if intento == intentos - 1:
                ctx.tel.llamadas_error += 1
                return None
            await asyncio.sleep(espera + random.uniform(0, espera * 0.5))
            espera *= 2
        except _ERRORES_REINTENTABLES:
            if intento == intentos - 1:
                ctx.tel.llamadas_error += 1
                return None
            await asyncio.sleep(espera + random.uniform(0, espera * 0.5))
            espera *= 2
        except Exception:
            ctx.tel.llamadas_error += 1
            return None
    ctx.tel.llamadas_error += 1
    return None


async def _mapear(coros: List[Any], limite: int) -> List[Any]:
    sem = asyncio.Semaphore(limite)

    async def _envuelto(c):
        async with sem:
            return await c

    return await asyncio.gather(*[_envuelto(c) for c in coros])


# ==============================================================================
# 7 · TONO
# ==============================================================================

SISTEMA_TONO = (
    "Eres analista senior de reputación corporativa. Evalúas el impacto "
    "reputacional DIRECTO de una noticia sobre una marca concreta, no el clima "
    "emocional del texto. Respondes únicamente con JSON válido."
)


def texto_para_llm(titulo: Any, cuerpo: Any, max_chars: int = 2400) -> str:
    """Texto para el prompt: titular una vez + cuerpo.

    El pipeline original enviaba al clasificador el mismo string usado para
    embeddings, que repite el titular TRES veces ("T. T. T. cuerpo") para
    ponderarlo en el vector. Esa repetición es correcta para el embedding y
    dañina para el prompt: consumía ~2 de cada 3 caracteres del titular y
    truncaba el cuerpo, que es justamente donde está el hecho reputacional.
    """
    t = str(titulo or "").strip()
    c = str(cuerpo or "").strip()
    if c.lower() in ("nan", "none"):
        c = ""
    disponible = max(200, max_chars - len(t) - 24)
    return f"TITULAR: {t}\nCUERPO: {c[:disponible]}".strip()


def _pe_texto_para_embedding(titulo: Any, resumen: Any, max_len: int = 1800) -> str:
    """Se mantiene la triplicación: aquí sí pondera el titular en el vector."""
    t = str(titulo or "").strip()
    r = str(resumen or "").strip()
    if r.lower() in ("nan", "none"):
        r = ""
    return f"{t}. {t}. {t}. {r}"[:max_len]


def _pe_normalizar_mencion(texto: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", unidecode(str(texto).lower()))).strip()


class DetectorMarca:
    """Detecta la marca con límites de palabra y mide su prominencia.

    Que la marca esté en el titular o aparezca cinco veces frente a una sola
    mención de pasada cambia por completo la lectura reputacional, así que esa
    señal se le entrega explícitamente al modelo.
    """

    def __init__(self, marca: str, aliases: Sequence[str]):
        self.marca = (marca or "").strip()
        self.aliases = [a.strip() for a in (aliases or []) if a and a.strip()]
        self.nombres = [n for n in [self.marca] + self.aliases if len(n) >= 3]
        self._patrones = [
            re.compile(rf"(?<![a-z0-9]){re.escape(_pe_normalizar_mencion(n))}(?![a-z0-9])")
            for n in self.nombres
            if len(_pe_normalizar_mencion(n)) >= 3
        ]

    def contar(self, texto: str) -> int:
        t = _pe_normalizar_mencion(texto)
        return sum(len(p.findall(t)) for p in self._patrones)

    def menciona(self, texto: str) -> bool:
        return self.contar(texto) > 0

    def en_titular(self, titulo: str) -> bool:
        return self.contar(titulo) > 0


def _prompt_tono(marca: str, aliases: Sequence[str], texto: str, en_titular: bool, n_menciones: int) -> str:
    alias_str = f"\nTAMBIÉN LLAMADA: {', '.join(aliases)}" if aliases else ""
    return (
        f"MARCA EVALUADA: {marca}{alias_str}\n"
        f"SEÑALES: aparece en el titular: {'SÍ' if en_titular else 'NO'} · "
        f"menciones en el texto: {n_menciones}\n\n"
        f"NOTICIA\n{texto}\n\n"
        "Analiza en este orden:\n"
        "1. hecho: el hecho principal que involucra a la marca (máx. 15 palabras).\n"
        "2. afectado: quién recibe el efecto de ese hecho.\n"
        "   \"marca\" | \"tercero\" | \"sector\" | \"nadie\"\n"
        "3. tono: efecto de ese hecho sobre la reputación de la marca.\n"
        f"   Negativo -> el hecho perjudica, cuestiona o expone a {marca}: demanda, multa,\n"
        "   fraude, falla propia, queja de clientes, investigación, pérdidas, retiro de\n"
        "   producto, despidos, incumplimiento, accidente en sus instalaciones.\n"
        f"   Positivo -> el hecho acredita un logro verificable de {marca}: premio,\n"
        "   crecimiento, inversión ejecutada, lanzamiento exitoso, innovación, expansión,\n"
        "   reconocimiento, alianza favorable, aporte social propio.\n"
        "   Neutro -> la marca aparece sin efecto sobre su imagen: mención en una lista,\n"
        "   dato de mercado, patrocinio menor, comunicado rutinario, contexto sectorial,\n"
        "   o el afectado es un tercero.\n"
        "4. confianza: 0.0 a 1.0.\n\n"
        "REGLAS\n"
        "- Si el afectado no es la marca, el tono es Neutro.\n"
        "- No infieras el tono por palabras emocionales ni por la gravedad general.\n"
        "- Una tragedia, crisis o polémica ajena a la marca es Neutro.\n"
        "- Que la noticia sea sobre el sector no la hace positiva ni negativa para la marca.\n"
        "- Ante duda razonable: Neutro con confianza baja.\n\n"
        'Responde JSON: {"hecho":"...","afectado":"marca|tercero|sector|nadie",'
        '"tono":"Positivo|Negativo|Neutro","confianza":0.0}'
    )


def _prompt_revision_tono(marca: str, texto: str, previo: dict) -> str:
    return (
        f"MARCA EVALUADA: {marca}\n\n"
        f"NOTICIA\n{texto}\n\n"
        f"Un primer analista concluyó: tono={previo.get('tono', 'Neutro')}, "
        f"hecho=\"{previo.get('hecho', '')}\", afectado={previo.get('afectado', 'nadie')}, "
        f"pero con baja confianza.\n\n"
        "Revísalo de forma crítica y decide el tono definitivo. Pregúntate:\n"
        f"- ¿El hecho recae sobre {marca} o sobre un tercero?\n"
        f"- ¿Un directivo de {marca} celebraría o lamentaría esta publicación, "
        "o le sería indiferente?\n"
        "- ¿La clasificación anterior confundió el clima emocional del texto con el "
        "impacto real sobre la marca?\n\n"
        "Si tras la revisión sigue habiendo duda razonable, responde Neutro.\n\n"
        'Responde JSON: {"motivo":"máx 20 palabras","tono":"Positivo|Negativo|Neutro","confianza":0.0}'
    )


class MotorTono:
    def __init__(self, marca: str, aliases: Sequence[str], ctx: EngineContext, cfg: ConfigPrecision):
        self.detector = DetectorMarca(marca, aliases)
        self.marca = marca
        self.aliases = list(aliases or [])
        self.ctx = ctx
        self.cfg = cfg

    async def _clasificar(self, titulo: str, cuerpo: str, tam_grupo: int) -> Tuple[str, float]:
        texto_completo = f"{titulo} {cuerpo}"
        n_men = self.detector.contar(texto_completo)

        # Sin mención literal de la marca no hay impacto reputacional que medir.
        # Se decide en local: ahorra la llamada y elimina el falso positivo por
        # "la noticia va de banca, luego afecta al banco".
        if n_men == 0:
            self.ctx.tel.tono_sin_mencion += 1
            return "Neutro", 1.0

        texto = texto_para_llm(titulo, cuerpo)
        datos = await chat_json_async(
            SISTEMA_TONO,
            _prompt_tono(self.marca, self.aliases, texto, self.detector.en_titular(titulo), n_men),
            self.ctx,
            max_tokens=180,
        )
        if not datos:
            self.ctx.tel.tono_baja_confianza_final += 1
            return "Neutro", 0.0

        tono = str(datos.get("tono", "Neutro")).strip().capitalize()
        if tono not in ("Positivo", "Negativo", "Neutro"):
            tono = "Neutro"
        afectado = str(datos.get("afectado", "")).strip().lower()
        try:
            confianza = float(datos.get("confianza", 0.5))
        except (TypeError, ValueError):
            confianza = 0.5
        confianza = min(max(confianza, 0.0), 1.0)

        # Red de seguridad: el modelo a veces marca "afectado: sector" y aun así
        # devuelve Negativo. La regla se aplica en código, no solo en el prompt.
        if afectado in ("tercero", "sector", "nadie") and tono != "Neutro":
            self.ctx.tel.tono_forzado_neutro_por_afectado += 1
            tono, confianza = "Neutro", max(confianza, 0.75)

        necesita_revision = confianza < self.cfg.confianza_minima_tono or (
            tono != "Neutro" and tam_grupo >= self.cfg.revisar_grupos_grandes
        )
        if necesita_revision:
            self.ctx.tel.tono_revisado += 1
            revision = await chat_json_async(
                SISTEMA_TONO,
                _prompt_revision_tono(self.marca, texto_para_llm(titulo, cuerpo, max_chars=3200), datos),
                self.ctx,
                max_tokens=140,
            )
            if revision:
                nuevo = str(revision.get("tono", tono)).strip().capitalize()
                if nuevo in ("Positivo", "Negativo", "Neutro"):
                    if nuevo != tono:
                        self.ctx.tel.tono_cambiado_en_revision += 1
                    tono = nuevo
                    try:
                        confianza = min(max(float(revision.get("confianza", confianza)), 0.0), 1.0)
                    except (TypeError, ValueError):
                        pass

        if confianza < 0.5:
            self.ctx.tel.tono_baja_confianza_final += 1
        return tono, confianza

    async def procesar(
        self,
        titulos: Sequence[str],
        resumenes: Sequence[str],
        id_identidad: List[int],
        embs: List[Optional[List[float]]],
        progreso: Optional[Callable[[float, str], None]] = None,
    ) -> Tuple[List[str], List[float]]:
        """Una clasificación por cluster de identidad, propagada al resto.

        A diferencia del original, la propagación NO es incondicional: una nota
        del cluster que no nombra la marca recibe Neutro, aunque el representante
        sea Negativo. Las republicaciones no siempre conservan la mención.
        """
        n = len(titulos)
        clusters: Dict[int, List[int]] = defaultdict(list)
        for i, c in enumerate(id_identidad):
            clusters[c].append(i)

        # El representante es la nota más completa del cluster, no la más cercana
        # al centroide: clasificar sobre un teletipo de dos líneas cuando existe
        # la versión desarrollada desperdicia la evidencia disponible.
        representantes: Dict[int, int] = {}
        for c, idxs in clusters.items():
            representantes[c] = max(idxs, key=lambda i: len(str(resumenes[i] or "")))

        ids = sorted(clusters.keys())
        if progreso:
            progreso(0.05, f"Tono · {len(ids)} grupos únicos de {n} noticias")

        coros = [
            self._clasificar(str(titulos[representantes[c]]), str(resumenes[representantes[c]]), len(clusters[c]))
            for c in ids
        ]
        resultados = await _mapear(coros, self.ctx.max_concurrencia)

        tonos = ["Neutro"] * n
        confianzas = [1.0] * n
        for c, (tono, conf) in zip(ids, resultados):
            for i in clusters[c]:
                if tono != "Neutro" and not self.detector.menciona(f"{titulos[i]} {resumenes[i]}"):
                    tonos[i], confianzas[i] = "Neutro", 1.0
                else:
                    tonos[i], confianzas[i] = tono, conf

        if progreso:
            progreso(1.0, "Tono completado")
        return tonos, confianzas


# ==============================================================================
# 8 · ETIQUETAS (subtema)
# ==============================================================================

SISTEMA_SUBTEMA = (
    "Eres editor jefe de un diario. Nombras el asunto de un grupo de noticias "
    "con una frase nominal breve en español, sin verbos conjugados ni sujeto. "
    "Respondes únicamente con JSON válido."
)


def capitalizar(etiqueta: str) -> str:
    if not etiqueta or not etiqueta.strip():
        return "Sin tema"
    # El modelo devuelve con frecuencia texto sin tildes ni eñes; la etiqueta
    # llega al Excel tal cual, así que se restauran aquí y no en la vista.
    e = corregir_tildes(etiqueta.strip())
    return e[0].upper() + e[1:]


def recortar_frase(texto: str, max_palabras: int = 7) -> str:
    """Recorta y elimina la cola incompleta ('Proyecto de terminal de' -> ...)."""
    if not texto:
        return ""
    palabras = texto.strip().split()
    if len(palabras) > max_palabras:
        palabras = palabras[:max_palabras]
    while palabras and unidecode(palabras[-1].lower().rstrip(".,;:!?")) in PALABRAS_CORTE_ETIQUETA:
        palabras.pop()
    return " ".join(palabras)


def frase_completa(texto: str) -> bool:
    if not texto or not texto.strip():
        return False
    palabras = texto.strip().split()
    if not palabras:
        return False
    ultima = unidecode(palabras[-1].lower().rstrip(".,;:!?"))
    return ultima not in PALABRAS_CORTE_ETIQUETA and len(ultima) > 1


def limpiar_etiqueta(bruto: str, marca: str = "", aliases: Sequence[str] = ()) -> str:
    """Normaliza la salida del modelo y borra la marca del texto."""
    if not bruto:
        return ""
    t = str(bruto).strip().strip("\"'`")
    for px in ("subtema:", "tema:", "categoría:", "categoria:", "category:", "asunto:"):
        if t.lower().startswith(px):
            t = t[len(px):].strip()
    t = re.sub(r"\s+", " ", t).strip(" .,;:")

    if marca:
        for nombre in [marca] + [a for a in aliases if a]:
            nn = unidecode(str(nombre).strip().lower())
            if len(nn) < 3:
                continue
            t = re.sub(rf"\b{re.escape(nn)}\b", "", t, flags=re.IGNORECASE)
            # También la forma con tildes tal cual aparece
            t = re.sub(rf"\b{re.escape(str(nombre).strip())}\b", "", t, flags=re.IGNORECASE)

    for frase in ("en colombia", "de colombia", "del pais", "del país",
                  "en el pais", "en el país", "a nivel nacional", "en todo el pais"):
        t = re.sub(rf"\b{re.escape(frase)}\b", "", t, flags=re.IGNORECASE)

    t = re.sub(r"\s+", " ", t).strip(" .,;:-")
    t = recortar_frase(t, max_palabras=7)
    return t.lower() if t else ""


def validar_subtema(etiqueta: str, cfg: ConfigPrecision) -> bool:
    if not etiqueta:
        return False
    palabras = etiqueta.split()
    if not (cfg.min_palabras_subtema <= len(palabras) <= cfg.max_palabras_subtema):
        return False
    if RE_VERBO_CONJUGADO.search(unidecode(etiqueta.lower())):
        return False
    if normalizar_texto(etiqueta) in {normalizar_texto(g) for g in ETIQUETAS_GENERICAS}:
        return False
    if not frase_completa(etiqueta):
        return False
    # Frases cortas necesitan nexo: "Tarifas energía" es telegráfico,
    # "Tarifas de energía" es la forma periodística correcta.
    if len(palabras) <= 4:
        if not any(unidecode(p.lower().rstrip(".,;:")) in NEXOS_VALIDOS for p in palabras[1:]):
            return False
    return True


def validar_tema(etiqueta: str, cfg: ConfigPrecision) -> bool:
    if not etiqueta:
        return False
    palabras = etiqueta.split()
    if not (cfg.min_palabras_tema <= len(palabras) <= cfg.max_palabras_tema):
        return False
    if re.match(r"^[0-9]", etiqueta):
        return False
    if RE_VERBO_CONJUGADO.search(unidecode(etiqueta.lower())):
        return False
    if normalizar_texto(etiqueta) in {normalizar_texto(g) for g in ETIQUETAS_GENERICAS}:
        return False
    return frase_completa(etiqueta)


@dataclass
class GrupoEtiquetable:
    id_grupo: int
    indices: List[int]
    titulos: List[str]
    resumenes: List[str]
    entidades: set
    centroide: Optional[np.ndarray] = None


def _ordenar_por_representatividad(
    indices: Sequence[int],
    embs: Sequence[Optional[List[float]]],
    centroide: Optional[np.ndarray],
    resumenes: Sequence[str],
) -> List[int]:
    """Ordena un grupo poniendo delante lo más representativo y mejor documentado.

    El prompt de etiquetado solo lee las primeras noticias del grupo. Si el orden
    es el de llegada, un asunto de 60 notas se nombra a partir de 6 titulares
    arbitrarios. Ordenando por cercanía al centroide (con la nota más extensa
    promovida al frente) la etiqueta describe el núcleo del asunto y deja de
    depender del orden del Excel.
    """
    idxs = list(indices)
    if len(idxs) <= 1:
        return idxs

    if centroide is not None:
        c = np.asarray(centroide, dtype=float)
        norma_c = float(np.linalg.norm(c)) or 1.0

        def _cercania(i: int) -> float:
            e = embs[i]
            if e is None:
                return -1.0
            v = np.asarray(e, dtype=float)
            nv = float(np.linalg.norm(v)) or 1.0
            return float(np.dot(v, c) / (nv * norma_c))

        orden = sorted(idxs, key=lambda i: (-_cercania(i), i))
    else:
        orden = sorted(idxs, key=lambda i: (-len(str(resumenes[i] or "")), i))

    # La nota más desarrollada aporta el contexto que a un teletipo le falta.
    mas_extensa = max(idxs, key=lambda i: (len(str(resumenes[i] or "")), -i))
    if orden and orden[0] != mas_extensa:
        orden.remove(mas_extensa)
        orden.insert(1 if len(orden) > 1 else 0, mas_extensa)
    return orden


def _fallback_etiqueta(titulos: Sequence[str], entidades: set, tel: Telemetria) -> str:
    """Etiqueta sin LLM. Se registra: un informe lleno de fallbacks es una alarma."""
    tel.etiquetas_fallback += 1
    palabras = []
    for t in list(titulos)[:6]:
        for w in normalizar_texto(t).split():
            if len(w) > 4 and w not in TOKENS_DEBILES:
                palabras.append(w)
    top = [w for w, _ in Counter(palabras).most_common(3)]
    ent = sorted(entidades, key=len, reverse=True)
    if len(top) >= 2:
        return capitalizar(f"{top[0]} de {top[1]}")
    if top and ent:
        return capitalizar(f"{top[0]} de {ent[0]}")
    if top:
        return capitalizar(f"Asuntos de {top[0]}")
    return "Cobertura informativa sin clasificar"


def _prompt_subtema(g: GrupoEtiquetable, vocabulario: Sequence[str]) -> str:
    titulares = list(dict.fromkeys(str(t)[:130] for t in g.titulos if t and str(t).strip().lower() != "nan"))[:6]
    resumenes = [str(r)[:220] for r in g.resumenes[:3] if r and len(str(r)) > 20]

    palabras = []
    for t in g.titulos[:12]:
        for w in normalizar_texto(t).split():
            if len(w) > 3 and w not in TOKENS_DEBILES:
                palabras.append(w)
    kw = ", ".join(w for w, _ in Counter(palabras).most_common(8))
    ents = ", ".join(sorted(g.entidades, key=len, reverse=True)[:6])

    bloque_voc = ""
    if vocabulario:
        bloque_voc = (
            "\n\nSUBTEMAS YA EXISTENTES EN ESTE INFORME:\n"
            + "\n".join(f"  · {s}" for s in vocabulario)
            + "\nSi este grupo trata EXACTAMENTE el mismo asunto que uno de ellos, "
            "devuélvelo tal cual. Si el asunto es distinto, aunque sea del mismo "
            "campo, crea uno nuevo. No fuerces la coincidencia."
        )

    return (
        f"GRUPO DE {len(g.titulos)} NOTICIA(S)\n\n"
        "TITULARES:\n" + "\n".join(f"  · {t}" for t in titulares)
        + (("\n\nCONTEXTO:\n" + "\n".join(f"  · {r}" for r in resumenes)) if resumenes else "")
        + (f"\n\nENTIDADES: {ents}" if ents else "")
        + f"\n\nPALABRAS CLAVE: {kw}"
        + bloque_voc
        + "\n\nEscribe el SUBTEMA: frase nominal de 4 a 7 palabras que nombre el "
        "ASUNTO concreto de estas noticias.\n\n"
        "FORMATO OBLIGATORIO\n"
        "- Empieza por sustantivo. Sin verbo conjugado. Sin sujeto ni cargo.\n"
        "- Une los conceptos con preposición (de, del, para, sobre, en, por).\n"
        "- Nombra el asunto, no el actor ni el género periodístico.\n"
        "- Sin nombres de empresas privadas. Con tildes y ñ correctas.\n"
        "- Las ciudades y regiones pueden aparecer si definen el asunto.\n\n"
        "EJEMPLOS\n"
        '  OK  "Ampliación de la red de acueducto"\n'
        '  OK  "Regulación de tarifas de energía"\n'
        '  OK  "Operación del Canal del Dique"\n'
        '  MAL "Alcalde presenta obra"      (verbo + cargo)\n'
        '  MAL "Gestión institucional"      (genérico)\n'
        '  MAL "Tarifas energía"            (sin preposición)\n\n'
        'Responde JSON: {"asunto":"qué ocurre, máx 12 palabras","subtema":"..."}'
    )


class EtiquetadorSubtemas:
    """Etiqueta en dos rondas paralelas.

    El original etiquetaba en bucle secuencial pasando al prompt las etiquetas ya
    aprobadas, lo que producía sesgo de anclaje: los primeros grupos fijaban el
    vocabulario y los siguientes reutilizaban esos nombres aunque no
    correspondieran, con un orden que además dependía del tamaño de grupo.

    Aquí:
      Ronda A · todos los grupos se etiquetan en paralelo y sin ver a los demás.
      Ronda B · las etiquetas equivalentes se detectan por embedding + contenido
                y se unifican con una sola llamada por familia.
    """

    def __init__(self, marca: str, aliases: Sequence[str], ctx: EngineContext, cfg: ConfigPrecision):
        self.marca = marca
        self.aliases = list(aliases or [])
        self.ctx = ctx
        self.cfg = cfg

    async def _etiquetar_grupo(self, g: GrupoEtiquetable, vocabulario: Sequence[str]) -> str:
        datos = await chat_json_async(
            SISTEMA_SUBTEMA, _prompt_subtema(g, vocabulario), self.ctx, max_tokens=140
        )
        if datos:
            etiqueta = limpiar_etiqueta(datos.get("subtema", ""), self.marca, self.aliases)
            if validar_subtema(etiqueta, self.cfg):
                self.ctx.tel.etiquetas_llm += 1
                return capitalizar(etiqueta)
            # Un solo reintento dirigido al defecto detectado, en vez de las
            # cuatro llamadas encadenadas del pipeline original.
            correccion = await self._reintentar(g, etiqueta)
            if correccion:
                self.ctx.tel.etiquetas_llm += 1
                return capitalizar(correccion)
        return _fallback_etiqueta(g.titulos, g.entidades, self.ctx.tel)

    async def _reintentar(self, g: GrupoEtiquetable, fallida: str) -> Optional[str]:
        problemas = []
        if fallida and RE_VERBO_CONJUGADO.search(unidecode(fallida.lower())):
            problemas.append("contiene un verbo conjugado")
        if fallida and len(fallida.split()) < self.cfg.min_palabras_subtema:
            problemas.append("es demasiado corta")
        if fallida and not frase_completa(fallida):
            problemas.append("termina en preposición o artículo")
        if fallida and normalizar_texto(fallida) in {normalizar_texto(x) for x in ETIQUETAS_GENERICAS}:
            problemas.append("es genérica y no describe el asunto")
        if not problemas:
            problemas.append("no cumple el formato de frase nominal con preposición")

        titulares = list(dict.fromkeys(str(t)[:130] for t in g.titulos))[:5]
        usuario = (
            f"Tu respuesta anterior fue \"{fallida}\" y {', '.join(problemas)}.\n\n"
            "TITULARES:\n" + "\n".join(f"  · {t}" for t in titulares) + "\n\n"
            "Reescribe el subtema como frase nominal de 4 a 7 palabras, empezando por "
            "sustantivo, uniendo los conceptos con preposición y sin ningún verbo "
            "conjugado. Debe describir el asunto concreto de estos titulares.\n\n"
            'Responde JSON: {"subtema":"..."}'
        )
        datos = await chat_json_async(SISTEMA_SUBTEMA, usuario, self.ctx, max_tokens=90, temperature=0.15)
        if not datos:
            return None
        etiqueta = limpiar_etiqueta(datos.get("subtema", ""), self.marca, self.aliases)
        return etiqueta if validar_subtema(etiqueta, self.cfg) else None

    async def _unificar(self, candidatas: Sequence[str], titulares: Sequence[str]) -> Optional[str]:
        usuario = (
            "Estos nombres describen el MISMO asunto:\n"
            + "\n".join(f"  · {c}" for c in candidatas)
            + "\n\nTITULARES DE REFERENCIA:\n"
            + "\n".join(f"  · {t[:110]}" for t in list(dict.fromkeys(titulares))[:6])
            + "\n\nElige el mejor de la lista o redacta uno nuevo que los cubra a todos: "
            "frase nominal de 4 a 7 palabras, empezando por sustantivo, con preposición, "
            "sin verbo conjugado y sin nombres de empresas.\n\n"
            'Responde JSON: {"subtema":"..."}'
        )
        datos = await chat_json_async(SISTEMA_SUBTEMA, usuario, self.ctx, max_tokens=90)
        if not datos:
            return None
        etiqueta = limpiar_etiqueta(datos.get("subtema", ""), self.marca, self.aliases)
        return capitalizar(etiqueta) if validar_subtema(etiqueta, self.cfg) else None

    async def _reformular_concreta(self, g: GrupoEtiquetable, generica: str) -> Optional[str]:
        """Segunda oportunidad cuando la etiqueta describe al corpus entero.

        No basta con pedir "sé más específico": se le devuelven las entidades
        fuertes del grupo (las que el filtro IDF considera discriminantes) y se
        le exige anclarse en el hecho, que es justo lo que distingue a este grupo
        de todos los demás.
        """
        titulares = list(dict.fromkeys(str(t)[:130] for t in g.titulos))[:6]
        ents = sorted(g.entidades, key=len, reverse=True)[:6]
        usuario = (
            f'La etiqueta "{generica}" describe igual de bien a estas noticias que '
            "al resto del informe, así que no sirve para distinguirlas.\n\n"
            "TITULARES:\n" + "\n".join(f"  · {t}" for t in titulares)
            + (f"\n\nELEMENTOS DISTINTIVOS: {', '.join(ents)}" if ents else "")
            + "\n\nEscribe un subtema que nombre el HECHO CONCRETO que comparten estas "
            "noticias y que no encajaría en una noticia cualquiera del informe: "
            "frase nominal de 4 a 7 palabras, empezando por sustantivo, unida con "
            "preposición, sin verbo conjugado y sin nombres de empresas privadas.\n"
            "Prohibido responder con categorías amplias (gestión, actualidad, "
            "desarrollo, impacto, panorama, situación).\n\n"
            'Responde JSON: {"subtema":"..."}'
        )
        datos = await chat_json_async(SISTEMA_SUBTEMA, usuario, self.ctx, max_tokens=90, temperature=0.2)
        if not datos:
            return None
        etiqueta = limpiar_etiqueta(datos.get("subtema", ""), self.marca, self.aliases)
        return capitalizar(etiqueta) if validar_subtema(etiqueta, self.cfg) else None

    async def _depurar_genericas(
        self,
        grupos: List[GrupoEtiquetable],
        candidatas: List[str],
        progreso: Optional[Callable[[float, str], None]] = None,
    ) -> List[str]:
        """Detecta etiquetas vagas sin listas negras: por geometría.

        Una etiqueta específica está mucho más cerca de su propio grupo que del
        corpus completo. Una etiqueta genérica ("Desarrollo del sector") está a
        una distancia parecida de todo. Comparando ambas similitudes se detectan
        generalizaciones que ninguna lista de palabras prohibidas anticipa,
        porque el criterio depende del corpus concreto de cada informe.
        """
        centros = [g.centroide for g in grupos if g.centroide is not None]
        if len(centros) < 3 or len(set(candidatas)) < 2:
            return candidatas

        embs_lab = self.ctx.embed(list(candidatas))
        global_c = np.mean(np.array(centros), axis=0).reshape(1, -1)

        sospechosas: List[int] = []
        for k, g in enumerate(grupos):
            if g.centroide is None or embs_lab[k] is None:
                continue
            v = np.asarray(embs_lab[k], dtype=float).reshape(1, -1)
            s_propio = float(cosine_similarity(v, np.asarray(g.centroide).reshape(1, -1))[0][0])
            s_global = float(cosine_similarity(v, global_c)[0][0])
            if s_propio < s_global + self.cfg.margen_especificidad:
                sospechosas.append(k)

        if not sospechosas:
            return candidatas
        if progreso:
            progreso(0.45, f"Subtemas · reformulando {len(sospechosas)} etiquetas poco específicas")

        nuevas = await _mapear(
            [self._reformular_concreta(grupos[k], candidatas[k]) for k in sospechosas],
            self.ctx.max_concurrencia,
        )
        salida = list(candidatas)
        for k, propuesta in zip(sospechosas, nuevas):
            if propuesta and normalizar_texto(propuesta) != normalizar_texto(candidatas[k]):
                salida[k] = propuesta
                self.ctx.tel.etiquetas_reformuladas_genericas += 1
        return salida

    async def etiquetar(
        self,
        grupos: List[GrupoEtiquetable],
        textos: Sequence[str],
        idx_ents: IndiceEntidades,
        progreso: Optional[Callable[[float, str], None]] = None,
    ) -> Dict[int, str]:
        if not grupos:
            return {}

        # ── Ronda A · en paralelo, sin contexto cruzado ─────────────────────
        if progreso:
            progreso(0.10, f"Subtemas · etiquetando {len(grupos)} grupos en paralelo")
        candidatas = await _mapear(
            [self._etiquetar_grupo(g, []) for g in grupos], self.ctx.max_concurrencia
        )

        # ── Prueba de especificidad · antes de consolidar ───────────────────
        # Se hace aquí y no al final: una etiqueta vaga arrastra a su familia
        # entera durante la unificación y contamina el tema que se derive de ella.
        candidatas = await self._depurar_genericas(grupos, list(candidatas), progreso)

        # ── Consolidación · qué etiquetas nombran lo mismo ──────────────────
        if progreso:
            progreso(0.55, "Subtemas · consolidando etiquetas equivalentes")

        unicas = list(dict.fromkeys(candidatas))
        mapa_final: Dict[str, str] = {u: u for u in unicas}

        if len(unicas) > 1:
            embs_lab = self.ctx.embed(list(unicas))
            validos = [k for k, e in enumerate(embs_lab) if e is not None]
            if len(validos) >= 2:
                M = np.array([embs_lab[k] for k in validos])
                sim = cosine_similarity(M)
                dsu = DSUPrecision(len(unicas))

                grupos_por_etiqueta: Dict[str, List[GrupoEtiquetable]] = defaultdict(list)
                for g, c in zip(grupos, candidatas):
                    grupos_por_etiqueta[c].append(g)
                frecuencia = Counter(candidatas)

                for a in range(len(validos)):
                    for b in range(a + 1, len(validos)):
                        s = float(sim[a][b])
                        if s < self.cfg.sim_unificar_etiquetas:
                            continue
                        ea, eb = unicas[validos[a]], unicas[validos[b]]
                        if dsu.find(validos[a]) == dsu.find(validos[b]):
                            continue
                        if hay_conflicto_accion(ea, eb):
                            self.ctx.tel.uniones_rechazadas_conflicto += 1
                            continue
                        # Dos etiquetas parecidas no bastan: se comprueba que el
                        # CONTENIDO que cubren también lo sea. "Inversión en vías"
                        # e "Inversión en salud" puntúan 0.91 como cadenas.
                        if self._contenidos_compatibles(
                            grupos_por_etiqueta[ea], grupos_por_etiqueta[eb], idx_ents
                        ):
                            dsu.union(validos[a], validos[b])

                for miembros in sorted(dsu.grupos().values(), key=lambda ms: (len(ms), ms[0])):
                    if len(miembros) < 2:
                        continue
                    familia = [unicas[m] for m in miembros]
                    titulares = []
                    for e in familia:
                        for g in grupos_por_etiqueta[e]:
                            titulares.extend(g.titulos[:3])
                    canonica = max(familia, key=lambda x: (frecuencia[x], len(x)))
                    if len(familia) <= 4:
                        propuesta = await self._unificar(familia, titulares)
                        if propuesta:
                            canonica = propuesta
                    for e in familia:
                        mapa_final[e] = canonica
                    self.ctx.tel.etiquetas_unificadas += len(familia) - 1

        if progreso:
            progreso(0.95, "Subtemas listos")
        return {g.id_grupo: capitalizar(mapa_final.get(c, c)) for g, c in zip(grupos, candidatas)}

    def _contenidos_compatibles(
        self,
        grupos_a: List[GrupoEtiquetable],
        grupos_b: List[GrupoEtiquetable],
        idx_ents: IndiceEntidades,
    ) -> bool:
        if not grupos_a or not grupos_b:
            return False
        ca = [g.centroide for g in grupos_a if g.centroide is not None]
        cb = [g.centroide for g in grupos_b if g.centroide is not None]
        if not ca or not cb:
            return False
        sim = float(
            cosine_similarity(
                np.mean(ca, axis=0).reshape(1, -1), np.mean(cb, axis=0).reshape(1, -1)
            )[0][0]
        )
        if sim < self.cfg.sim_asunto:
            return False
        ents_a: set = set()
        ents_b: set = set()
        for g in grupos_a:
            ents_a |= g.entidades
        for g in grupos_b:
            ents_b |= g.entidades
        txt_a = " ".join(t for g in grupos_a for t in g.titulos[:8])
        txt_b = " ".join(t for g in grupos_b for t in g.titulos[:8])
        return jaccard(ents_a, ents_b) >= self.cfg.jaccard_entidades_min or (
            overlap_distintivo(txt_a, txt_b) >= self.cfg.overlap_tokens_min
        )


# ==============================================================================
# 9 · TEMAS
# ==============================================================================

SISTEMA_TEMA = (
    "Eres editor jefe de un diario y construyes la taxonomía de secciones de un "
    "informe de medios. Respondes únicamente con JSON válido."
)


async def _nombrar_tema(
    subtemas: Sequence[str],
    titulares: Sequence[str],
    ctx: EngineContext,
    cfg: ConfigPrecision,
    marca: str,
    aliases: Sequence[str],
) -> Optional[str]:
    palabras = []
    for t in list(titulares)[:15]:
        for w in normalizar_texto(t).split():
            if len(w) > 3 and w not in TOKENS_DEBILES:
                palabras.append(w)
    kw = ", ".join(w for w, _ in Counter(palabras).most_common(6))

    usuario = (
        "SUBTEMAS A AGRUPAR:\n" + "\n".join(f"  · {s}" for s in list(subtemas)[:10])
        + "\n\nTITULARES DE REFERENCIA:\n"
        + "\n".join(f"  · {t[:100]}" for t in list(dict.fromkeys(titulares))[:5])
        + f"\n\nPALABRAS CLAVE: {kw}\n\n"
        "Escribe el TEMA: categoría editorial de 2 a 4 palabras que englobe esos subtemas.\n\n"
        "REGLAS\n"
        "- Más general que los subtemas, pero sin perder el asunto que los une.\n"
        "- Prohibidas las secciones vagas de una palabra (Economía, Política, Actualidad).\n"
        "- Sin números, cantidades ni nombres propios de personas o empresas.\n"
        "- Sustantivo + complemento o adjetivo. Sin verbo conjugado. Con tildes y ñ.\n"
        "- No copies literalmente ninguno de los subtemas.\n\n"
        "EJEMPLOS\n"
        '  OK  "Infraestructura vial"   "Regulación financiera"\n'
        '  OK  "Movilidad urbana"       "Salud pública territorial"\n'
        '  MAL "Economía"               (sección vaga)\n'
        '  MAL "Nuevo acuerdo firmado"  (titular)\n\n'
        'Responde JSON: {"tema":"..."}'
    )
    datos = await chat_json_async(SISTEMA_TEMA, usuario, ctx, max_tokens=60)
    if not datos:
        return None
    tema = limpiar_etiqueta(datos.get("tema", ""), marca, aliases)
    tema = recortar_frase(tema, max_palabras=cfg.max_palabras_tema)
    return capitalizar(tema) if validar_tema(tema, cfg) else None


def _tema_igual_a_subtema(tema: str, subtemas: Sequence[str]) -> bool:
    tn = normalizar_texto(tema)
    if not tn:
        return True
    for s in subtemas:
        sn = normalizar_texto(s)
        if not sn:
            continue
        if SequenceMatcher(None, tn, sn).ratio() >= 0.82:
            return True
        if tn in sn or sn in tn:
            return True
    return False


async def construir_temas(
    subtemas: List[str],
    textos: Sequence[str],
    embs: List[Optional[List[float]]],
    idx_ents: IndiceEntidades,
    ctx: EngineContext,
    cfg: ConfigPrecision,
    marca: str,
    aliases: Sequence[str],
    num_temas_max: Optional[int] = None,
    progreso: Optional[Callable[[float, str], None]] = None,
) -> List[str]:
    """NIVEL 2 · agrupa subtemas afines bajo una categoría editorial.

    Garantía estructural: cada subtema pertenece a exactamente un tema. En el
    original el tema se decidía por noticia y luego se reconciliaba por voto,
    de modo que dos noticias con el mismo subtema podían acabar en temas
    distintos según el orden de las fases de validación.
    """
    unicos = list(dict.fromkeys(subtemas))
    if len(unicos) <= 1:
        return [capitalizar(s) for s in subtemas]

    if progreso:
        progreso(0.05, f"Temas · agrupando {len(unicos)} subtemas")

    indices_por_sub: Dict[str, List[int]] = defaultdict(list)
    for i, s in enumerate(subtemas):
        indices_por_sub[s].append(i)

    # Representación de cada subtema: etiqueta + centroide de su contenido.
    centroides: Dict[str, np.ndarray] = {}
    for s in unicos:
        vecs = [embs[i] for i in indices_por_sub[s][:50] if embs[i] is not None]
        if vecs:
            centroides[s] = np.mean(vecs, axis=0)

    embs_lab = ctx.embed(list(unicos))
    validos = [s for s in unicos if s in centroides]
    if len(validos) < 2:
        return [capitalizar(s) for s in subtemas]

    pos_lab = {s: k for k, s in enumerate(unicos)}
    M_cont = np.array([centroides[s] for s in validos])
    sim = cosine_similarity(M_cont)

    if all(embs_lab[pos_lab[s]] is not None for s in validos):
        M_lab = np.array([embs_lab[pos_lab[s]] for s in validos])
        # El contenido manda; la etiqueta solo matiza. Al revés, dos subtemas con
        # nombres parecidos y contenidos distintos acabarían en el mismo tema.
        sim = 0.70 * sim + 0.30 * cosine_similarity(M_lab)

    np.fill_diagonal(sim, 1.0)
    dist = np.clip(1.0 - sim, 0.0, 2.0)
    np.fill_diagonal(dist, 0.0)

    if progreso:
        progreso(0.25, "Temas · clustering")

    etiquetas = AgglomerativeClustering(
        n_clusters=None,
        distance_threshold=1.0 - cfg.sim_tema,
        metric="precomputed",
        linkage="average",
    ).fit(dist).labels_

    provisional: Dict[int, List[int]] = defaultdict(list)
    for k, lbl in enumerate(etiquetas):
        provisional[lbl].append(k)

    dsu = DSUPrecision(len(validos))
    textos_rep = {s: " ".join(str(textos[i])[:250] for i in indices_por_sub[s][:8]) for s in validos}
    ents_por_sub = {
        s: set().union(*[idx_ents.fuertes[i] for i in indices_por_sub[s][:15]]) if indices_por_sub[s] else set()
        for s in validos
    }

    for miembros in provisional.values():
        if len(miembros) < 2:
            continue
        for a in range(len(miembros)):
            for b in range(a + 1, len(miembros)):
                ka, kb = miembros[a], miembros[b]
                sa, sb = validos[ka], validos[kb]
                s_val = float(sim[ka][kb])
                if s_val < cfg.piso_absoluto_tema:
                    continue
                if hay_conflicto_accion(f"{sa} {textos_rep[sa]}", f"{sb} {textos_rep[sb]}"):
                    ctx.tel.uniones_rechazadas_conflicto += 1
                    continue
                ov = overlap_distintivo(f"{sa} {textos_rep[sa]}", f"{sb} {textos_rep[sb]}")
                jac = jaccard(ents_por_sub[sa], ents_por_sub[sb])
                if s_val >= cfg.sim_tema and (ov >= cfg.overlap_tokens_min_tema or jac >= 0.20):
                    dsu.union(ka, kb)
                else:
                    ctx.tel.uniones_rechazadas_overlap += 1

    familias: Dict[int, List[str]] = defaultdict(list)
    for k, s in enumerate(validos):
        familias[dsu.find(k)].append(s)

    # `num_temas_max` era un parámetro fantasma en el original: se calculaba y se
    # mostraba en pantalla, pero el clustering usaba distance_threshold y nunca
    # lo aplicaba. Aquí se aplica de verdad, fusionando solo los pares más
    # similares y nunca por debajo del piso: el tope no puede forzar una fusión
    # falsa.
    if num_temas_max and len(familias) > num_temas_max:
        fusionadas = _reducir_familias(familias, validos, sim, cfg, num_temas_max, ctx.tel)
        familias = fusionadas

    if progreso:
        progreso(0.45, f"Temas · nombrando {len(familias)} categorías")

    claves = list(familias.keys())
    coros = []
    for clave in claves:
        subs = familias[clave]
        titulares = []
        for s in subs:
            for i in indices_por_sub[s][:6]:
                titulares.append(str(textos[i]).split(". ")[0][:120])
        coros.append(_nombrar_tema(subs, titulares, ctx, cfg, marca, aliases))
    nombres = await _mapear(coros, ctx.max_concurrencia)

    sub_a_tema: Dict[str, str] = {}
    usados: set = set()
    for clave, nombre in zip(claves, nombres):
        subs = familias[clave]
        if not nombre or _tema_igual_a_subtema(nombre, subs) or normalizar_texto(nombre) in usados:
            nombre = _tema_fallback(subs, ctx.tel)
        usados.add(normalizar_texto(nombre))
        for s in subs:
            sub_a_tema[s] = nombre

    # Subtemas sin embedding válido: tema propio derivado de su propio nombre.
    for s in unicos:
        if s not in sub_a_tema:
            sub_a_tema[s] = capitalizar(recortar_frase(s, max_palabras=cfg.max_palabras_tema) or s)

    if progreso:
        progreso(1.0, "Temas listos")
    return [sub_a_tema[s] for s in subtemas]


def _reducir_familias(
    familias: Dict[int, List[str]],
    validos: List[str],
    sim: np.ndarray,
    cfg: ConfigPrecision,
    tope: int,
    tel: Telemetria,
) -> Dict[int, List[str]]:
    pos = {s: k for k, s in enumerate(validos)}
    actuales = {k: list(v) for k, v in familias.items()}

    while len(actuales) > tope:
        mejor: Optional[Tuple[float, Any, Any]] = None
        claves = list(actuales.keys())
        for a in range(len(claves)):
            for b in range(a + 1, len(claves)):
                ia = [pos[s] for s in actuales[claves[a]]]
                ib = [pos[s] for s in actuales[claves[b]]]
                s_val = float(np.mean(sim[np.ix_(ia, ib)]))
                if mejor is None or s_val > mejor[0]:
                    mejor = (s_val, claves[a], claves[b])
        if mejor is None or mejor[0] < cfg.piso_absoluto_tema:
            tel.avisar(
                f"Se alcanzó el tope de {tope} temas pero quedaron {len(actuales)}: "
                "fusionar más habría unido categorías no relacionadas."
            )
            break
        _, ka, kb = mejor
        actuales[ka].extend(actuales[kb])
        del actuales[kb]

    return actuales


def _tema_fallback(subtemas: Sequence[str], tel: Telemetria) -> str:
    tel.etiquetas_fallback += 1
    palabras = []
    for s in subtemas:
        for w in normalizar_texto(s).split():
            if len(w) > 3 and w not in TOKENS_DEBILES:
                palabras.append(w)
    top = [w for w, _ in Counter(palabras).most_common(2)]
    if len(top) >= 2:
        return capitalizar(f"{top[0]} y {top[1]}")
    if top:
        return capitalizar(top[0])
    return "Cobertura general"


# ==============================================================================
# 10 · COHERENCIA Y ARMONIZACIÓN
# ==============================================================================

def reasignar_por_coherencia(
    subtemas: List[str],
    embs: List[Optional[List[float]]],
    id_asunto: List[int],
    cfg: ConfigPrecision,
    tel: Telemetria,
) -> List[str]:
    """Mueve un ASUNTO COMPLETO a otro subtema solo si la mejora es clara.

    La unidad de reasignación es el asunto, no la noticia suelta. Mover notas
    individuales rompía justo la garantía que persigue el motor: dos noticias del
    mismo hecho podían acabar en subtemas distintos porque una de ellas quedaba
    marginalmente más cerca del centroide vecino. Se compara el centroide del
    asunto contra los centroides de subtema y se exige margen y mínimo absoluto.
    """
    n = len(subtemas)
    unicos = list(dict.fromkeys(subtemas))
    if len(unicos) < 2:
        return subtemas

    centroides = {}
    for s in unicos:
        vecs = [embs[i] for i in range(n) if subtemas[i] == s and embs[i] is not None]
        if vecs:
            centroides[s] = np.mean(vecs, axis=0)
    subs_validos = [s for s in unicos if s in centroides]
    if len(subs_validos) < 2:
        return subtemas

    por_asunto: Dict[int, List[int]] = defaultdict(list)
    for i, g in enumerate(id_asunto):
        por_asunto[g].append(i)

    # Centroide de cada asunto: un asunto sin ningún embedding no se toca.
    asuntos: List[int] = []
    vec_asunto: List[np.ndarray] = []
    for g in sorted(por_asunto.keys()):
        vecs = [embs[i] for i in por_asunto[g] if embs[i] is not None]
        if not vecs:
            continue
        asuntos.append(g)
        vec_asunto.append(np.mean(vecs, axis=0))
    if not asuntos:
        return subtemas

    sim = cosine_similarity(np.array(vec_asunto), np.array([centroides[s] for s in subs_validos]))
    pos_sub = {s: k for k, s in enumerate(subs_validos)}

    salida = list(subtemas)
    MARGEN = 0.06        # la alternativa debe ganar por un margen real
    MINIMO = 0.62        # y superar un mínimo absoluto de pertenencia

    for fila, g in enumerate(asuntos):
        idxs = por_asunto[g]
        actual = Counter(subtemas[i] for i in idxs).most_common(1)[0][0]
        if actual not in pos_sub:
            continue
        s_actual = float(sim[fila][pos_sub[actual]])
        mejor_k = int(np.argmax(sim[fila]))
        destino = subs_validos[mejor_k]
        if destino == actual:
            continue
        s_mejor = float(sim[fila][mejor_k])
        if s_mejor >= MINIMO and s_mejor >= s_actual + MARGEN:
            for i in idxs:
                salida[i] = destino
            tel.subtemas_reasignados_coherencia += 1

    return armonizar(salida, id_asunto)


def armonizar_tono(
    tonos: List[str],
    confianzas: List[float],
    grupos: List[int],
    menciona: List[bool],
    tel: Telemetria,
) -> Tuple[List[str], List[float]]:
    """Unifica el tono dentro de cada asunto, ponderando por confianza.

    Un voto mayoritario simple no sirve aquí: las notas sin mención literal de la
    marca son Neutro por construcción y, al ser mayoría, arrastrarían a Neutro un
    asunto claramente negativo. Solo votan las notas que nombran la marca, con
    peso igual a su confianza. Si nadie la nombra, o hay empate, se queda Neutro:
    ante la duda, el tono más conservador.
    """
    por_grupo: Dict[int, List[int]] = defaultdict(list)
    for i, g in enumerate(grupos):
        por_grupo[g].append(i)

    salida = list(tonos)
    salida_conf = list(confianzas)
    for g in sorted(por_grupo.keys()):
        idxs = por_grupo[g]
        if len(idxs) < 2:
            continue
        votos: Dict[str, float] = defaultdict(float)
        for i in idxs:
            if not menciona[i]:
                continue
            votos[tonos[i]] += max(float(confianzas[i]), 0.01)
        if not votos:
            continue
        mejor = max(votos.values())
        empatados = sorted(t for t, v in votos.items() if abs(v - mejor) < 1e-9)
        ganador = "Neutro" if len(empatados) > 1 else empatados[0]
        if any(tonos[i] != ganador for i in idxs if menciona[i]):
            tel.tono_armonizado_por_asunto += 1
        for i in idxs:
            # Una nota que no nombra la marca sigue siendo Neutro: no hereda el
            # tono del asunto solo por pertenecer a él.
            if not menciona[i]:
                salida[i], salida_conf[i] = "Neutro", 1.0
            else:
                salida[i] = ganador
                salida_conf[i] = max(salida_conf[i], 0.0)
    return salida, salida_conf


def armonizar(valores: List[str], grupos: List[int]) -> List[str]:
    """Fuerza un único valor por grupo (voto mayoritario, desempate por frecuencia global).

    Esta es la garantía que pedía el flujo: si dos noticias son la misma noticia,
    su tono, tema y subtema deben coincidir, pase lo que pase en las fases
    intermedias.
    """
    por_grupo: Dict[int, List[int]] = defaultdict(list)
    for i, g in enumerate(grupos):
        por_grupo[g].append(i)
    global_freq = Counter(valores)
    salida = list(valores)
    for idxs in por_grupo.values():
        if len(idxs) < 2:
            continue
        conteo = Counter(valores[i] for i in idxs)
        ganador = max(conteo.items(), key=lambda kv: (kv[1], global_freq[kv[0]], len(kv[0])))[0]
        for i in idxs:
            salida[i] = ganador
    return salida


# ==============================================================================
# 11 · ORQUESTADOR
# ==============================================================================

async def analizar_corpus(
    titulos: Sequence[str],
    resumenes: Sequence[str],
    marca: str,
    aliases: Sequence[str],
    ctx: EngineContext,
    cfg: Optional[ConfigPrecision] = None,
    progreso: Optional[Callable[[float, str], None]] = None,
    calcular_tono: bool = True,
    calcular_subtemas: bool = True,
    calcular_temas: bool = True,
    num_temas_max: Optional[int] = None,
) -> ResultadoAnalisis:
    """Punto de entrada único del motor."""
    titulos = [str(t or "") for t in titulos]
    resumenes = [str(r if r is not None and str(r).lower() != "nan" else "") for r in resumenes]
    n = len(titulos)
    tel = ctx.tel

    if n == 0:
        return ResultadoAnalisis([], [], [], [], [], [], tel)

    cfg = (cfg or ConfigPrecision()).escalar_por_corpus(n)

    def _p(frac: float, msg: str) -> None:
        if progreso:
            progreso(min(max(frac, 0.0), 1.0), msg)

    # ── Embeddings y señales léxicas ────────────────────────────────────────
    _p(0.02, "Preparando representaciones...")
    textos_emb = [_pe_texto_para_embedding(titulos[i], resumenes[i]) for i in range(n)]
    embs = ctx.embed(textos_emb)
    if sum(1 for e in embs if e is None) > n * 0.2:
        tel.avisar(
            "Más del 20% de los embeddings falló: la agrupación será conservadora "
            "y muchas noticias quedarán sin agrupar."
        )

    entidades = [extraer_entidades(titulos[i], resumenes[i]) for i in range(n)]
    idx_ents = IndiceEntidades(entidades, cfg)

    # ── NIVEL 0 · identidad ─────────────────────────────────────────────────
    _p(0.08, "Nivel 0 · detectando republicaciones...")
    id_identidad = construir_identidad(titulos, resumenes, textos_emb, embs, idx_ents, cfg, tel)

    # ── NIVEL 1 · asunto ────────────────────────────────────────────────────
    _p(0.14, "Nivel 1 · agrupando noticias del mismo hecho...")
    id_asunto = construir_asunto(textos_emb, embs, id_identidad, idx_ents, cfg, tel)

    n_ident = len(set(id_identidad))
    n_asunto = len(set(id_asunto))
    _p(0.18, f"{n} noticias · {n_ident} únicas · {n_asunto} asuntos")

    # ── Tono ────────────────────────────────────────────────────────────────
    tonos = ["N/A"] * n
    confianzas = [0.0] * n
    if calcular_tono:
        motor = MotorTono(marca, aliases, ctx, cfg)
        tonos, confianzas = await motor.procesar(
            titulos, resumenes, id_identidad, embs,
            progreso=lambda f, m: _p(0.20 + 0.25 * f, m),
        )
        # El tono se unifica en los dos niveles: primero la republicación exacta,
        # después el hecho completo. Sin el segundo paso, dos coberturas del mismo
        # hecho podían salir con tonos opuestos por haberse clasificado aparte.
        tonos = armonizar(tonos, id_identidad)
        menciona = [motor.detector.menciona(f"{titulos[i]} {resumenes[i]}") for i in range(n)]
        tonos, confianzas = armonizar_tono(tonos, confianzas, id_asunto, menciona, tel)

    # ── Subtemas ────────────────────────────────────────────────────────────
    subtemas = ["N/A"] * n
    if calcular_subtemas:
        _p(0.46, "Construyendo subtemas...")
        grupos: List[GrupoEtiquetable] = []
        por_asunto: Dict[int, List[int]] = defaultdict(list)
        for i, g in enumerate(id_asunto):
            por_asunto[g].append(i)

        # UN grupo por asunto, sin trocear. Trocear un asunto grande hacía que
        # cada trozo se etiquetara por separado y noticias del mismo hecho
        # acabaran con subtemas distintos. En su lugar se ordena el grupo por
        # representatividad: el prompt solo lee las primeras noticias, así que
        # basta con que las primeras sean las más centrales y las más completas.
        for g, idxs in sorted(por_asunto.items(), key=lambda kv: (-len(kv[1]), kv[0])):
            vecs = [embs[i] for i in idxs if embs[i] is not None]
            centroide = np.mean(vecs, axis=0) if vecs else None
            orden = _ordenar_por_representatividad(idxs, embs, centroide, resumenes)
            ents: set = set()
            for i in orden[:20]:
                ents |= idx_ents.fuertes[i]
            grupos.append(
                GrupoEtiquetable(
                    id_grupo=g,
                    indices=idxs,
                    titulos=[titulos[i] for i in orden],
                    resumenes=[resumenes[i] for i in orden],
                    entidades=ents,
                    centroide=centroide,
                )
            )

        etiquetador = EtiquetadorSubtemas(marca, aliases, ctx, cfg)
        mapa = await etiquetador.etiquetar(
            grupos, textos_emb, idx_ents,
            progreso=lambda f, m: _p(0.46 + 0.28 * f, m),
        )
        subtemas = ["Sin clasificar"] * n
        for g in grupos:
            etiqueta = mapa.get(g.id_grupo, "Sin clasificar")
            for i in g.indices:
                subtemas[i] = etiqueta

        _p(0.76, "Verificando coherencia noticia ↔ subtema...")
        subtemas = reasignar_por_coherencia(subtemas, embs, id_asunto, cfg, tel)
        # Doble red: primero el hecho, luego la republicación. Tras esto es
        # imposible que dos noticias del mismo asunto tengan subtemas distintos.
        subtemas = armonizar(subtemas, id_asunto)
        subtemas = armonizar(subtemas, id_identidad)

    # ── Temas ───────────────────────────────────────────────────────────────
    temas = ["N/A"] * n
    if calcular_temas and calcular_subtemas:
        _p(0.80, "Construyendo temas...")
        temas = await construir_temas(
            subtemas, textos_emb, embs, idx_ents, ctx, cfg, marca, aliases,
            num_temas_max=num_temas_max,
            progreso=lambda f, m: _p(0.80 + 0.18 * f, m),
        )
        # La jerarquía es estricta: el tema depende del subtema, nunca al revés.
        sub_a_tema: Dict[str, str] = {}
        for s, t in zip(subtemas, temas):
            sub_a_tema.setdefault(s, t)
        temas = [sub_a_tema[s] for s in subtemas]
        temas = armonizar(temas, id_asunto)
        temas = armonizar(temas, id_identidad)

    if tel.tasa_fallo_llm > 0.10:
        tel.avisar(
            f"El {tel.tasa_fallo_llm * 100:.0f}% de las llamadas al modelo falló. "
            "Revisa la clave de API y los límites de cuota antes de dar por buenos "
            "estos resultados."
        )

    _p(1.0, "Análisis completado")
    return ResultadoAnalisis(
        tonos=tonos,
        confianza_tono=confianzas,
        subtemas=subtemas,
        temas=temas,
        id_identidad=id_identidad,
        id_asunto=id_asunto,
        telemetria=tel,
    )

# ============ FIN DEL MOTOR DE PRECISIÓN ============


# ======================================
# Puente con el motor de precisión
# ======================================
def motor_precision_activo() -> bool:
    return bool(USAR_MOTOR_PRECISION and MOTOR_PRECISION_DISPONIBLE)


def _contexto_motor():
    """Inyecta en el motor las funciones de la app (embeddings y contador de costos).

    El motor no importa Streamlit: recibe aquí el acceso al caché de embeddings y
    al acumulador de tokens, de modo que el costo mostrado en pantalla sigue
    siendo el real y el caché se comparte con el resto del pipeline.
    """
    def _contar(tin: int, tout: int) -> None:
        st.session_state['tokens_input']  += tin
        st.session_state['tokens_output'] += tout

    return EngineContext(
        embed=get_embeddings_batch,
        modelo=OPENAI_MODEL_CLASIFICACION,
        on_tokens=_contar,
        max_concurrencia=CONCURRENT_REQUESTS,
        seed=7,
    )


async def analizar_con_motor(titulos, resumenes, bn, ba, pb=None,
                             calcular_tono=True, calcular_subtemas=True,
                             calcular_temas=True, num_temas_max=None):
    """Ejecuta el motor y deja la telemetría en sesión para el panel de calidad."""
    for k in ('telemetria_motor', 'avisos_motor', 'grupos_motor'):
        st.session_state.pop(k, None)
    ctx = _contexto_motor()

    def _progreso(frac, msg):
        if pb is not None:
            try:
                pb.progress(min(max(float(frac), 0.0), 1.0), text=str(msg)[:110])
            except Exception:
                pass

    resultado = await analizar_corpus(
        titulos=list(titulos),
        resumenes=list(resumenes),
        marca=bn,
        aliases=list(ba or []),
        ctx=ctx,
        progreso=_progreso,
        calcular_tono=calcular_tono,
        calcular_subtemas=calcular_subtemas,
        calcular_temas=calcular_temas,
        num_temas_max=num_temas_max,
    )
    st.session_state['telemetria_motor'] = resultado.telemetria.resumen()
    st.session_state['avisos_motor']     = list(resultado.telemetria.avisos)
    st.session_state['grupos_motor']     = {
        "noticias": len(resultado.tonos) or len(resultado.subtemas),
        "identidades": len(set(resultado.id_identidad)),
        "asuntos": len(set(resultado.id_asunto)),
        "subtemas": len({s for s in resultado.subtemas if s not in ("N/A",)}),
        "temas": len({t for t in resultado.temas if t not in ("N/A",)}),
    }
    return resultado


def render_panel_calidad():
    """Panel de diagnóstico: sin esto no hay forma de saber si un informe lleno
    de 'Neutro' refleja la realidad o una tanda de errores de API."""
    tel = st.session_state.get('telemetria_motor')
    if not tel:
        return
    for aviso in st.session_state.get('avisos_motor', []):
        st.warning(aviso)

    # Un embedding perdido significa una noticia que no pudo agruparse con nadie:
    # es un hueco en la clasificación, no un detalle técnico que ocultar.
    fallidos = st.session_state.get('embeddings_fallidos', 0)
    if fallidos:
        detalle = "; ".join(st.session_state.get('embeddings_errores', [])[:3])
        st.warning(
            f"{fallidos} noticia(s) se quedaron sin embedding y pudieron no agruparse "
            f"con sus similares. Primeros errores: {detalle}"
        )
    g = st.session_state.get('grupos_motor', {})
    if g:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Noticias", g.get("noticias", 0))
        c2.metric("Noticias únicas", g.get("identidades", 0),
                  help="Grupos de identidad: republicaciones de la misma noticia.")
        c3.metric("Asuntos", g.get("asuntos", 0),
                  help="Noticias distintas que cubren el mismo hecho.")
        c4.metric("Temas", g.get("temas", 0))
    with st.expander("Diagnóstico de calidad del análisis"):
        filas = [{"Indicador": k, "Valor": v} for k, v in tel.items()]
        st.dataframe(pd.DataFrame(filas), use_container_width=True, hide_index=True)


# ======================================
# Proceso principal
# ======================================
async def run_full_process_async(df_file, bn, ba, tpkl, epkl, mode, xlsx_bytes=None, cliente="", voceros="", enable_scraping=False):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    for _k in ('telemetria_motor', 'avisos_motor', 'grupos_motor',
               'embeddings_fallidos', 'embeddings_lotes_fallidos', 'embeddings_errores'):
        st.session_state.pop(_k, None)
    get_embedding_cache().clear()
    t0 = time.time()
    
    if "API" in mode:
        try:
            openai.api_key=st.secrets["OPENAI_API_KEY"]
            openai.aiosession.set(None)
        except (KeyError, FileNotFoundError):
            st.error("OPENAI_API_KEY no encontrado en st.secrets.")
            st.stop()
        except Exception as e:
            st.error(f"No se pudo inicializar el cliente de OpenAI: {type(e).__name__}: {e}")
            st.stop()
            
    with st.status("Paso 1 · Carga de Configuración y Dossier", expanded=True) as s:
        region_map, internet_map = load_config_from_sheets()

        wb_in = load_workbook(df_file, data_only=True)
        df_normalized = read_and_normalize_dossier(wb_in.active, region_map, internet_map)

        medios_sin_region = sorted(set(
            df_normalized.loc[df_normalized['Región'] == 'N/A', 'Medio']
            .astype(str).str.strip()
        ) - {'', 'nan', 'None'})
        if medios_sin_region:
            st.session_state["medios_sin_mapear"] = medios_sin_region
        
        rows_expanded = []
        for idx, row_series in df_normalized.iterrows():
            menciones = [m.strip() for m in str(row_series['Menciones - Empresa']).split(';') if m.strip()]
            if not menciones:
                row_dict = row_series.to_dict()
                row_dict['Menciones - Empresa'] = ""
                row_dict['original_index'] = idx
                row_dict['expanded_index'] = len(rows_expanded)
                row_dict['is_duplicate'] = False
                rows_expanded.append(row_dict)
            else:
                for m in menciones:
                    row_dict = row_series.to_dict()
                    row_dict['Menciones - Empresa'] = m
                    row_dict['original_index'] = idx
                    row_dict['expanded_index'] = len(rows_expanded)
                    row_dict['is_duplicate'] = False
                    rows_expanded.append(row_dict)

        km = {
            "idnoticia": "ID Noticia",
            "fecha": "Fecha",
            "hora": "Hora",
            "medio": "Medio",
            "tipodemedio": "Tipo de Medio",
            "seccion_programa": "Sección - Programa",
            "region": "Región",
            "titulo": "Título",
            "autor_conductor": "Autor - Conductor",
            "nro_pagina": "Nro. Pagina",
            "dimension": "Dimensión",
            "duracion_caracteres": "Duración - Nro. Caracteres",
            "cpe": "CPE",
            "tier": "Tier",
            "audiencia": "Audiencia",
            "tono": "Tono",
            "tonoiai": "Tono IA",
            "tema": "Tema",
            "subtema": "Subtema",
            "link_nota": "Link Nota",
            "resumen": "Resumen - Aclaracion",
            "link_streaming": "Link (Streaming - Imagen)",
            "menciones": "Menciones - Empresa",
            "idduplicada": "ID duplicada"
        }
        
        rows = detectar_duplicados_avanzado(rows_expanded, km)
        for row in rows:
            if row["is_duplicate"]:
                row["Tono IA"] = "Duplicada"
                row["Tema"] = "-"
                row["Subtema"] = "-"
                
        s.update(label="✓ Paso 1 completado", state="complete")
        
    with st.status("Paso 2 · Normalización", expanded=True) as s:
        s.update(label="✓ Paso 2 · Mapeos y normalizaciones aplicados", state="complete")
        
    gc.collect()
    ta = [r for r in rows if not r.get("is_duplicate")]
    
    if ta:
        df = pd.DataFrame(ta)
        df["_txt"] = df.apply(
            lambda r: texto_para_embedding(str(r.get(km["titulo"], "")), str(r.get(km["resumen"], ""))),
            axis=1
        )
        with st.status("Embeddings...", expanded=True) as s:
            _ = get_embeddings_batch(df["_txt"].tolist())
            s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")

        titulos_l   = df[km["titulo"]].fillna('').astype(str).tolist()
        resumenes_l = df[km["resumen"]].fillna('').astype(str).tolist()
        usar_motor  = motor_precision_activo() and "Solo Modelos PKL" not in mode and not tpkl and not epkl

        if usar_motor:
            # Ruta unificada: tono, subtema y tema salen de la MISMA jerarquía de
            # grupos, así dos noticias iguales no pueden recibir valores distintos.
            with st.status("Paso 3-4 · Análisis de precisión", expanded=True) as s:
                pb = st.progress(0)
                if "API" in mode or "Híbrido" in mode:
                    resultado = await analizar_con_motor(
                        titulos_l, resumenes_l, bn, ba, pb, num_temas_max=15
                    )
                    df[km["tonoiai"]] = resultado.tonos
                    df[km["subtema"]] = resultado.subtemas
                    df[km["tema"]]    = resultado.temas
                else:
                    df[km["tonoiai"]] = ["N/A"] * len(ta)
                    df[km["subtema"]] = ["N/A"] * len(ta)
                    df[km["tema"]]    = ["N/A"] * len(ta)
                s.update(label="✓ Paso 3-4 · Tono, subtema y tema coherentes", state="complete")
        else:
            with st.status("Paso 3 · Tono (Reputación)", expanded=True) as s:
                pb = st.progress(0)
                if ("PKL" in mode or tpkl) and tpkl:
                    res = analizar_tono_con_pkl(df["_txt"].tolist(), tpkl)
                    if res is None: st.stop()
                elif "API" in mode or "Híbrido" in mode:
                    res = await ClasificadorTono(bn, ba).procesar_lote_async(
                        df["_txt"], pb, df[km["resumen"]], df[km["titulo"]]
                    )
                else:
                    res = [{"tono": "N/A"}] * len(ta)
                df[km["tonoiai"]] = [r["tono"] for r in res]
                s.update(label="✓ Paso 3 · Tono (Reputación)", state="complete")

            with st.status("Paso 4 · Clasificación", expanded=True) as s:
                pb = st.progress(0)
                if "Solo Modelos PKL" in mode:
                    subtemas = ["N/A"] * len(ta)
                    temas    = ["N/A"] * len(ta)
                else:
                    subtemas = ClasificadorSubtema(bn, ba).procesar_lote(
                        df["_txt"], pb, df[km["resumen"]], df[km["titulo"]]
                    )
                    temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb)
                df[km["subtema"]] = subtemas
                if epkl:
                    tp = analizar_temas_con_pkl(df["_txt"].tolist(), epkl)
                    if tp: df[km["tema"]] = tp
                else:
                    df[km["tema"]] = temas
                s.update(label="✓ Paso 4 · Clasificación", state="complete")
            
        rm2 = df.set_index("expanded_index").to_dict("index")
        for idx, row in enumerate(rows):
            if not row.get("is_duplicate"):
                row.update(rm2.get(row.get("expanded_index"), {}))
                
    gc.collect()
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    
    with st.status("Paso 5 · Informe", expanded=True) as s:
        st.session_state["output_data"]     = generate_output_excel(rows, km)
        st.session_state["output_filename"] = f"Informe_IA_{bn.replace(' ', '_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        st.session_state["processing_complete"] = True
        st.session_state.update({
            "brand_name": bn, "brand_aliases": ba,
            "total_rows": len(rows), "unique_rows": len(ta), "duplicates": len(rows) - len(ta),
            "process_duration": f"{time.time() - t0:.0f}s",
            "process_cost": f"${ci + co + ce:.4f} USD",
            "cache_stats": get_embedding_cache().stats()
        })
        s.update(label=f"✓ Completado · {get_embedding_cache().stats()}", state="complete")

async def run_quick_async(df, tc, sc, bn, al):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    for _k in ('telemetria_motor', 'avisos_motor', 'grupos_motor',
               'embeddings_fallidos', 'embeddings_lotes_fallidos', 'embeddings_errores'):
        st.session_state.pop(_k, None)
    get_embedding_cache().clear()
    df['_txt'] = df.apply(lambda r: texto_para_embedding(str(r.get(tc, "")), str(r.get(sc, ""))), axis=1)
    with st.status("Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ {get_embedding_cache().stats()}", state="complete")

    if motor_precision_activo():
        with st.status("Análisis de precisión", expanded=True) as s:
            pb = st.progress(0)
            resultado = await analizar_con_motor(
                df[tc].fillna('').astype(str).tolist(),
                df[sc].fillna('').astype(str).tolist(),
                bn, al, pb, num_temas_max=15
            )
            df['Tono IA'] = resultado.tonos
            df['Subtema'] = resultado.subtemas
            df['Tema']    = resultado.temas
            s.update(label="✓ Tono, subtema y tema coherentes", state="complete")
    else:
        with st.status("Tono", expanded=True) as s:
            pb = st.progress(0)
            res = await ClasificadorTono(bn, al).procesar_lote_async(df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''))
            df['Tono IA'] = [r["tono"] for r in res]
            s.update(label="✓ Tono", state="complete")
        with st.status("Clasificación", expanded=True) as s:
            pb = st.progress(0)
            subtemas = ClasificadorSubtema(bn, al).procesar_lote(df["_txt"], pb, df[sc].fillna(''), df[tc].fillna(''))
            df['Subtema'] = subtemas
            temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb)
            df['Tema'] = temas
            s.update(label="✓ Clasificación", state="complete")
    df.drop(columns=['_txt'], inplace=True)
    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M
    st.session_state['quick_cost'] = f"${ci + co + ce:.4f} USD"
    return df

def gen_quick_excel(df):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as w:
        df.to_excel(w, index=False, sheet_name='Analisis')
    return buf.getvalue()

def render_quick_tab():
    st.markdown('<div class="sec-label">Análisis rápido</div>', unsafe_allow_html=True)
    if 'quick_result' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Completado</div>'
            '<div class="success-sub">Listo para descargar</div></div></div>',
            unsafe_allow_html=True
        )
        st.metric("Costo", st.session_state.get('quick_cost', "$0.00"))
        render_panel_calidad()
        st.dataframe(st.session_state.quick_result.head(10), use_container_width=True)
        st.download_button(
            "Descargar",
            data=gen_quick_excel(st.session_state.quick_result),
            file_name="Analisis_Rapido_IA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )
        if st.button("Nuevo análisis"):
            for k in ('quick_result', 'quick_df', 'quick_name', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return
    if 'quick_df' not in st.session_state:
        st.markdown("Sube un Excel con columnas de título y resumen.")
        f = st.file_uploader("Excel", type=["xlsx"], label_visibility="collapsed", key="qu")
        if f:
            try:
                st.session_state.quick_df   = pd.read_excel(f)
                st.session_state.quick_name = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error: {e}")
    else:
        st.success(f"**{st.session_state.quick_name}** cargado")
        with st.form("qf"):
            cols = st.session_state.quick_df.columns.tolist()
            c1, c2 = st.columns(2)
            tc = c1.selectbox("Col. título",  cols, 0)
            sc = c2.selectbox("Col. resumen", cols, 1 if len(cols) > 1 else 0)
            bn  = st.text_input("Marca",       placeholder="Ej: Bancolombia")
            bat = st.text_input("Alias (;)",   placeholder="Ej: Grupo Bancolombia;Ban")
            if st.form_submit_button("Analizar", use_container_width=True, type="primary"):
                if not bn:
                    st.error("Indica la marca.")
                else:
                    try:
                        openai.api_key = st.secrets["OPENAI_API_KEY"]
                        openai.aiosession.set(None)
                    except (KeyError, FileNotFoundError):
                        st.error("OPENAI_API_KEY no encontrada en st.secrets.")
                        st.stop()
                    except Exception as e:
                        st.error(f"No se pudo inicializar el cliente de OpenAI: {type(e).__name__}: {e}")
                        st.stop()
                    al = [a.strip() for a in bat.split(";") if a.strip()]
                    with st.spinner("Procesando..."):
                        st.session_state.quick_result = asyncio.run(
                            run_quick_async(st.session_state.quick_df.copy(), tc, sc, bn, al)
                        )
                    st.rerun()
        if st.button("Otro archivo"):
            for k in ('quick_df', 'quick_name', 'quick_result', 'quick_cost'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# EXCEL PERSONALIZADO (Mantiene formato original + 3 columnas al final)
# ======================================
async def run_custom_excel_async(file_bytes, tc, sc, bn, al, mode="API de OpenAI", tpkl=None, epkl=None):
    st.session_state.update({'tokens_input': 0, 'tokens_output': 0, 'tokens_embedding': 0})
    for _k in ('telemetria_motor', 'avisos_motor', 'grupos_motor',
               'embeddings_fallidos', 'embeddings_lotes_fallidos', 'embeddings_errores'):
        st.session_state.pop(_k, None)
    get_embedding_cache().clear()
    t0 = time.time()

    # Cargar archivo usando openpyxl para conservar estilos y formato original
    buf_in = io.BytesIO(file_bytes)
    wb = load_workbook(buf_in)
    ws = wb.active

    # Cargar DataFrame solo para extraer textos e índices
    buf_in.seek(0)
    df = pd.read_excel(buf_in)

    df['_txt'] = df.apply(
        lambda r: texto_para_embedding(str(r.get(tc, "")), str(r.get(sc, ""))),
        axis=1
    )

    with st.status("Paso 1 · Generando Embeddings...", expanded=True) as s:
        _ = get_embeddings_batch(df['_txt'].tolist())
        s.update(label=f"✓ Embeddings listos · {get_embedding_cache().stats()}", state="complete")

    usar_motor = (
        motor_precision_activo()
        and not tpkl and not epkl
        and "Solo Modelos PKL" not in mode
        and ("API" in mode or "Híbrido" in mode)
    )

    if usar_motor:
        with st.status("Paso 2 · Análisis de precisión (tono, subtema y tema)...", expanded=True) as s:
            pb = st.progress(0)
            resultado = await analizar_con_motor(
                df[tc].fillna('').astype(str).tolist(),
                df[sc].fillna('').astype(str).tolist(),
                bn, al, pb, num_temas_max=15
            )
            tonos    = resultado.tonos
            subtemas = resultado.subtemas
            temas    = resultado.temas
            df['Tono IA'] = tonos
            df['Subtema'] = subtemas
            df['Tema']    = temas
            s.update(label="✓ Tono, subtema y tema coherentes", state="complete")
    else:
        # --- PASO 2: TONO ---
        with st.status("Paso 2 · Evaluando Tono (Reputación)...", expanded=True) as s:
            pb = st.progress(0)
            if tpkl:
                # Si se subió PKL de Sentimiento/Tono, usarlo directamente
                res = analizar_tono_con_pkl(df["_txt"].tolist(), tpkl)
                if res is None: st.stop()
                tonos = [r["tono"] for r in res]
            elif "API" in mode or "Híbrido" in mode:
                res = await ClasificadorTono(bn, al).procesar_lote_async(
                    df["_txt"], pb, df[sc].fillna(''), df[tc].fillna('')
                )
                tonos = [r["tono"] for r in res]
            else:
                tonos = ["N/A"] * len(df)
            df['Tono IA'] = tonos
            s.update(label="✓ Tono IA evaluado", state="complete")

        # --- PASO 3: SUBTEMAS Y TEMAS ---
        with st.status("Paso 3 · Clasificando Subtemas y Temas...", expanded=True) as s:
            pb = st.progress(0)

            # Subtemas
            if "Solo Modelos PKL" in mode:
                subtemas = ["N/A"] * len(df)
            else:
                subtemas = ClasificadorSubtema(bn, al).procesar_lote(
                    df["_txt"], pb, df[sc].fillna(''), df[tc].fillna('')
                )

            # Temas
            if epkl:
                # Si se subió PKL de Temas, usar las predicciones directas del modelo
                tp = analizar_temas_con_pkl(df["_txt"].tolist(), epkl)
                if tp:
                    temas = tp
                else:
                    temas = ["N/A"] * len(df)
            elif "Solo Modelos PKL" in mode:
                temas = ["N/A"] * len(df)
            else:
                temas = consolidar_temas(subtemas, df["_txt"].tolist(), pb)

            df['Subtema'] = subtemas
            df['Tema']    = temas
            s.update(label="✓ Clasificación completada", state="complete")

    # Escribir las 3 columnas adicionales al final en la hoja openpyxl respetando el formato original
    max_col = ws.max_column
    col_tono    = max_col + 1
    col_tema    = max_col + 2
    col_subtema = max_col + 3

    # Encabezados en negrita
    font_bold = Font(bold=True)
    ws.cell(row=1, column=col_tono, value="Tono IA").font = font_bold
    ws.cell(row=1, column=col_tema, value="Tema").font = font_bold
    ws.cell(row=1, column=col_subtema, value="Subtema").font = font_bold

    # Asignar valores por fila manteniendo la coincidencia exacta
    for idx, row_data in df.iterrows():
        r = idx + 2
        ws.cell(row=r, column=col_tono, value=str(row_data['Tono IA']))
        ws.cell(row=r, column=col_tema, value=str(row_data['Tema']))
        ws.cell(row=r, column=col_subtema, value=str(row_data['Subtema']))

    buf_out = io.BytesIO()
    wb.save(buf_out)

    ci = (st.session_state['tokens_input']     / 1e6) * PRICE_INPUT_1M
    co = (st.session_state['tokens_output']    / 1e6) * PRICE_OUTPUT_1M
    ce = (st.session_state['tokens_embedding'] / 1e6) * PRICE_EMBEDDING_1M

    cost_str = f"${ci + co + ce:.4f} USD"
    time_str = f"{time.time() - t0:.0f}s"

    return buf_out.getvalue(), df, cost_str, time_str


def render_custom_excel_tab():
    st.markdown('<div class="sec-label">Análisis de Excel Personalizado</div>', unsafe_allow_html=True)
    st.caption("Sube cualquier archivo Excel (.xlsx). Al finalizar se descargarán los mismos datos y formato original con 3 nuevas columnas añadidas al final: **Tono IA**, **Tema** y **Subtema**.")

    if 'custom_result_bytes' in st.session_state:
        st.markdown(
            '<div class="success-banner"><div class="success-icon">✓</div>'
            '<div><div class="success-title">Análisis de Excel Finalizado</div>'
            '<div class="success-sub">Se han añadido las 3 columnas al final del Excel original manteniendo su formato.</div></div></div>',
            unsafe_allow_html=True
        )
        c1, c2 = st.columns(2)
        c1.metric("Costo estimado", st.session_state.get('custom_cost', "$0.00"))
        c2.metric("Tiempo de ejecución", st.session_state.get('custom_time', "0s"))

        render_panel_calidad()

        if 'custom_df_preview' in st.session_state:
            st.markdown("##### Vista previa del archivo (primeras filas con columnas añadidas):")
            st.dataframe(st.session_state.custom_df_preview.head(10), use_container_width=True)

        st.download_button(
            "⬇ Descargar Excel Actualizado",
            data=st.session_state.custom_result_bytes,
            file_name=f"Analisis_{st.session_state.get('custom_filename', 'Personalizado.xlsx')}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary"
        )

        if st.button("Nuevo análisis personalizado"):
            for k in ('custom_result_bytes', 'custom_df', 'custom_filename', 'custom_cost', 'custom_time', 'custom_df_preview'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()
        return

    if 'custom_df' not in st.session_state:
        f = st.file_uploader("Sube cualquier archivo Excel (.xlsx)", type=["xlsx"], key="custom_uploader")
        if f:
            try:
                bytes_data = f.getvalue()
                df_temp = pd.read_excel(io.BytesIO(bytes_data))
                st.session_state.custom_df       = df_temp
                st.session_state.custom_bytes    = bytes_data
                st.session_state.custom_filename = f.name
                st.rerun()
            except Exception as e:
                st.error(f"Error al leer el archivo Excel: {e}")
    else:
        st.success(f"📁 Archivo cargado: **{st.session_state.custom_filename}** ({len(st.session_state.custom_df)} filas)")

        cols = st.session_state.custom_df.columns.tolist()

        with st.form("custom_form"):
            st.markdown('<div class="sec-label">Selección de Columnas</div>', unsafe_allow_html=True)
            c_col1, c_col2 = st.columns(2)
            tc = c_col1.selectbox("Columna que contiene el TÍTULO", cols, index=0)
            sc = c_col2.selectbox("Columna que contiene el RESUMEN / CUERPO", cols, index=1 if len(cols) > 1 else 0)

            st.markdown('<div class="sec-label">Configuración del Análisis</div>', unsafe_allow_html=True)
            cl, cr = st.columns([3, 2])
            with cl:
                bn  = st.text_input("Marca principal", placeholder="Ej: Bancolombia", key="custom_bn")
                bat = st.text_input("Alias (separados por ;)", placeholder="Ej: Grupo Bancolombia;Ban", key="custom_ba")
            with cr:
                mode = st.radio(
                    "Modo de análisis",
                    ["API de OpenAI", "Híbrido (PKL + API)", "Solo Modelos PKL"],
                    index=0, key="custom_mode"
                )

            tpkl, epkl = None, None
            st.markdown('<div class="sec-label">Modelos PKL (Opcionales)</div>', unsafe_allow_html=True)
            p1, p2 = st.columns(2)
            tpkl = p1.file_uploader("Modelo Sentimiento / Tono (.pkl)", type=["pkl"], key="custom_tpkl")
            epkl = p2.file_uploader("Modelo Temas (.pkl)", type=["pkl"], key="custom_epkl")

            if st.form_submit_button("▶ Iniciar análisis personalizado", use_container_width=True, type="primary"):
                if not bn.strip():
                    st.error("Ingresa el nombre de la marca principal.")
                elif "Solo Modelos PKL" in mode and not (tpkl or epkl):
                    st.error("Seleccionaste 'Solo Modelos PKL', por favor adjunta al menos un archivo .pkl para continuar.")
                else:
                    if "API" in mode or "Híbrido" in mode:
                        try:
                            openai.api_key = st.secrets["OPENAI_API_KEY"]
                            openai.aiosession.set(None)
                        except (KeyError, FileNotFoundError):
                            st.error("OPENAI_API_KEY no encontrada en st.secrets.")
                            st.stop()
                        except Exception as e:
                            st.error(f"No se pudo inicializar el cliente de OpenAI: {type(e).__name__}: {e}")
                            st.stop()

                    al = [a.strip() for a in bat.split(";") if a.strip()]

                    with st.spinner("Procesando Excel personalizado..."):
                        res_bytes, res_df, cost_str, time_str = asyncio.run(
                            run_custom_excel_async(
                                st.session_state.custom_bytes,
                                tc, sc, bn, al,
                                mode=mode, tpkl=tpkl, epkl=epkl
                            )
                        )

                        st.session_state.custom_result_bytes = res_bytes
                        st.session_state.custom_df_preview   = res_df
                        st.session_state.custom_cost         = cost_str
                        st.session_state.custom_time         = time_str
                        st.rerun()

        if st.button("Subir otro archivo Excel"):
            for k in ('custom_df', 'custom_bytes', 'custom_filename', 'custom_result_bytes', 'custom_cost', 'custom_time', 'custom_df_preview'):
                if k in st.session_state:
                    del st.session_state[k]
            st.rerun()


# ======================================
# Main
# ======================================
def main():
    load_custom_css()
    if not check_password(): return

    st.markdown("""
    <div class="app-header">
        <div class="app-header-icon">◈</div>
        <div class="app-header-text">
            <div class="app-header-title">Análisis de Noticias - API</div>
            <div class="app-header-version">v18.2 · 😼 Realizado por Johnathan Cortés 🕵️‍♂️ </div>
        </div>
        <div class="app-header-badge">IA</div>
    </div>""", unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["Análisis Completo", "Análisis Rápido", "Excel Personalizado"])

    with tab1:
        if not st.session_state.get("processing_complete", False):
            col_cfg1, col_cfg2 = st.columns([4, 1])
            with col_cfg1:
                st.markdown(
                    '<span class="config-badge">⚙ Configuración: Google Sheets (Regiones / Internet)</span>',
                    unsafe_allow_html=True
                )
            with col_cfg2:
                if st.button("↻ Refrescar config", use_container_width=True):
                    refresh_config_cache()
                    st.success("Config recargada")

            st.markdown('<div class="sec-label">Configuración</div>', unsafe_allow_html=True)
            cl, cr = st.columns([3, 2])
            with cl:
                bn  = st.text_input("Marca principal", placeholder="Ej: Bancolombia", key="bn")
                bat = st.text_input("Alias (separados por ;)", placeholder="Ej: Grupo Bancolombia;Ban", key="ba")
            with cr:
                mode = st.radio(
                    "Modo de análisis",
                    ["API de OpenAI", "Híbrido (PKL + API)", "Solo Modelos PKL"],
                    index=0, key="mode"
                )

            tpkl, epkl = None, None
            if "PKL" in mode:
                st.markdown('<div class="sec-label">Modelos PKL</div>', unsafe_allow_html=True)
                p1, p2 = st.columns(2)
                tpkl = p1.file_uploader(
                    "Modelo de Sentimiento (.pkl)", type=["pkl"], key="tpkl",
                    help="Pipeline sklearn para clasificar tono: -1/0/1 o Negativo/Neutro/Positivo"
                )
                epkl = p2.file_uploader(
                    "Modelo de Temas (.pkl)", type=["pkl"], key="epkl",
                    help="Pipeline sklearn para clasificar temas"
                )

            with st.form("main_form"):
                st.markdown('<div class="sec-label">Archivo de entrada</div>', unsafe_allow_html=True)
                st.markdown("""
                <div class="upload-zone" style="grid-template-columns:1fr">
                    <div class="upload-zone-card">
                        <div class="upload-zone-icon uz-dossier">📋</div>
                        <div class="upload-zone-text">
                            <div class="upload-zone-title">Dossier</div>
                            <div class="upload-zone-desc">Sube las noticias en el nuevo formato .xlsx a analizar</div>
                        </div>
                    </div>
                </div>""", unsafe_allow_html=True)
                f1 = st.file_uploader("Dossier", type=["xlsx"], label_visibility="collapsed", key="f1")

                # Los parámetros mostrados son los que el motor usa de verdad;
                # antes se imprimían las constantes del pipeline antiguo, que ya
                # no intervienen cuando el motor de precisión está activo.
                _cfg_ui = ConfigPrecision()
                st.markdown(
                    f'<div class="cluster-info">'
                    f'<b>Motor de precisión</b> · Identidad={_cfg_ui.sim_identidad} '
                    f'· Asunto={_cfg_ui.sim_asunto} (piso {_cfg_ui.piso_absoluto_asunto}) '
                    f'· Tema={_cfg_ui.sim_tema} (piso {_cfg_ui.piso_absoluto_tema}) · '
                    f'<b>Jaccard entidades={_cfg_ui.jaccard_entidades_min}</b> '
                    f'· Overlap={_cfg_ui.overlap_tokens_min} '
                    f'· Unif. etiquetas={_cfg_ui.sim_unificar_etiquetas} '
                    f'· Especificidad=+{_cfg_ui.margen_especificidad} '
                    f'· Confianza tono≥{_cfg_ui.confianza_minima_tono} '
                    f'(se endurecen automáticamente en corpus pequeños)'
                    f'</div>',
                    unsafe_allow_html=True
                )

                if st.form_submit_button("▶ Iniciar análisis", use_container_width=True, type="primary"):
                    if not all([f1, bn.strip()]):
                        st.error("Completa todos los campos.")
                    else:
                        al = [a.strip() for a in bat.split(";") if a.strip()]
                        cur_mode = st.session_state.get("mode", "API de OpenAI")
                        cur_tpkl = st.session_state.get("tpkl")
                        cur_epkl = st.session_state.get("epkl")
                        asyncio.run(run_full_process_async(f1, bn, al, cur_tpkl, cur_epkl, cur_mode,
                                                         xlsx_bytes=None, cliente="", voceros="",
                                                         enable_scraping=False))
                        st.rerun()
        else:
            total = st.session_state.total_rows
            uniq  = st.session_state.unique_rows
            dups  = st.session_state.duplicates
            dur   = st.session_state.process_duration
            cost  = st.session_state.get("process_cost", "$0.00")
            st.markdown(
                '<div class="success-banner"><div class="success-icon">✓</div>'
                '<div><div class="success-title">Análisis completado</div>'
                '<div class="success-sub">Informe listo para descargar</div></div></div>',
                unsafe_allow_html=True
            )

            medios_sin_mapear = st.session_state.get("medios_sin_mapear")
            if medios_sin_mapear:
                st.warning(
                    "⚠️ Los siguientes medios no tienen región asignada en el Sheets de "
                    f"'Regiones' (quedaron como N/A): {', '.join(medios_sin_mapear)}. "
                    "Agrégalos en el Google Sheets para que se mapeen automáticamente la próxima vez."
                )

            st.markdown(f"""
            <div class="metrics-grid">
              <div class="metric-card m-total"><div class="metric-val" style="color:var(--text)">{total}</div><div class="metric-lbl">Total</div></div>
              <div class="metric-card m-unique"><div class="metric-val" style="color:var(--green)">{uniq}</div><div class="metric-lbl">Únicas</div></div>
              <div class="metric-card m-dup"><div class="metric-val" style="color:var(--amber)">{dups}</div><div class="metric-lbl">Duplicados</div></div>
              <div class="metric-card m-time"><div class="metric-val" style="color:var(--blue)">{dur}</div><div class="metric-lbl">Tiempo</div></div>
              <div class="metric-card m-cost"><div class="metric-val" style="color:var(--accent)">{cost}</div><div class="metric-lbl">Costo</div></div>
            </div>""", unsafe_allow_html=True)
            if 'cache_stats' in st.session_state: st.caption(f"📊 {st.session_state['cache_stats']}")
            render_panel_calidad()
            c1, c2 = st.columns(2)
            c1.download_button(
                "⬇ Descargar informe",
                data=st.session_state.output_data,
                file_name=st.session_state.output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            if c2.button("Nuevo análisis", use_container_width=True):
                pwd = st.session_state.get("password_correct")
                st.session_state.clear()
                st.session_state.password_correct = pwd
                st.rerun()

    with tab2:
        render_quick_tab()

    with tab3:
        render_custom_excel_tab()

    st.markdown(
        '<div class="footer">v18.2 · Análisis de Noticias con IA · Johnathan Cortés ©</div>',
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
