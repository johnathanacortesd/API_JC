"""
Core pipeline helpers for API_JC.

This module is intentionally Streamlit-free so the working product can be
repaired and regression-tested without rewriting the UI, PKL branches or
dossier controls. app.py remains the product surface.
"""
from __future__ import annotations

import datetime
import hashlib
import html
import io
import json
import re
import threading
import time
from collections import Counter, defaultdict
from concurrent.futures import ThreadPoolExecutor
from copy import deepcopy
from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from unidecode import unidecode

try:
    from sklearn.metrics.pairwise import cosine_similarity
except Exception:  # pragma: no cover
    cosine_similarity = None


_COUNTER_LOCK = threading.Lock()

OPENAI_MODEL_EMBEDDING = "text-embedding-3-small"
OPENAI_MODEL_CLASIFICACION = "gpt-4.1-nano-2025-04-14"
CHAT_BATCH_SIZE = 30
CHAT_PARALLEL_BATCHES = 6
# A 30-item classification request with ~1 KB of context per item plus a
# 100-item embedding batch can legitimately take 40–90 s server-side.
# 25 s made both paths time out and retry (×3 with backoff), which is what
# made large dossiers feel stuck.
REQUEST_TIMEOUT_S = 120
MAX_RETRIES = 2
MAX_BUCKET = 48
MAX_EXACT_BUCKET = 20
MAX_PAIRS_TOTAL = 12000
MAX_CMP_CHARS = 400
SIMILARITY_THRESHOLD_TITULOS = 0.84
SIMILARITY_THRESHOLD_TITULOS_BCAST = 0.86
SIMILARITY_THRESHOLD_RESUMEN = 0.92
SIMILARITY_THRESHOLD_SEMANTIC = 0.88
SIMILARITY_SEMANTIC_ALONE = 0.93
MIN_OVERLAP_GRUPO = 0.30
JACCARD_TITULO_GRUPO = 0.70
EMBED_CANDIDATE_MIN = 0.82
EMBED_NEIGHBORS = 16
CUERPO_SCAN_CHARS = 8000
MAX_PALABRAS_SUBTEMA = 6
MIN_PALABRAS_SUBTEMA = 3
MAX_PALABRAS_TEMA = 5

OUTPUT_COLUMNS = [
    "ID Noticia",
    "Fecha",
    "Hora",
    "Medio",
    "Tipo de Medio",
    "Sección - Programa",
    "Región",
    "Título",
    "Tono IA",
    "Tema",
    "Subtema",
    "Link Nota",
    "Resumen - Aclaracion",
    "Link (Streaming - Imagen)",
    "Menciones - Empresa",
    "ID duplicada",
    "Cuerpo Completo",
    "Contexto analizado",
    "Coincidencia marca",
    "Origen coincidencia",
    "Tono",
    "Grupo noticia",
]

KEYMAP = {
    "idnoticia": "ID Noticia",
    "fecha": "Fecha",
    "hora": "Hora",
    "medio": "Medio",
    "tipodemedio": "Tipo de Medio",
    "seccion_programa": "Sección - Programa",
    "region": "Región",
    "titulo": "Título",
    "tonoiai": "Tono IA",
    "tema": "Tema",
    "subtema": "Subtema",
    "link_nota": "Link Nota",
    "resumen": "Resumen - Aclaracion",
    "link_streaming": "Link (Streaming - Imagen)",
    "menciones": "Menciones - Empresa",
    "idduplicada": "ID duplicada",
    "cuerpo": "Cuerpo Completo",
    "contexto": "Contexto analizado",
    "coincidencia": "Coincidencia marca",
    "origen": "Origen coincidencia",
    "tono": "Tono",
    "grupo": "Grupo noticia",
}

TRACKING_PARAMS = {
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content",
    "utm_id", "utm_cid", "utm_reader", "utm_name", "utm_social",
    "fbclid", "gclid", "gclsrc", "dclid", "msclkid", "twclid", "yclid",
    "mc_cid", "mc_eid", "igshid", "_ga", "_gl", "ref", "ref_src",
    "feature", "ncid", "cmpid", "s",
}

STOPWORDS_ES = set("""
a ante bajo cabe con contra de desde durante en entre hacia hasta mediante
para por segun sin so sobre tras y o u e la el los las un una unos unas lo
al del se su sus le les mi mis tu tus nuestro nuestros vuestra vuestras este
esta estos estas ese esa esos esas aquel aquella aquellos aquellas que cual
cuales quien quienes cuyo cuya cuyos cuyas como cuando donde cual es son fue
fueron era eran sera seran seria serian he ha han habia hay hubo habra
habria estoy estan estaba estaban estamos estar estare estaria
estuvieron estarian estuvo asi ya mas menos tan tanto cada muy todo toda todos
todas ser haber hacer tener poder deber ir dar ver saber querer llegar pasar
encontrar creer decir poner salir volver seguir llevar sentir cambiar
""".split())

_TRAILING_INCOMPLETE = {
    "de", "del", "la", "el", "los", "las", "un", "una", "unos", "unas", "al",
    "su", "sus", "en", "con", "sin", "por", "para", "sobre", "ante", "bajo",
    "contra", "desde", "entre", "hacia", "hasta", "mediante", "tras", "y", "o",
    "u", "e", "lo", "que", "se", "como", "donde", "cuando", "cual", "cuyo",
    "si", "cómo", "como", "qué", "cuál", "dentro", "fuera", "encima", "debajo",
    "segun", "según", "via", "vs", "versus", "a",
}

_CHANNEL_PREFIXES = {
    "video", "en vivo", "vivo", "streaming", "audio", "podcast", "radio",
    "tv", "television", "televisión",
}

_QUESTION_TAILS = {
    "como", "cómo", "si", "que", "qué", "cual", "cuál", "cuando", "cuándo",
    "donde", "dónde", "por", "para",
}

_CONJUGATED_TAILS = {
    "ascienden", "asciende", "saben", "sabe", "destacan", "destaca",
    "confirman", "confirma", "anuncian", "anuncia", "presentan", "presenta",
    "lanzan", "lanza", "reportan", "reporta", "informan", "informa",
    "advierten", "advierte", "revelan", "revela", "ocurren", "ocurre",
    "registran", "registra", "mueren", "muere", "resultan", "resulta",
    "quedan", "queda", "llegan", "llega", "inician", "inicia",
    "tendra", "tendran", "tiene", "tienen", "sera", "seran", "esta", "estan",
    "hay", "va", "van", "hace", "hacen", "dice", "dicen", "pide", "piden",
    "busca", "buscan", "deja", "dejan", "cae", "caen", "sube", "suben",
    "crece", "crecen", "gana", "ganan", "pierde", "pierden", "abre", "abren",
    "cierra", "cierran", "entrega", "entregan", "recibe", "reciben",
    "vive", "viven", "cumple", "cumplen", "ofrece", "ofrecen", "logra", "logran",
    "quiere", "quieren", "puede", "pueden", "debe", "deben", "sigue", "siguen",
    "vuelve", "vuelven", "sale", "salen", "pasa", "pasan", "pone", "ponen",
    "firma", "firman", "aprueba", "aprueban", "rechaza", "rechazan",
    "denuncia", "denuncian", "alerta", "alertan", "celebra", "celebran",
    "invierte", "invierten", "impulsa", "impulsan", "lidera", "lideran",
    "explica", "explican", "analiza", "analizan", "promueve", "promueven",
    "realiza", "realizan", "suspende", "suspenden", "activa", "activan",
    "refuerza", "refuerzan", "amplia", "amplian", "reduce", "reducen",
    "asegura", "aseguran", "afirma", "afirman", "inaugura", "inauguran",
    "es", "son", "fue", "fueron", "era", "eran", "sido", "estara", "estaran",
    "habra", "podra", "podran", "debera", "deberan", "hara", "haran",
    "dara", "daran", "ira", "iran", "vendra", "vendran", "llegara", "llegaran",
}

_SUFIJO_VERBO_RE = re.compile(
    r"(?:ar[áa]n?|er[áa]n?|ir[áa]n?|r[áa]s|aron|ieron|aban|[áa]bamos|[íi]amos|"
    r"iendo|yendo|[óo]|amos|emos|imos)$"
)
_NOMBRES_TERMINADOS_O = {
    "balance", "gobierno", "proyecto", "programa", "convenio", "acuerdo", "aumento",
    "cierre", "cambio", "riesgo", "impacto", "reto", "hecho", "caso", "costo",
    "presupuesto", "desarrollo", "manejo", "apoyo", "respaldo", "rechazo", "anuncio",
    "lanzamiento", "reconocimiento", "premio", "logro", "trabajo", "empleo", "comercio",
    "consumo", "transporto", "plazo", "estado", "mercado", "operativo", "dispositivo",
    "estudio", "informe", "resultado", "registro", "aniversario", "concurso",
    "torneo", "campeonato", "partido", "público", "privado", "sector", "puesto",
    "terremoto", "sismo", "incendio", "secuestro", "hurto", "robo", "homicidio",
    "acueducto", "alcantarillado", "aeropuerto", "puerto", "metro", "recorrido",
    "ingreso", "descuento", "subsidio", "bono", "seguro", "crédito", "credito",
    "servicio", "comercio", "negocio", "consorcio", "municipio", "departamento",
    "territorio", "barrio", "centro", "colegio", "liceo", "instituto", "museo", "teatro",
    "concierto", "festival", "evento", "encuentro", "foro", "congreso", "diálogo", "dialogo",
    "conflicto", "ataque", "atentado", "operativo", "hallazgo", "rescate", "cuerpo",
    "gasto", "ahorro", "pago", "cobro", "recaudo", "trámite", "tramite", "periodo",
    "modelo", "diseño", "diseno", "producto", "alimento", "medicamento", "tratamiento",
    "nombramiento", "reglamento", "documento", "monumento", "instrumento",
}


def _es_verbo_conjugado(palabra: str) -> bool:
    """Heuristic: known conjugated forms or verbal suffixes (tendrá, anunció,
    ascienden, cayeron). Common nouns ending in -o are whitelisted."""
    raw = str(palabra).lower().strip(".,;:!?¿¡\"'")
    w = unidecode(raw)
    if not w:
        return False
    if w in _CONJUGATED_TAILS:
        return True
    # Accented preterite / future (anunció, cayó, tendrá) is decided on the
    # raw form first: 'anunció' must not be whitelisted as the noun 'anuncio'.
    if re.search(r"(?:ó|á)$", raw) and len(raw) >= 4 and raw not in {"está", "acá", "allá", "quizá", "ojalá", "sofá", "papá", "mamá", "bogotá", "cúcuta", "panamá", "canadá"}:
        return True
    if w in _NOMBRES_TERMINADOS_O or w in STOPWORDS_ES:
        return False
    if len(w) <= 3:
        return False
    if re.search(r"(?:ar|er|ir)[áa]n?$", w) and len(w) >= 6:
        return True
    if re.search(r"(?:aron|ieron|eron|aban|iendo|yendo)$", w):
        return True
    return False


def _parece_adjetivo(palabra: str) -> bool:
    w = unidecode(str(palabra).lower().strip(".,;:"))
    return bool(re.search(
        r"(?:al|ales|ar|ares|ico|ica|icos|icas|ivo|iva|ivos|ivas|oso|osa|osos|osas|"
        r"ble|bles|ante|antes|ente|entes|ario|aria|arios|arias|ense|enses|eno|ena|"
        r"ano|ana|anos|anas|ial|iales|orio|oria|orios|orias|ista|istas|il|iles|"
        r"estre|estres|uro|ura|uros|uras|ero|era|eros|eras)$", w))

_NEXOS = {
    "de", "del", "para", "sobre", "en", "con", "por", "ante", "hacia",
    "entre", "sin", "al", "las", "los",
}

_ACCIONES_OPUESTAS = [
    ({"aprobacion", "aprueba", "apoyo", "acuerdo", "aval", "respaldo"},
     {"rechazo", "rechaza", "desacuerdo", "oposicion", "critica"}),
    ({"aumento", "crecimiento", "alza", "subida", "incremento"},
     {"caida", "reduccion", "baja", "disminucion", "descenso"}),
    ({"apertura", "inauguracion", "inicio", "lanzamiento", "estreno"},
     {"cierre", "suspension", "fin", "clausura", "cancelacion"}),
    ({"exito", "logro", "triunfo", "premio", "reconocimiento"},
     {"fracaso", "derrota", "problema", "crisis", "sancion"}),
    ({"demanda", "denuncia", "investigacion", "sancion", "multa"},
     {"absolucion", "archivo", "exoneracion"}),
]

_TOKENS_DEBILES = STOPWORDS_ES | {
    "noticia", "noticias", "informe", "informacion", "comunicado", "anuncio",
    "colombia", "pais", "nacional", "regional", "local", "sector",
    "empresa", "empresas", "nuevo", "nueva", "plan", "programa", "proyecto",
    "actividad", "gestion", "tema", "caso",
}


# ---------------------------------------------------------------------------
# Instrumentation
# ---------------------------------------------------------------------------

class ComparisonCounter:
    """Counts blocked pairwise comparisons (must stay << n²)."""

    def __init__(self) -> None:
        self.n = 0

    def add(self, k: int = 1) -> None:
        self.n += int(k)

    def reset(self) -> None:
        self.n = 0


class CallCounter:
    def __init__(self) -> None:
        self.chat = 0
        self.embed = 0
        self.embed_items = 0
        self.chat_items = 0

    def reset(self) -> None:
        self.chat = self.embed = self.embed_items = self.chat_items = 0


class ProgressTracker:
    STAGES = (
        "Limpieza", "Duplicados", "Contexto", "Embedding único",
        "Agrupación", "Tono", "Tema/Subtema", "Excel",
    )

    def __init__(self, on_stage: Optional[Callable[[Dict[str, Any]], None]] = None) -> None:
        self.t0 = time.time()
        self._t_stage = self.t0
        self.events: List[Dict[str, Any]] = []
        self.calls = CallCounter()
        self.comparisons = ComparisonCounter()
        self._current = None
        self.on_stage = on_stage

    def stage(self, name: str, extra: str = "") -> Dict[str, Any]:
        now = time.time()
        stage_s = now - self._t_stage
        self._t_stage = now
        elapsed = now - self.t0
        ev = {
            "stage": name,
            "elapsed_s": round(elapsed, 3),
            "stage_s": round(stage_s, 3),
            "chat_calls": self.calls.chat,
            "embed_calls": self.calls.embed,
            "comparisons": self.comparisons.n,
            "extra": extra,
            "label": (
                f"{name} · {stage_s:.1f}s · total {elapsed:.1f}s · "
                f"chat={self.calls.chat} · emb={self.calls.embed}"
                + (f" · {extra}" if extra else "")
            ),
        }
        ev["index"] = len(self.events) + 1
        ev["total"] = len(self.STAGES)
        self.events.append(ev)
        self._current = ev
        if self.on_stage is not None:
            try:
                self.on_stage(ev)
            except Exception:
                pass
        return ev

    def summary(self) -> str:
        return " | ".join(e["label"] for e in self.events)


# ---------------------------------------------------------------------------
# BUSCARV (VLOOKUP) — do not rename or drop this branch
# ---------------------------------------------------------------------------

def construir_mapa_buscarv(pares: Iterable) -> Dict[str, str]:
    """Excel BUSCARV / VLOOKUP: primera columna = clave, segunda = valor."""
    mapping: Dict[str, str] = {}
    if pares is None:
        return mapping
    if hasattr(pares, "iloc"):
        df = pares.dropna(how="all")
        keys = df.iloc[:, 0].astype(str).str.lower().str.strip()
        vals = df.iloc[:, 1].astype(str)
        for k, v in zip(keys, vals):
            if k and k != "nan":
                mapping[k] = v
        return mapping
    for item in pares:
        if isinstance(item, dict):
            k = str(item.get("clave", item.get("key", ""))).lower().strip()
            v = item.get("valor", item.get("value", ""))
        else:
            if len(item) < 2:
                continue
            k = str(item[0]).lower().strip()
            v = item[1]
        if k and k != "nan":
            mapping[k] = "" if v is None else str(v)
    return mapping


def aplicar_buscarv(clave: Any, mapping: Dict[str, str], si_no_encuentra: Any = "N/A") -> Any:
    """Exact API_JC lookup: lower+strip key, return mapped value or default."""
    if clave is None:
        return si_no_encuentra
    k = str(clave).lower().strip()
    if not k or k == "nan":
        return si_no_encuentra
    return mapping.get(k, si_no_encuentra)


def aplicar_buscarv_dossier(medios: Sequence[Any], region_map: Dict[str, str],
                            internet_map: Dict[str, str],
                            tipos: Optional[Sequence[Any]] = None
                            ) -> Tuple[List[Any], List[Any]]:
    """
    Same two BUSCARV branches the dossier already uses:
    - Región = BUSCARV(Medio, sheet Regiones)
    - Medio (solo Internet) = BUSCARV(Medio, sheet Internet)
    """
    regiones = [aplicar_buscarv(m, region_map, "N/A") for m in medios]
    medios_out = list(medios)
    if tipos is not None:
        for i, (medio, tipo) in enumerate(zip(medios, tipos)):
            if normalizar_tipo_medio(tipo) == "Internet":
                mapped = aplicar_buscarv(medio, internet_map, None)
                if mapped is not None:
                    medios_out[i] = mapped
    return regiones, medios_out


# ---------------------------------------------------------------------------
# Titles, URLs, hyperlinks
# ---------------------------------------------------------------------------

def titulo_original(valor: Any) -> Any:
    """Output Título must stay the original cell; never split on | : - HTML."""
    if isinstance(valor, dict):
        return valor.get("value", valor)
    return valor


_PREFIJO_CANAL_RE = re.compile(
    r"^\s*(?:video|en vivo|vivo|audio|streaming|podcast|galer[ií]a|fotos?|"
    r"opini[oó]n|editorial|exclusivo|urgente|[uú]ltima hora|breaking|columna)\s*[\|:\-–—]\s*",
    re.I,
)


def normalize_title_for_comparison(title: Any) -> str:
    """Internal-only copy for matching: strips HTML, channel prefixes
    ('Video |', 'En vivo:') and trailing source segments ('| El País',
    '- Semana'). The output Título is never touched by this."""
    if not isinstance(title, str):
        title = "" if title is None else str(title)
    t = re.sub(r"<[^>]+>", " ", title)
    t = html.unescape(t) if hasattr(html, "unescape") else t
    for _ in range(2):
        t = _PREFIJO_CANAL_RE.sub("", t)
    # Trailing " | Fuente" / " - Fuente" / " – Fuente" of up to 5 words.
    t = re.sub(r"\s+[\|–—-]\s+(?:[^\|–—-]\S*\s*){1,5}$", "", t).strip()
    cleaned = unidecode(t)
    return re.sub(r"\W+", " ", cleaned).lower().strip()


def extraer_hipervinculo_celda(cell: Any) -> Optional[str]:
    """Read the embedded Excel hyperlink target (pandas drops this)."""
    if cell is None:
        return None
    hyper = getattr(cell, "hyperlink", None)
    if hyper is not None:
        target = getattr(hyper, "target", None) or getattr(hyper, "location", None)
        if target:
            return str(target)
    value = getattr(cell, "value", cell)
    if isinstance(value, str) and "HYPERLINK" in value.upper():
        m = re.search(r'HYPERLINK\(\s*"([^"]+)"', value, flags=re.I)
        if m:
            return m.group(1)
    return None


def valor_con_hipervinculo(cell: Any) -> Any:
    url = extraer_hipervinculo_celda(cell)
    value = getattr(cell, "value", cell)
    if url:
        return {"value": value if value not in (None, "") else "Link", "url": url}
    return value


def url_de_celda_link(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, dict):
        return str(val.get("url") or "")
    if isinstance(val, str) and val.startswith("http"):
        return val
    return ""


def normalize_url(url: Any) -> str:
    """Strip tracking/fragment/trailing slash and lowercase host; keep path."""
    if not url:
        return ""
    raw = str(url).strip()
    if not raw:
        return ""
    if "://" not in raw:
        raw = "http://" + raw
    try:
        p = urlparse(raw)
    except Exception:
        return raw.strip().lower().rstrip("/")
    host = (p.hostname or "").lower()
    if host.startswith("www."):
        host = host[4:]
    path = p.path or ""
    if path != "/" and path.endswith("/"):
        path = path.rstrip("/")
    q = [(k, v) for k, v in parse_qsl(p.query, keep_blank_values=True)
         if k.lower() not in TRACKING_PARAMS]
    query = urlencode(q, doseq=True)
    rebuilt = urlunparse((p.scheme.lower() or "http", host, path, "", query, ""))
    rebuilt = re.sub(r"^https?://", "", rebuilt)
    return rebuilt.rstrip("/")


def normalizar_tipo_medio(tipo_raw: Any) -> str:
    if not isinstance(tipo_raw, str):
        tipo_raw = str(tipo_raw or "")
    t = unidecode(tipo_raw.strip().lower())
    return {
        "online": "Internet", "internet": "Internet",
        "diario": "Prensa", "prensa": "Prensa",
        "am": "Radio", "fm": "Radio", "radio": "Radio",
        "aire": "Televisión", "cable": "Televisión", "tv": "Televisión",
        "television": "Televisión", "televisión": "Televisión",
        "revista": "Revistas", "revistas": "Revistas",
    }.get(t, tipo_raw.strip().title() or "Otro")


def normalizar_fecha(val: Any) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    if hasattr(val, "strftime"):
        try:
            return val.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(val).strip()
    if not s or s.lower() in {"nan", "nat", "none"}:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except Exception:
            continue
    try:
        import pandas as pd
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if ts is not None and not (hasattr(ts, "isna") and ts.isna()):
            return ts.strftime("%Y-%m-%d")
    except Exception:
        pass
    return s[:10]


def normalizar_hora(val: Any) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    if hasattr(val, "strftime"):
        try:
            return val.strftime("%H:%M")
        except Exception:
            pass
    s = str(val).strip()
    if not s or s.lower() in {"nan", "none"}:
        return ""
    m = re.search(r"(\d{1,2})[:.](\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s


def _norm_text(texto: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", unidecode(str(texto or "").lower()))).strip()


def _tokens_distintivos(texto: str, min_len: int = 4) -> set:
    return {
        t for t in _norm_text(texto).split()
        if len(t) >= min_len and t not in _TOKENS_DEBILES and not t.isdigit()
    }


_SUFIJOS_ES = (
    "aciones", "iciones", "amiento", "imiento", "adores", "edores", "idores",
    "mente", "aron", "ieron", "aban", "ian", "ando", "iendo", "ados", "idos",
    "adas", "idas", "ados", "acion", "icion", "ador", "edor", "idor", "aron",
    "ara", "era", "ira", "aran", "eran", "iran", "ado", "ida", "ido", "ada",
    "an", "en", "es", "os", "as", "ar", "er", "ir", "o", "a", "e", "s",
)


def _stem_es(tok: str) -> str:
    """Light Spanish stemmer for matching only (destacan/destacaron → destac)."""
    t = unidecode(str(tok).lower())
    if len(t) <= 4:
        return t
    for suf in _SUFIJOS_ES:
        if t.endswith(suf) and len(t) - len(suf) >= 4:
            return t[: len(t) - len(suf)]
    return t


def _stems(tokens: Iterable[str]) -> set:
    return {_stem_es(t) for t in tokens}


def _overlap_distintivo(a: str, b: str) -> float:
    ta, tb = _tokens_distintivos(a), _tokens_distintivos(b)
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / max(1, min(len(ta), len(tb)))


def _hay_conflicto_accion(a: str, b: str) -> bool:
    ta, tb = _tokens_distintivos(a, min_len=3), _tokens_distintivos(b, min_len=3)
    for ga, gb in _ACCIONES_OPUESTAS:
        if (ta & ga and tb & gb) or (ta & gb and tb & ga):
            return True
    return False


def _ratio(a: str, b: str, threshold: float = 0.0) -> float:
    """Character SequenceMatcher for SHORT strings (titles), capped in length.

    Reserved for tiny n (tema consolidation) and gray-zone title pairs.
    Never call this on resúmenes or on every grouping candidate.
    """
    from difflib import SequenceMatcher
    if not a or not b:
        return 0.0
    a, b = a[:MAX_CMP_CHARS], b[:MAX_CMP_CHARS]
    sm = SequenceMatcher(None, a, b)
    if threshold > 0 and sm.real_quick_ratio() < threshold:
        return 0.0
    if threshold > 0 and sm.quick_ratio() < threshold:
        return 0.0
    return sm.ratio()


MAX_CMP_WORDS = 80


def _jaccard_sets(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    inter = len(a & b)
    if not inter:
        return 0.0
    return inter / len(a | b)


def _overlap_sets(a: set, b: set) -> float:
    if not a or not b:
        return 0.0
    return len(a & b) / min(len(a), len(b))


def _ratio_palabras(a_words: Sequence[str], b_words: Sequence[str], threshold: float = 0.0) -> float:
    """Word-set similarity for resúmenes (Jaccard / overlap).

    SequenceMatcher on 80-token lists was ~1.5 ms/pair and made Agrupación
    take several seconds on a 400-row dossier. Set overlap is the same
    "same text" signal at microseconds.
    """
    if not a_words or not b_words:
        return 0.0
    sa, sb = set(a_words[:MAX_CMP_WORDS]), set(b_words[:MAX_CMP_WORDS])
    jac = _jaccard_sets(sa, sb)
    ov = _overlap_sets(sa, sb)
    score = jac if ov < 0.90 else max(jac, ov)
    if threshold > 0 and score < threshold * 0.45:
        return 0.0
    return score


# ---------------------------------------------------------------------------
# Duplicates (NOT the same as Grupo noticia)
# ---------------------------------------------------------------------------

def _mencion_key(row: Dict[str, Any], km: Dict[str, str]) -> str:
    return _norm_text(row.get(km.get("menciones", "Menciones - Empresa"), ""))


def detectar_duplicados(rows: Sequence[Dict[str, Any]],
                        km: Optional[Dict[str, str]] = None) -> List[Dict[str, Any]]:
    """
    Duplicate = exact same publication, decided before any LLM call. The
    rules are strict and match the client spec; title equality alone NEVER
    marks a duplicate (that is Grupo noticia's job):

    1. Internet / Prensa / Revistas (and any non-AV row): duplicate ONLY if
       the same normalized article URL AND the same `Menciones - Empresa`.
       The article URL is the embedded hyperlink of `Link Nota`;
       `Link (Streaming - Imagen)` is consulted only when both rows have no
       Link Nota (some imports carry the article URL in that field). Same
       title with a different URL is NOT a duplicate.
    2. Radio / Televisión: duplicate ONLY if the same `Menciones - Empresa`
       AND the same `Medio` AND the same `Hora` (same broadcast slot).
       Fecha is deliberately NOT part of the key and a shared clip/link URL
       alone is NOT enough (client rule: "de resto no").

    Mención is part of every key so the dossier expansion (one row per
    company) never marks two companies of the same note as duplicates.
    """
    km = km or KEYMAP
    processed = deepcopy(list(rows))
    seen_url: Dict[Tuple[str, str], int] = {}
    seen_bcast: Dict[Tuple[str, str, str], int] = {}

    def _marcar(row: Dict[str, Any], orig_idx: int) -> None:
        orig = processed[orig_idx]
        row["is_duplicate"] = True
        row["_dup_of_index"] = orig_idx
        row[km.get("idduplicada", "ID duplicada")] = orig.get(km.get("idnoticia", "ID Noticia"), "")
        row["Tono IA"] = "Duplicada"
        row["Tema"] = "-"
        row["Subtema"] = "-"

    for i, row in enumerate(processed):
        row.setdefault("is_duplicate", False)
        if row.get("is_duplicate"):
            continue
        tipo = normalizar_tipo_medio(row.get(km.get("tipodemedio", "Tipo de Medio"), ""))
        mencion = _mencion_key(row, km)
        if not mencion:
            # Without a company mention there is no way to say the SAME
            # note was logged twice for the same company.
            continue

        if tipo in ("Radio", "Televisión"):
            medio = _norm_text(row.get(km.get("medio", "Medio"), ""))
            hora = normalizar_hora(row.get(km.get("hora", "Hora"), ""))
            if not (medio and hora):
                continue
            slot = (mencion, medio, hora)
            if slot in seen_bcast:
                _marcar(row, seen_bcast[slot])
            else:
                seen_bcast[slot] = i
            continue

        # Internet / Prensa / Revistas / Otro: article URL + mención.
        link_nota = normalize_url(url_de_celda_link(row.get(km.get("link_nota", "Link Nota"))))
        streaming = normalize_url(url_de_celda_link(row.get(km.get("link_streaming", "Link (Streaming - Imagen)"))))
        if link_nota:
            key = (link_nota, mencion)
            if key in seen_url:
                _marcar(row, seen_url[key])
            else:
                seen_url[key] = i
        elif streaming:
            key = (streaming, mencion)
            if key in seen_url:
                _marcar(row, seen_url[key])
            else:
                seen_url[key] = i
        # No URL at all → nothing to compare, never a duplicate.
    return processed


# ---------------------------------------------------------------------------
# Contexto analizado
# ---------------------------------------------------------------------------

def _lista_alias(marca: str, aliases=None) -> List[str]:
    nombres: List[str] = []
    if marca:
        nombres.extend(str(marca).split(";"))
    if isinstance(aliases, str):
        nombres.extend(aliases.split(";"))
    else:
        nombres.extend(str(a) for a in (aliases or []))
    out, seen = [], set()
    for n in nombres:
        k = _norm_text(n)
        if k and k not in seen:
            seen.add(k)
            out.append(n.strip())
    return out


def _compile_alias_matcher(marca: str, aliases=None) -> List[Tuple[str, str, "re.Pattern", List[str]]]:
    compiled = []
    for nombre in _lista_alias(marca, aliases):
        kn = _norm_text(nombre)
        if not kn:
            continue
        pat = re.compile(rf"(?<![a-z0-9]){re.escape(kn)}(?![a-z0-9])")
        toks = [t for t in kn.split() if len(t) >= 3 and t not in {"de", "del", "la", "el", "los", "las", "y"}]
        compiled.append((nombre.strip(), kn, pat, toks))
    return compiled


def _menciona_compiled(norm: str, compiled) -> bool:
    if not norm or not compiled:
        return False
    words = None
    for _n, _kn, pat, toks in compiled:
        if pat.search(norm):
            return True
        if len(toks) >= 2:
            if words is None:
                words = set(norm.split())
            if sum(t in words for t in toks) >= max(2, int(np.ceil(len(set(toks)) * 0.6))):
                return True
    return False


def _menciona(texto: str, marca: str, aliases=None) -> bool:
    return _menciona_compiled(_norm_text(texto), _compile_alias_matcher(marca, aliases))


def _ventanas_mencion_compiled(texto: str, norm: str, compiled, ventana: int = 220) -> List[str]:
    if not texto or not compiled:
        return []
    hits = []
    for _n, _kn, pat, _toks in compiled:
        for m in pat.finditer(norm):
            ratio = (m.start() / max(len(norm), 1))
            center = int(ratio * len(texto))
            lo, hi = max(0, center - ventana), min(len(texto), center + ventana)
            fragment = texto[lo:hi].strip()
            if fragment:
                hits.append(fragment)
    if hits:
        return hits
    partes = re.split(r"(?<=[\.\!\?\n])\s+", texto)
    out = []
    for p in partes:
        p = p.strip()
        if p and _menciona_compiled(_norm_text(p), compiled):
            out.append(p)
    return out


def _ventanas_mencion(texto: str, marca: str, aliases=None, ventana: int = 220) -> List[str]:
    if not texto:
        return []
    return _ventanas_mencion_compiled(
        texto, _norm_text(texto), _compile_alias_matcher(marca, aliases), ventana
    )


def extraer_contexto_analizado(titulo: Any, resumen: Any, marca: str,
                               aliases=None, cuerpo: Any = "",
                               _matcher=None) -> Dict[str, str]:
    """
    Coherent Colombian-Spanish paragraph from brand mention windows.
    Título + Resumen first; Cuerpo Completo only if needed.
    Fallback: Resumen, then full Título. Never mutates the original title cell.
    """
    tit = "" if titulo is None else str(titulo)
    res = "" if resumen is None else str(resumen)
    cue = "" if cuerpo is None else str(cuerpo)
    compiled = _matcher if _matcher is not None else _compile_alias_matcher(marca, aliases)
    tit_n = _norm_text(tit)
    res_n = _norm_text(res)
    title_hit = _menciona_compiled(tit_n, compiled)
    res_hit = _menciona_compiled(res_n, compiled)
    # Cuerpo Completo is only a fallback. Scanning 410 long bodies with
    # unidecode+regex was ~6s of the pipeline; skip it when the brand
    # already appears in Título or Resumen.
    cue_n = ""
    if cue and not (title_hit or res_hit):
        cue_n = _norm_text(cue[:CUERPO_SCAN_CHARS])
        cue_hit = _menciona_compiled(cue_n, compiled)
    else:
        cue_hit = False

    coincidencia = ""
    for n, kn, pat, _toks in compiled:
        if (kn and (
            pat.search(tit_n) or pat.search(res_n) or (cue_n and pat.search(cue_n))
        )):
            coincidencia = n
            break
    if not coincidencia and (title_hit or res_hit or cue_hit):
        coincidencia = marca or ""

    origen_parts = []
    if title_hit:
        origen_parts.append("Título")
    if res_hit:
        origen_parts.append("Resumen")
    if cue_hit and not (title_hit or res_hit):
        origen_parts.append("Cuerpo Completo")
    origen = ", ".join(origen_parts)

    bloques: List[str] = []
    if title_hit or res_hit:
        bloques.extend(_ventanas_mencion_compiled(tit, tit_n, compiled))
        bloques.extend(_ventanas_mencion_compiled(res, res_n, compiled))
        if not bloques:
            if title_hit:
                bloques.append(tit.strip())
            if res_hit:
                bloques.append(res.strip())
    elif cue_hit:
        bloques.extend(_ventanas_mencion_compiled(cue[:CUERPO_SCAN_CHARS], cue_n, compiled)[:3])
    else:
        if res.strip():
            bloques.append(res.strip())
            origen = origen or "Resumen"
        elif tit.strip():
            bloques.append(tit.strip())
            origen = origen or "Título"

    vistos, out = set(), []
    for b in bloques:
        k = _norm_text(b)
        if k and k not in vistos:
            vistos.add(k)
            out.append(re.sub(r"\s+", " ", b).strip())
    contexto = " ".join(out).strip()
    if contexto and contexto[-1] not in ".!?…":
        contexto = contexto.rstrip(" ,;") + "."
    return {
        "contexto": contexto[:1800],
        "coincidencia": coincidencia,
        "origen": origen,
    }


# ---------------------------------------------------------------------------
# Blocked grouping (Grupo noticia) — not n²
# ---------------------------------------------------------------------------

class DSU:
    def __init__(self, n: int) -> None:
        self.p = list(range(n))
        self.rank = [0] * n

    def find(self, i: int) -> int:
        while self.p[i] != i:
            self.p[i] = self.p[self.p[i]]
            i = self.p[i]
        return i

    def union(self, i: int, j: int) -> None:
        ri, rj = self.find(i), self.find(j)
        if ri == rj:
            return
        if self.rank[ri] < self.rank[rj]:
            ri, rj = rj, ri
        self.p[rj] = ri
        if self.rank[ri] == self.rank[rj]:
            self.rank[ri] += 1

    def grupos(self, n: int) -> Dict[int, List[int]]:
        c: Dict[int, List[int]] = defaultdict(list)
        for i in range(n):
            c[self.find(i)].append(i)
        return dict(c)


def _prefix(text: str, words: int) -> str:
    toks = _norm_text(text).split()
    return " ".join(toks[:words])


def _candidate_pairs(n: int, buckets: Dict[str, List[int]],
                     counter: Optional[ComparisonCounter] = None,
                     token_sets: Optional[Sequence[set]] = None,
                     unit_matrix: Optional[np.ndarray] = None,
                     ) -> List[Tuple[int, int]]:
    """Blocked candidate generation (Python-level work stays << n²).

    - Exact buckets (title prefix, resumen prefix, URL) always meet.
      Date+time buckets are capped: same hour alone is too coarse.
    - Token buckets: pairs sharing ≥2 distinctive tokens, or 1 rare token
      (df ≤ 6), via posting-list co-occurrence counts.
    - Embedding neighbours: top-K per row from one matrix product, not
      every pair above a cosine floor (that exploded to ~8k pairs).
    """
    pares = set()
    for key, idxs in buckets.items():
        if len(idxs) < 2:
            continue
        if key.startswith("k:"):
            continue
        orden = sorted(set(idxs))
        if key.startswith("d:") and len(orden) > MAX_EXACT_BUCKET:
            continue
        if len(orden) > MAX_BUCKET:
            orden = orden[:MAX_BUCKET]
        for a in range(len(orden)):
            for b in range(a + 1, len(orden)):
                pares.add((orden[a], orden[b]))

    if token_sets is not None:
        posting: Dict[str, List[int]] = defaultdict(list)
        for i, toks in enumerate(token_sets):
            for t in toks:
                posting[t].append(i)
        shared: Dict[Tuple[int, int], int] = defaultdict(int)
        for tok, idxs in posting.items():
            df = len(idxs)
            if df < 2 or df > MAX_BUCKET:
                continue
            rare = df <= 6
            for a in range(df):
                ia = idxs[a]
                for b in range(a + 1, df):
                    key = (ia, idxs[b]) if ia < idxs[b] else (idxs[b], ia)
                    shared[key] += 1
                    if rare:
                        shared[key] += 1
        for key, cnt in shared.items():
            if cnt >= 2:
                pares.add(key)
            if len(pares) >= MAX_PAIRS_TOTAL:
                break

    if unit_matrix is not None and len(unit_matrix) >= 2 and len(pares) < MAX_PAIRS_TOTAL:
        M = unit_matrix
        k = min(EMBED_NEIGHBORS, len(M) - 1)
        if k >= 1:
            sims = M @ M.T
            np.fill_diagonal(sims, -2.0)
            # argpartition is O(n) per row; full argsort was wasteful.
            neigh = np.argpartition(-sims, kth=k - 1, axis=1)[:, :k]
            rows = np.arange(len(M))[:, None]
            mask = sims[rows, neigh] >= EMBED_CANDIDATE_MIN
            ii = np.repeat(np.arange(len(M)), k)[mask.ravel()]
            jj = neigh.ravel()[mask.ravel()]
            for i, j in zip(ii.tolist(), jj.tolist()):
                if i < j:
                    pares.add((i, j))
                if len(pares) >= MAX_PAIRS_TOTAL:
                    break

    if counter:
        counter.add(len(pares))
    return sorted(pares)


def agrupar_noticias_bloqueado(
    titulos: Sequence[Any],
    resumenes: Sequence[Any],
    contextos: Optional[Sequence[Any]] = None,
    embeddings: Optional[Sequence[Any]] = None,
    urls: Optional[Sequence[Any]] = None,
    fechas: Optional[Sequence[Any]] = None,
    horas: Optional[Sequence[Any]] = None,
    counter: Optional[ComparisonCounter] = None,
) -> Dict[int, List[int]]:
    """
    Same-fact grouping by similar title OR similar resumen OR high semantic
    similarity on Contexto analizado. Blocked by title prefix, resumen
    prefix, distinctive tokens, normalized URL and date+time.
    """
    n = len(titulos)
    dsu = DSU(n)
    tit_n = [normalize_title_for_comparison(t) for t in titulos]
    res_n = [_norm_text(r) for r in resumenes]
    ctx_n = [_norm_text(c) for c in (contextos or [""] * n)]
    url_n = [normalize_url(u) for u in (urls or [""] * n)]

    # Per-row precomputation: tokens/sets built once, not once per pair.
    res_pref = [_prefix(r, 10) for r in res_n]
    texto_row = [" ".join(x for x in (tit_n[i], res_n[i], ctx_n[i]) if x) for i in range(n)]
    tok_conf = [_tokens_distintivos(texto_row[i], min_len=3) for i in range(n)]
    tok_ov = [_tokens_distintivos(ctx_n[i] or texto_row[i]) for i in range(n)]
    tok_tit = [_stems(_tokens_distintivos(tit_n[i])) for i in range(n)]
    tit_clave = []
    for i in range(n):
        ordered = []
        seen = set()
        for w in tit_n[i].split():
            if len(w) < 4 or w in _TOKENS_DEBILES or w.isdigit():
                continue
            st = _stem_es(w)
            if st in seen:
                continue
            seen.add(st)
            ordered.append(st)
            if len(ordered) >= 4:
                break
        tit_clave.append(" ".join(ordered))
    # Blocking tokens: headline tokens plus the first distinctive tokens of
    # the resumen/contexto, so paraphrased headlines still meet.
    tok_block = []
    for i in range(n):
        s = set(tok_tit[i])
        extra = [t for t in (res_n[i] or ctx_n[i]).split() if len(t) >= 5 and t not in _TOKENS_DEBILES]
        s.update(_stems(extra[:10]))
        tok_block.append(s)

    # Bitmask of opposite-action tokens so the pair loop is a few AND tests.
    action_bits = [0] * n
    for i in range(n):
        bits = 0
        for k, (ga, gb) in enumerate(_ACCIONES_OPUESTAS):
            if tok_conf[i] & ga:
                bits |= 1 << (2 * k)
            if tok_conf[i] & gb:
                bits |= 1 << (2 * k + 1)
        action_bits[i] = bits
    opp_pairs = [(1 << (2 * k), 1 << (2 * k + 1)) for k in range(len(_ACCIONES_OPUESTAS))]

    buckets: Dict[str, List[int]] = defaultdict(list)
    for i in range(n):
        tp = _prefix(tit_n[i], 6)
        rp = _prefix(res_n[i] or ctx_n[i], 12)
        if tp:
            buckets[f"t:{tp}"].append(i)
        if tit_clave[i] and tit_clave[i].count(" ") >= 2:
            buckets[f"c:{tit_clave[i]}"].append(i)
        if rp:
            buckets[f"r:{rp}"].append(i)
        if url_n[i]:
            buckets[f"u:{url_n[i]}"].append(i)
        if fechas is not None and horas is not None:
            fh = f"{normalizar_fecha(fechas[i])}|{normalizar_hora(horas[i])}"
            if fh != "|" and not fh.endswith("|"):
                buckets[f"d:{fh}"].append(i)

    embs = list(embeddings) if embeddings is not None else [None] * n
    unit_matrix = None
    dim = None
    present = []
    for k, e in enumerate(embs):
        if e is None:
            continue
        v = np.asarray(e, dtype=np.float32).ravel()
        nrm = float(np.linalg.norm(v))
        if nrm > 0:
            present.append((k, v / nrm))
            dim = len(v)
    if dim is not None:
        unit_matrix = np.zeros((n, dim), dtype=np.float32)
        for k, u in present:
            if len(u) == dim:
                unit_matrix[k] = u

    pares = _candidate_pairs(n, buckets, counter, token_sets=tok_block, unit_matrix=unit_matrix)

    sem_all = None
    if unit_matrix is not None and pares:
        I = np.fromiter((p[0] for p in pares), dtype=np.intp, count=len(pares))
        J = np.fromiter((p[1] for p in pares), dtype=np.intp, count=len(pares))
        sem_all = np.einsum("ij,ij->i", unit_matrix[I], unit_matrix[J])

    def _contenido(i: int, j: int) -> bool:
        a, b = tit_n[i], tit_n[j]
        return bool(a and b and min(len(a), len(b)) >= 25 and (a in b or b in a))

    def _conflicto(i: int, j: int) -> bool:
        a, b = action_bits[i], action_bits[j]
        for ga, gb in opp_pairs:
            if (a & ga and b & gb) or (a & gb and b & ga):
                return True
        return False

    for k, (i, j) in enumerate(pares):
        semantic = float(sem_all[k]) if sem_all is not None else 0.0
        same_url = bool(url_n[i] and url_n[i] == url_n[j])
        inter_t = tok_tit[i] & tok_tit[j]
        shared_t = len(inter_t)
        jac_t = (shared_t / len(tok_tit[i] | tok_tit[j])) if (tok_tit[i] and tok_tit[j] and shared_t) else 0.0
        titulo_cerca = (
            _contenido(i, j)
            or (
                shared_t >= 4
                and tok_tit[i] and tok_tit[j]
                and (tok_tit[i] <= tok_tit[j] or tok_tit[j] <= tok_tit[i])
            )
            or (tit_clave[i] and tit_clave[i] == tit_clave[j] and tit_clave[i].count(" ") >= 3)
            or (jac_t >= 0.80 and shared_t >= 4)
        )
        # Opposite-action guard must not split a near-identical headline
        # (one title is a prefix of the other, extra money/source clause).
        if _conflicto(i, j) and not titulo_cerca:
            continue
        overlap = _overlap_sets(tok_ov[i], tok_ov[j])
        # Same-resumen = same opening, not bag-of-words Jaccard.
        # A 22-word vocabulary over 200 tokens yields Jaccard ~1 for
        # unrelated notes and used to collapse a 400-row dossier into 1 group.
        sim_r = 0.0
        if res_pref[i] and res_pref[i] == res_pref[j]:
            sim_r = 0.95
        elif res_n[i][:72] and res_n[i][:72] == res_n[j][:72]:
            sim_r = 0.95
        mismo_hecho = (
            same_url
            or titulo_cerca
            or (jac_t >= JACCARD_TITULO_GRUPO and shared_t >= 4)
            or (jac_t >= 0.80 and shared_t >= 4)
            or (jac_t >= 0.50 and shared_t >= 4 and semantic >= 0.82)
            or sim_r >= SIMILARITY_THRESHOLD_RESUMEN
            or (semantic >= SIMILARITY_SEMANTIC_ALONE and shared_t >= 4)
            or (semantic >= SIMILARITY_THRESHOLD_SEMANTIC and overlap >= MIN_OVERLAP_GRUPO and shared_t >= 3)
        )
        # Gray zone only: near-paraphrase titles that set metrics miss.
        # SequenceMatcher stays off the common path (was ~1.5 ms × 8k pairs).
        if (
            not mismo_hecho
            and tit_n[i] and tit_n[j]
            and shared_t >= 4
            and semantic >= 0.75
            and jac_t >= 0.40
        ):
            if _ratio(tit_n[i], tit_n[j], SIMILARITY_THRESHOLD_TITULOS) >= SIMILARITY_THRESHOLD_TITULOS:
                mismo_hecho = True
        if mismo_hecho:
            dsu.union(i, j)
    return dsu.grupos(n)


def ids_grupo(grupos: Dict[int, List[int]]) -> List[str]:
    out = [""] * sum(len(v) for v in grupos.values())
    for numero, idxs in enumerate(grupos.values(), start=1):
        gid = f"G{numero:05d}"
        for i in idxs:
            out[i] = gid
    return out


# ---------------------------------------------------------------------------
# Tema / Subtema quality
# ---------------------------------------------------------------------------

_BAD_COLLAGE = re.compile(
    r"^(?:\S+\s+){2,}\S+$"
)


def _palabras(etiqueta: str) -> List[str]:
    return [p for p in re.split(r"\s+", (etiqueta or "").strip()) if p]


_INICIO_PROHIBIDO = {
    "multiples", "multiple", "diversos", "diversas", "varios", "varias",
    "algunos", "algunas", "escalables", "escalable", "destacada", "destacado",
    "destacadas", "destacados", "elegida", "elegido", "premiada", "premiado",
}


def _ultima_palabra(etiqueta: str) -> str:
    words = _palabras(etiqueta)
    if not words:
        return ""
    return unidecode(words[-1].lower().rstrip(".,;:!?¿¡\"'$"))


def _termina_en_cifra_o_monto(etiqueta: str) -> bool:
    last = _ultima_palabra(etiqueta)
    if not last:
        return False
    if re.search(r"\d", last) or last in {"millones", "millon", "mil", "usd", "us", "dolares", "pesos", "cop"}:
        return True
    if re.search(r"\$|\bus\$?\s*\d|\d[\d.,]*\s*(millones|mil)?\s*$", etiqueta, flags=re.I):
        return True
    return False


def validar_subtema(etiqueta: str) -> bool:
    """Complete topical noun phrase, 3–6 words. Report heading, not a fragment."""
    if not etiqueta or not str(etiqueta).strip():
        return False
    et = str(etiqueta).strip().strip(" .;:¡!¿?")
    if "," in et or ";" in et or ":" in et:
        return False
    if _es_generica(et):
        return False
    words = _palabras(et)
    if not (MIN_PALABRAS_SUBTEMA <= len(words) <= MAX_PALABRAS_SUBTEMA):
        return False
    last = _ultima_palabra(et)
    if last in _TRAILING_INCOMPLETE or last in _QUESTION_TAILS:
        return False
    if _termina_en_cifra_o_monto(et):
        return False
    if _es_verbo_conjugado(words[-1]) or _es_verbo_conjugado(words[0]):
        return False
    if any(_es_verbo_conjugado(w) for w in words):
        return False
    if any(unidecode(w.lower()) in {"como", "cómo"} for w in words):
        return False
    first = unidecode(words[0].lower().rstrip(".,;:"))
    if first in _DROP_LEAD | _INICIO_PROHIBIDO | {"que", "y", "o", "pero"}:
        return False
    head = " ".join(unidecode(w.lower()) for w in words[:2])
    if first in _CHANNEL_PREFIXES or head in _CHANNEL_PREFIXES:
        return False
    if "|" in et or et.lower().startswith("video"):
        return False
    # Participle + preposition is a headline leftover, not a topic
    # ('Destacada entre las sociólogas…').
    if len(words) >= 2 and re.search(r"(?:ada|ado|idas|idos|ida|ido)$", first):
        if unidecode(words[1].lower()) in _NEXOS | {"entre", "como"}:
            return False
    nexos = [unidecode(w.lower()) for w in words[1:] if unidecode(w.lower()) in _NEXOS]
    content = [w for w in words if unidecode(w.lower()) not in _NEXOS | STOPWORDS_ES]
    if not nexos:
        if not all(_parece_adjetivo(w) for w in words[1:]):
            return False
    if len(content) >= 4 and not nexos:
        return False
    if "?" in et or "¿" in et:
        return False
    return True


def validar_tema(etiqueta: str) -> bool:
    """Tema: coherent thematic noun phrase, 2–5 words, no verbs, no
    commas, no numbers, no keyword collage."""
    if not etiqueta or not str(etiqueta).strip():
        return False
    et = str(etiqueta).strip()
    if "?" in et or "¿" in et or "|" in et or "," in et or ";" in et or _es_generica(et):
        return False
    words = _palabras(et)
    if not (2 <= len(words) <= MAX_PALABRAS_TEMA):
        return False
    last = _ultima_palabra(et)
    if last in _TRAILING_INCOMPLETE or last in _QUESTION_TAILS:
        return False
    if _termina_en_cifra_o_monto(et):
        return False
    if any(_es_verbo_conjugado(w) for w in words):
        return False
    if unidecode(words[0].lower()) in _CHANNEL_PREFIXES | _DROP_LEAD:
        return False
    content = [w for w in words if unidecode(w.lower()) not in _NEXOS | STOPWORDS_ES | {"y", "e"}]
    nexos = [w for w in words if unidecode(w.lower()) in _NEXOS | {"y", "e"}]
    if not nexos and len(words) >= 2:
        if not all(_parece_adjetivo(w) for w in words[1:]):
            return False
    if len(content) >= 4 and not nexos:
        return False
    return True


def _capitalizar(frase: str) -> str:
    frase = re.sub(r"\s+", " ", (frase or "").strip().strip(" ,;:."))
    if not frase:
        return ""
    return frase[0].upper() + frase[1:]


_ETIQUETAS_GENERICAS = {
    "cobertura informativa general", "cobertura informativa", "informacion general",
    "actualidad", "noticias", "varios", "sin tema", "otros", "general", "cobertura",
    "actualidad general", "noticias generales", "informacion", "tema general",
    "cobertura de informacion relevante", "cobertura de la noticia", "hecho noticioso",
    "noticia sobre la marca", "mencion de la marca", "actividad corporativa",
    "gestion corporativa", "gestion institucional",
}


def _es_generica(etiqueta: str) -> bool:
    return _norm_text(etiqueta) in _ETIQUETAS_GENERICAS


_VERBOS_TITULAR = re.compile(
    r"\b(presenta|presentan|presento|anuncia|anuncian|anuncio|destaca|destacan|"
    r"confirma|confirman|confirmo|pide|piden|pidio|lanza|lanzan|lanzo|abre|abren|abrio|"
    r"inaugura|inauguran|inauguro|entrega|entregan|entrego|reporta|reportan|reporto|"
    r"advierte|advierten|advirtio|revela|revelan|revelo|asegura|aseguran|aseguro|"
    r"afirma|afirman|afirmo|alerta|alertan|alerto|denuncia|denuncian|denuncio|"
    r"rechaza|rechazan|rechazo|aprueba|aprueban|aprobo|firma|firman|firmo|"
    r"recibe|reciben|recibio|celebra|celebran|celebro|invierte|invierten|invirtio|"
    r"impulsa|impulsan|impulso|lidera|lideran|lidero|logra|logran|logro|"
    r"busca|buscan|busco|explica|explican|explico|analiza|analizan|analizo|"
    r"promueve|promueven|promovio|realiza|realizan|realizo|ofrece|ofrecen|ofrecio|"
    r"suspende|suspenden|suspendio|cierra|cierran|cerro|activa|activan|activo|"
    r"refuerza|refuerzan|reforzo|amplia|amplian|amplio|reduce|reducen|redujo|"
    r"compra|compro|compraron|adquiere|adquirio|adquirieron)\b"
)

_DROP_LEAD = {"otro", "otra", "otros", "otras", "este", "esta", "estos", "estas",
              "nuevo", "nueva", "nuevos", "nuevas", "asi", "hoy", "ayer", "ahora"}

_VERBO_A_EVENTO = [
    (re.compile(r"\b(adquiri[oó]|adquiere|adquirieron|compra|compr[oó]|compraron|comprado)\b", re.I), "Adquisición"),
    (re.compile(r"\b(invirti[oó]|invierte|invierten|inversion)\b", re.I), "Inversión"),
    (re.compile(r"\b(lanz[oó]|lanza|lanzan|lanzamiento)\b", re.I), "Lanzamiento"),
    (re.compile(r"\b(inaugur[oó]|inaugura|inauguran)\b", re.I), "Inauguración"),
    (re.compile(r"\b(anunci[oó]|anuncia|anuncian)\b", re.I), "Anuncio"),
    (re.compile(r"\b(firm[oó]|firma|firman|convenio)\b", re.I), "Convenio"),
    (re.compile(r"\b(aprob[oó]|aprueba|aprueban)\b", re.I), "Aprobación"),
    (re.compile(r"\b(rechaz[oó]|rechaza|rechazan)\b", re.I), "Rechazo"),
    (re.compile(r"\b(reconoc|premi[oó]|galardon|destacad[oa]s? entre|eligi[oó])\b", re.I), "Reconocimiento"),
    (re.compile(r"\b(present[oó]|presenta|presentan)\b", re.I), "Presentación"),
    (re.compile(r"\b(entreg[oó]|entrega|entregan)\b", re.I), "Entrega"),
]


def _limpiar_titular(titulo: str) -> str:
    t = re.sub(r"<[^>]+>", " ", str(titulo or ""))
    t = re.sub(r"^\s*(video|en vivo|audio|streaming|podcast|galeria|galería|fotos?)\s*[\|:\-–—]\s*",
               "", t, flags=re.I)
    t = t.split("|")[-1] if "|" in t else t
    return re.sub(r"\s+", " ", t).strip(" .:;-–—")


def _entidad_inicial(titulo: str) -> str:
    """Company/person named at the start of the headline, before a comma."""
    limpio = _limpiar_titular(titulo)
    if not limpio:
        return ""
    cabeza = re.split(r"[,:;–—]", limpio, maxsplit=1)[0].strip()
    words = [w for w in cabeza.split() if w]
    if 1 <= len(words) <= 4 and not any(_es_verbo_conjugado(w) for w in words):
        return " ".join(words)
    return words[0] if words else ""


def _entidad_en_resto(resto: str) -> str:
    """First proper-noun-like token in the remainder, ignoring money."""
    limpio = re.sub(r"\bpor\s+US\$?\s*[\d.,]+\s*(millones?)?", " ", resto or "", flags=re.I)
    limpio = re.sub(r"US\$?\s*[\d.,]+|\$(?:\d[\d.,]*)", " ", limpio)
    limpio = re.sub(r"\b\d[\d.,]*\b", " ", limpio)
    for w in limpio.split():
        wl = unidecode(w.lower().strip(".,;:"))
        if len(wl) < 3 or wl in STOPWORDS_ES | _DROP_LEAD | {"ia"}:
            continue
        if w[:1].isupper() or (wl[0].isalpha() and wl not in STOPWORDS_ES):
            return w.strip(".,;:")
    return ""


def _evento_del_verbo(texto: str) -> str:
    plano = unidecode(texto or "")
    for pat, evento in _VERBO_A_EVENTO:
        if pat.search(plano) or pat.search(texto or ""):
            return evento
    return ""


def _nominalizar_titular(titulo: str) -> str:
    """Headline → 3–6 word topic ('Adquisición de InterPositive por Netflix')."""
    limpio = _limpiar_titular(titulo)
    if not limpio:
        return ""
    entidad = _entidad_inicial(limpio)
    rel = re.match(
        r"^(.{2,80}?),\s+(?:la|el|los|las)\s+.+?\s+que\s+(\S+)\s+(.+)$",
        limpio, flags=re.I | re.S,
    )
    if rel:
        entidad = entidad or rel.group(1).strip()
        evento = _evento_del_verbo(rel.group(2)) or _evento_del_verbo(rel.group(3)) or _evento_del_verbo(limpio)
        otra = _entidad_en_resto(rel.group(3))
        if evento and entidad:
            if otra and unidecode(otra.lower()) != unidecode(entidad.split()[0].lower()):
                frase = f"{evento} de {entidad} por {otra}"
            else:
                frase = f"{evento} de {entidad}"
            frase = _capitalizar(" ".join(_recortar_sintagma(frase.split(), MAX_PALABRAS_SUBTEMA)))
            if validar_subtema(frase):
                return frase
    evento = _evento_del_verbo(limpio)
    if evento and entidad:
        otra = ""
        m = re.search(r"\b(?:por|de)\s+([A-ZÁÉÍÓÚÑ][\wÁÉÍÓÚÑáéíóúñ]+)", limpio)
        if m and unidecode(m.group(1).lower()) not in STOPWORDS_ES:
            otra = m.group(1)
        if otra and unidecode(otra.lower()) != unidecode(entidad.split()[0].lower()):
            frase = f"{evento} de {entidad} por {otra}"
        else:
            frase = f"{evento} de {entidad}"
        frase = _capitalizar(" ".join(_recortar_sintagma(frase.split(), MAX_PALABRAS_SUBTEMA)))
        if validar_subtema(frase):
            return frase
    return ""


def _objeto_tras_verbo(titulo: str) -> str:
    """'Alcalde presenta balance de seguridad en Cali' → 'Balance de seguridad en Cali'."""
    limpio = _limpiar_titular(titulo)
    if not limpio:
        return ""
    tokens = limpio.split()
    plano = unidecode(limpio.lower())
    verbo_pos = None
    m = _VERBOS_TITULAR.search(plano)
    if m:
        verbo_pos = len(plano[: m.start()].split())
    else:
        for k, tok in enumerate(tokens):
            if _es_verbo_conjugado(tok):
                verbo_pos = k
                break
    if verbo_pos is None or verbo_pos >= len(tokens) - 1:
        return ""
    resto = " ".join(tokens[verbo_pos + 1:]).strip(" ,.:;")
    resto = re.split(r"[\.\?\!¿¡]|\s[-–—]\s|:\s", resto)[0].strip()
    words = resto.split()
    while words and unidecode(words[0].lower()) in _DROP_LEAD | {"que", "a", "al", "el", "la", "los", "las", "un", "una", "unos", "unas", "su", "sus"}:
        words.pop(0)
    if not words:
        return ""
    # Money / proper-noun fragments are not topics ('Netflix por US$587 millones').
    cand_raw = " ".join(words)
    if _termina_en_cifra_o_monto(cand_raw) or re.search(r"\$|\d", cand_raw):
        return ""
    words = _recortar_sintagma(words, MAX_PALABRAS_SUBTEMA)
    frase = _capitalizar(" ".join(words))
    return frase if validar_subtema(frase) else ""


def fallback_subtema(contexto: str, titulo: str = "") -> str:
    """
    Deterministic 3–6 word topical noun phrase from the HEADLINE.
    Context is a mention window and often a product blurb — never the
    source of the object-after-verb heuristic.
    """
    # 1) Nominalize the event named in the title.
    nom = _nominalizar_titular(titulo)
    if nom:
        return nom

    # 2) Object of the headline verb, only if it is already a complete topic.
    desde_titular = _objeto_tras_verbo(titulo)
    if desde_titular:
        return desde_titular

    # 3) A verb-less headline is already a noun phrase.
    limpio = _limpiar_titular(titulo)
    if limpio:
        cand_words = re.split(r"[\.\?\!¿¡]|\s[-–—]\s|:\s", limpio)[0].split()
        while cand_words and unidecode(cand_words[0].lower()) in _DROP_LEAD | {"el", "la", "los", "las", "un", "una"}:
            cand_words.pop(0)
        cand = _capitalizar(" ".join(_recortar_sintagma(cand_words, MAX_PALABRAS_SUBTEMA)))
        if cand and validar_subtema(cand) and not any(_es_verbo_conjugado(w) for w in cand.split()):
            return cand

    texto = " ".join(x for x in (titulo, contexto) if x).strip()
    norm = _norm_text(texto)

    # 4) Generic topical patterns (no brand- or client-specific rules).
    if re.search(r"\b(terremoto|sismo|temblor)\b", norm) and re.search(
        r"\b(victima|victimas|muerto|muertos|herido|heridos|asciend|fallecid)", norm
    ):
        return "Balance de víctimas del terremoto"
    if re.search(r"\b(sismo|terremoto|temblor)\b", norm) and re.search(
        r"\b(recomend|saber|prepar|simulacro|que hacer)\b", norm
    ):
        return "Recomendaciones ante sismos"
    if re.search(r"\b(terremoto|sismo|temblor)\b", norm):
        return "Cobertura del sismo en Colombia"
    if re.search(r"\b(jornada|fecha)\b", norm) and re.search(r"\b(resultado|resultados|liga|torneo|partido)\b", norm):
        return "Resultados de la jornada deportiva"
    if re.search(r"\b(adquisicion|adquiri|compr[oó]|compra de)\b", norm) and re.search(r"\b(ia|inteligencia|startup|tecnolog)\b", norm):
        ent = _entidad_inicial(titulo) or "la compañía"
        frase = f"Adquisición de {ent}"
        if validar_subtema(frase):
            return _capitalizar(frase)

    # 5) Last resort: entity coverage or thematic umbrella. NEVER join the
    # first bare keywords with 'de' ('Tecnología de innovación de obras'):
    # that keyword collage is exactly what the client rejects.
    ent = _entidad_inicial(titulo)
    if ent and len(ent.split()) <= 3:
        frase = f"Cobertura sobre {ent}"
        if validar_subtema(frase):
            return _capitalizar(frase)
    # 6) Thematic umbrella from the fixed taxonomy of headline + context, so
    # even the last resort names the actual issue ('Cobertura de salud
    # pública') instead of gluing headline words with 'de'.
    tema_fb = tema_por_taxonomia("", titulo, contexto)
    if tema_fb and validar_tema(tema_fb):
        frase = _capitalizar(f"Cobertura de {tema_fb.lower()}")
        if validar_subtema(frase):
            return frase
    return "Cobertura de información relevante"


TAXONOMIA_TEMAS: Dict[str, Tuple[str, ...]] = {
    "Emergencias y desastres naturales": (
        "sismo", "terremoto", "temblor", "inundacion", "deslizamiento", "avalancha",
        "incendio", "emergencia", "damnificado", "rescate", "bombero", "huracan",
        "lluvia", "ola invernal", "creciente", "evacuacion", "tragedia", "victima",
        "simulacro", "alerta roja", "socorro", "topos",
    ),
    "Seguridad y orden público": (
        "homicidio", "seguridad", "policia", "captura", "robo", "hurto", "atentado",
        "violencia", "extorsion", "delincuencia", "disidencia", "ejercito", "secuestro",
        "masacre", "asesinato", "sicario", "banda", "criminal", "explosivo", "fleteo",
        "riña", "rina", "microtrafico", "narcotrafico", "orden publico",
    ),
    "Salud pública": (
        "salud", "hospital", "eps", "medicamento", "vacuna", "paciente", "enfermedad",
        "clinica", "epidemia", "dengue", "medico", "urgencias", "brote", "sanitario",
        "cancer", "covid", "mental", "ambulancia",
    ),
    "Educación y formación": (
        "educacion", "colegio", "universidad", "estudiante", "docente", "icfes",
        "matricula", "beca", "profesor", "escolar", "academico", "rector", "pae",
        "campus", "graduacion", "investigacion cientifica",
    ),
    "Economía y empresas": (
        "economia", "inflacion", "empresa", "inversion", "dolar", "precio", "comercio",
        "exportacion", "impuesto", "tributaria", "banco", "credito", "pib", "mercado",
        "negocio", "ventas", "consumo", "industria", "emprendimiento", "financiero",
        "bolsa", "arancel", "importacion", "ganancia", "utilidades", "cifras economicas",
        "adquisicion", "compra", "fusion", "millones",
    ),
    "Empleo y trabajo": (
        "empleo", "trabajador", "sindicato", "salario", "desempleo", "paro", "huelga",
        "laboral", "contratacion", "despido", "jornada laboral", "pension",
        "pensionado", "pensionados", "mesada", "mesadas", "jubilacion", "jubilaciones",
    ),
    "Infraestructura y movilidad": (
        "via", "vias", "carretera", "movilidad", "transporte", "obra", "puente",
        "aeropuerto", "metro", "transmilenio", "mio", "trafico", "trancon", "peaje",
        "infraestructura", "pavimento", "ciclorruta", "semaforo", "tunel", "autopista",
        "terminal", "vial", "malla viaria", "hueco",
    ),
    "Política y gobierno": (
        "alcalde", "alcaldia", "gobernador", "gobernacion", "congreso", "senado",
        "presidente", "eleccion", "gobierno", "ministro", "ministerio", "decreto",
        "reforma", "concejo", "asamblea", "candidato", "partido politico", "campaña",
        "campana", "plan de desarrollo", "consejo de ministros", "politica",
        "gabinete", "posesion", "renuncia", "nombramiento", "agenda",
    ),
    "Justicia y procesos judiciales": (
        "fiscalia", "juez", "procuraduria", "contraloria", "condena", "demanda",
        "tutela", "corte suprema", "corte constitucional", "alta corte", "consejo de estado",
        "tribunal", "investigacion", "sancion", "imputacion",
        "audiencia", "carcel", "sentencia", "juicio", "abogado", "delito",
    ),
    "Actualidad deportiva": (
        "futbol", "liga", "partido", "jornada", "seleccion", "torneo", "atleta",
        "ciclismo", "deportivo", "campeonato", "gol", "estadio", "tecnico", "jugador",
        "america de cali", "deportivo cali", "diporto", "clasico", "medalla", "maraton",
        "deporte", "olimpico", "baloncesto", "tenis",
    ),
    "Cultura y entretenimiento": (
        "festival", "concierto", "feria", "cultura", "musica", "cine", "arte",
        "artista", "teatro", "salsa", "exposicion", "libro", "literatura", "danza",
        "petronio", "carnaval", "premio cultural", "museo", "cantante",
    ),
    "Medio ambiente": (
        "ambiente", "ambiental", "contaminacion", "deforestacion", "rio", "fauna",
        "clima", "reciclaje", "agua", "cambio climatico", "humedal", "bosque",
        "biodiversidad", "residuos", "basuras", "arbol", "mineria ilegal", "emisiones",
    ),
    "Servicios públicos": (
        "energia", "acueducto", "alcantarillado", "tarifa", "tarifas", "luz", "gas", "internet",
        "emcali", "epm", "corte de agua", "cortes de agua", "corte del servicio",
        "suspension del servicio", "racionamiento", "apagon", "factura",
        "servicio publico", "servicios publicos", "aseo", "recoleccion de basuras",
    ),
    "Responsabilidad social y solidaridad": (
        "solidaridad", "donacion", "voluntario", "fundacion", "ayuda humanitaria",
        "comunidad", "beneficiario", "vulnerable", "social", "jornada de salud",
        "brigada", "campaña solidaria", "entrega de ayudas",
    ),
    "Tecnología e innovación": (
        "tecnologia", "digital", "inteligencia artificial", "innovacion", "app",
        "ciberseguridad", "software", "plataforma", "startup", "datos", "conectividad",
        "adquisicion", "ia", "netflix", "algoritmo", "aplicacion",
    ),
    "Turismo y destinos": (
        "turismo", "turista", "hotel", "destino", "temporada", "visitante", "viajero",
    ),
    "Vivienda y ordenamiento territorial": (
        "vivienda", "pot", "urbanismo", "predio", "arriendo", "construccion de vivienda",
        "barrio", "invasion", "titulacion", "subsidio de vivienda",
    ),
}

TEMA_POR_DEFECTO = "Actualidad institucional"


def tema_por_taxonomia(subtema: str, contexto: str = "", titulo: str = "") -> str:
    """Deterministic tema: best-matching thematic category by keyword hits
    on subtema (weight 3), título (2) and contexto (1). Never a collage."""
    campos = (
        (_norm_text(subtema), 3),
        (_norm_text(titulo), 2),
        (_norm_text(contexto)[:1200], 1),
    )
    puntajes: Dict[str, float] = defaultdict(float)
    for tema, claves in TAXONOMIA_TEMAS.items():
        for texto, peso in campos:
            if not texto:
                continue
            padded = f" {texto} "
            tokens = set(texto.split())
            for k in claves:
                # Token boundary so 'cine' does not match 'cineastas'.
                hit = (
                    k in tokens
                    or f" {k} " in padded
                    or f" {k}s " in padded
                    or (len(k) >= 5 and any(t.startswith(k) and len(t) <= len(k) + 2 for t in tokens))
                )
                if hit:
                    puntajes[tema] += peso * (2 if " " in k else 1)
    if not puntajes:
        return ""
    mejor = max(puntajes.items(), key=lambda kv: kv[1])
    return mejor[0] if mejor[1] >= 2 else ""


def fallback_tema(subtema: str, contexto: str = "", titulo: str = "") -> str:
    tema = tema_por_taxonomia(subtema, contexto, titulo)
    if tema and validar_tema(tema):
        return tema
    return TEMA_POR_DEFECTO


def _recortar_sintagma(words: List[str], max_palabras: int) -> List[str]:
    """Cut a long noun phrase at a phrase boundary, never mid-complement.

    'Solidaridad de los caleños durante las labores de rescate humanitario
    en Cali' → stops before the last preposition that would be left
    dangling, so the result is still a complete phrase.
    """
    if len(words) <= max_palabras:
        out = list(words)
    else:
        adjuntos = {"en", "durante", "tras", "ante", "contra", "con", "para", "por",
                    "sobre", "hacia", "desde", "hasta", "y", "e", "segun", "según"}
        complementos = {"de", "del"}
        corte = None
        # Prefer dropping a whole adjunct ('… durante las labores …', '… en Cali')
        # over splitting a 'de' complement ('labores de rescate').
        for k in range(min(max_palabras, len(words) - 1), 2, -1):
            if unidecode(words[k].lower()) in adjuntos:
                corte = k
                break
        if corte is None:
            for k in range(min(max_palabras, len(words) - 1), 2, -1):
                if unidecode(words[k].lower()) in complementos:
                    corte = k
                    break
        out = words[:corte] if corte and corte >= 3 else words[:max_palabras]
    while out and unidecode(out[-1].lower().rstrip(".,;:")) in _TRAILING_INCOMPLETE | _QUESTION_TAILS | _DROP_LEAD | {"durante", "tras", "ante", "contra", "y", "e"}:
        out.pop()
    return out


def etiquetar_gramatical(contexto: str, titulo: str = "",
                         chat_fn: Optional[Callable] = None) -> Tuple[str, str]:
    sub = fallback_subtema(contexto, titulo)
    tema = fallback_tema(sub, contexto, titulo)
    if validar_subtema(sub) and validar_tema(tema):
        return tema, sub
    return (
        tema if validar_tema(tema) else TEMA_POR_DEFECTO,
        sub if validar_subtema(sub) else "Cobertura de información relevante",
    )


# ---------------------------------------------------------------------------
# Batched classification
# ---------------------------------------------------------------------------

def _parse_chat_payload(raw: Any) -> Dict[str, Any]:
    if raw is None:
        return {}
    if isinstance(raw, dict):
        return raw
    text = raw if isinstance(raw, str) else str(raw)
    text = text.strip()
    try:
        return json.loads(text)
    except Exception:
        m = re.search(r"\{.*\}", text, flags=re.S)
        if m:
            try:
                return json.loads(m.group(0))
            except Exception:
                return {}
    return {}


def clasificar_lotes(
    items: Sequence[Dict[str, Any]],
    marca: str,
    aliases=None,
    chat_fn: Optional[Callable] = None,
    batch_size: int = CHAT_BATCH_SIZE,
    call_counter: Optional[CallCounter] = None,
) -> List[Dict[str, str]]:
    """
    One JSON ChatCompletion per 25–40 items. Never one call per news/group
    unless a single leftover batch remains. Heuristic first; LLM only when
    the noun phrase cannot be formed grammatically. The model returns
    tono + subtema (tema is NOT requested: it is derived later from the
    consolidated subtema against the fixed ≤20-category taxonomy).
    """
    results: List[Dict[str, str]] = [{"tono": "Neutro", "tema": "", "subtema": ""} for _ in items]
    pending: List[int] = []
    for i, it in enumerate(items):
        ctx = str(it.get("contexto") or "")
        titulo = str(it.get("titulo") or "")
        tema, sub = etiquetar_gramatical(ctx, titulo, chat_fn=None)
        results[i]["tema"] = tema
        results[i]["subtema"] = sub
        results[i]["tono"] = it.get("tono") or "Neutro"
        if chat_fn is not None:
            # Every representative goes to the batch: tone is model-based and
            # the model's labels replace the heuristic ones when they validate.
            pending.append(i)

    if chat_fn is None or not pending:
        return results

    aliases_str = ", ".join(_lista_alias(marca, aliases)[1:6])
    # Model-side only: an ungrammatical fragment reaches the sheet only if
    # both the model and the deterministic fallback fail the validator.
    # Tema is NOT requested from the model anymore: it is derived from the
    # consolidated subtema against the fixed ≤20-category taxonomy, which is
    # the exact "subtemas → temas (≤20)" flow the client asked for.
    rejected: Dict[int, List[str]] = defaultdict(list)

    def _prompt(payload: List[Dict[str, Any]], reparacion: bool) -> str:
        cabecera = (
            f"Eres un analista senior de medios en Colombia. Marca analizada: '{marca}'"
            + (f" (alias: {aliases_str})" if aliases_str else "")
            + ".\nPara cada ítem el campo 'titulo' nombra el HECHO de la noticia y 'contexto' es el "
            "párrafo en el que se menciona a la marca (si la marca no aparece, es el resumen o el título). "
            "El subtema debe reflejar el hecho del TÍTULO, no un detalle técnico del contexto.\n"
        )
        formato = (
            "Devuelve SOLO JSON: {\"items\":[{\"id\":0,\"tono\":\"Positivo|Negativo|Neutro\","
            "\"subtema\":\"...\"}]}\n"
        )
        reglas = (
            "SUBTEMA (3 a 6 palabras, máximo 6): encabezado de reporte que dice DE QUÉ TRATÓ la "
            "noticia, frase nominal COMPLETA en español de Colombia, "
            "núcleo sustantivo + complemento con de/del/en/para/sobre/ante/por. Sin comas. "
            "No termine en número, cifra, millón(es), signo $, preposición ni artículo. "
            "Usa el TÍTULO para nombrar el hecho; el contexto solo aclara. "
            "PROHIBIDO: recortar el titular, copiar el objeto de un verbo ('Netflix por US$587 millones'), "
            "detalles de producto ('Múltiples y escalables a todas las fases de producción'), "
            "partícipios colgados ('Destacada entre las sociólogas… dentro'), verbos conjugados, "
            "colas interrogativas, prefijos de canal, y PROHIBIDO unir palabras sueltas con 'de' "
            "('Tecnología de innovación de obras', 'Aumento de tarifas de servicios').\n"
            "  Correcto: 'Adquisición de InterPositive por Netflix', "
            "'Solidaridad de caleños en rescate', "
            "'Inversión en vías del Cauca'.\n"
            "  Incorrecto: 'Netflix por US$587 millones', "
            "'Múltiples y escalables a todas las fases de producción', "
            "'Destacada entre las sociólogas más influyentes del país dentro', "
            "'Terremoto en colombia ascienden'.\n"
            "TONO: impacto reputacional DIRECTO sobre la marca o sus alias (no el tono general de la noticia). "
            "Positivo si la marca logra, gana, aporta, es reconocida; Negativo si es cuestionada, "
            "sancionada, afectada o responsable de un daño; Neutro si NO se menciona la marca, "
            "si solo se menciona de paso sin rol, o si el hecho no cambia su imagen.\n"
        )
        extra = ""
        if reparacion:
            extra = (
                "REPARACIÓN: el subtema de estos ítems fue RECHAZADO por incompleto, cortado o "
                "por ser una unión de palabras clave con 'de'. Redacta de nuevo SOLO el subtema como "
                "frase nominal completa basada en el título; el campo 'rechazadas' muestra lo que NO debes repetir.\n"
            )
        return cabecera + formato + reglas + extra + f"ÍTEMS:\n{json.dumps(payload, ensure_ascii=False)}"

    def _aplicar(i: int, row: Dict[str, Any]) -> bool:
        it = items[i]
        tono = str(row.get("tono") or results[i]["tono"]).strip().title()
        if tono not in ("Positivo", "Negativo", "Neutro"):
            tono = results[i]["tono"] if results[i]["tono"] in ("Positivo", "Negativo", "Neutro") else "Neutro"
        sub = _capitalizar(str(row.get("subtema") or "").strip().strip('"\'.'))
        ok_sub = validar_subtema(sub)
        if not ok_sub and sub:
            rejected[i].append(sub)
        results[i]["tono"] = tono
        if ok_sub:
            results[i]["subtema"] = sub
        return ok_sub

    def _run_batch(chunk_idx: List[int], reparacion: bool = False) -> List[int]:
        payload = []
        for i in chunk_idx:
            it = items[i]
            entry = {
                "id": i,
                "titulo": str(it.get("titulo") or "")[:220],
                "contexto": str(it.get("contexto") or "")[:900],
            }
            if reparacion:
                entry["rechazadas"] = rejected.get(i, [])[:4]
            payload.append(entry)
        if call_counter:
            with _COUNTER_LOCK:
                call_counter.chat += 1
                call_counter.chat_items += len(chunk_idx)
        fallidos: List[int] = []
        try:
            raw = chat_fn(_prompt(payload, reparacion))
            data = _parse_chat_payload(raw)
            rows = data.get("items") or data.get("resultados") or []
            by_id = {}
            for row in rows:
                try:
                    by_id[int(row.get("id"))] = row
                except Exception:
                    continue
            for i in chunk_idx:
                ok_sub = _aplicar(i, by_id.get(i, {}))
                if not ok_sub:
                    fallidos.append(i)
        except Exception:
            fallidos.extend(chunk_idx)
        return fallidos

    def _run_all(idx_list: List[int], reparacion: bool) -> List[int]:
        chunks = [idx_list[s:s + batch_size] for s in range(0, len(idx_list), batch_size)]
        if not chunks:
            return []
        workers = max(1, min(CHAT_PARALLEL_BATCHES, len(chunks)))
        fallidos: List[int] = []
        if workers == 1:
            for ch in chunks:
                fallidos.extend(_run_batch(ch, reparacion))
        else:
            # Bounded concurrency: a handful of batches in flight, never one
            # request per news item and never an unbounded retry storm.
            with ThreadPoolExecutor(max_workers=workers) as ex:
                for res in ex.map(lambda ch: _run_batch(ch, reparacion), chunks):
                    fallidos.extend(res)
        return fallidos

    fallidos = _run_all(pending, reparacion=False)
    # One bounded repair round, still batched, only for rejected labels.
    if fallidos:
        _run_all(sorted(set(fallidos)), reparacion=True)
    for i in range(len(items)):
        ctx = str(items[i].get("contexto") or "")
        tit = str(items[i].get("titulo") or "")
        if not validar_subtema(results[i]["subtema"]):
            results[i]["subtema"] = fallback_subtema(ctx, tit)
        if not validar_subtema(results[i]["subtema"]):
            results[i]["subtema"] = "Cobertura de información relevante"
        # Tema is always derived from the FINAL subtema against the fixed
        # taxonomy: this is what guarantees ≤20 stable themes per report.
        results[i]["tema"] = fallback_tema(results[i]["subtema"], ctx, tit)
        results[i]["tema"] = _capitalizar(results[i]["tema"])
        results[i]["subtema"] = _capitalizar(results[i]["subtema"])
    return results


def consolidar_temas_lote(labels: List[Dict[str, str]],
                          items: Optional[Sequence[Dict[str, Any]]] = None) -> List[Dict[str, str]]:
    """Subtemas → Temas (≤20), per client spec.

    1. Subtema canonicalization: near-identical subtema labels (same event
       wording that two different Grupo-noticia groups produced) collapse to
       the most frequent spelling, so "dos o más noticias iguales o
       similares o que comparten patrones" end up with the SAME Subtema.
       Merging is deliberately lexical (exact / near-exact wording, stem
       equality, no opposite-action pairs): embedding-based glue was removed
       because it merged distinct events that shared a paraphrase.
    2. Tema: every canonical subtema maps to ONE theme of the fixed
       taxonomy (17 categories + default = ≤18 ≤ 20). The taxonomy is the
       20-category cap "con base en los subtemas generados".
    """
    if not labels:
        return labels
    n = len(labels)
    subs = [str(l.get("subtema") or "") for l in labels]
    normed = [_norm_text(s) for s in subs]

    parent = list(range(n))

    def find(x: int) -> int:
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a: int, b: int) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for i in range(n):
        a = normed[i]
        if not a:
            continue
        for j in range(i + 1, n):
            if find(i) == find(j):
                continue
            b = normed[j]
            if not b:
                continue
            same = (a == b)
            if not same and len(a) >= 6 and len(b) >= 6:
                # Near-identical wording ('atencion de quejas de usuarios'
                # vs 'atencion de quejas de usuario'), not concept-glue.
                if _ratio(a, b) >= 0.96:
                    same = True
            if not same:
                # Same content stems: plural/tense/typography variants only.
                stems_a = _stems(a.split())
                stems_b = _stems(b.split())
                if stems_a and stems_a == stems_b:
                    same = True
            if same and not _hay_conflicto_accion(a, b):
                union(i, j)

    grupos = defaultdict(list)
    for i in range(n):
        grupos[find(i)].append(i)
    freq = Counter(subs)
    canon: Dict[int, str] = {}
    for root, idxs in grupos.items():
        vals = [subs[k] for k in idxs if subs[k]]
        if not vals:
            continue
        canon[root] = max(vals, key=lambda s: (freq.get(s, 0), len(s)))

    # Tema is per canonical subtema (one subtema ⇒ one tema), derived from
    # the member whose contexto/título best represent the merged group.
    tema_root: Dict[int, str] = {}
    for root, idxs in grupos.items():
        sub = canon.get(root, "")
        if not sub:
            continue
        best = None
        best_len = -1
        for i in idxs:
            if items is not None and i < len(items):
                ctx = str(items[i].get("contexto") or "")
                tit = str(items[i].get("titulo") or "")
                score = len(ctx) + len(tit)
                if score > best_len:
                    best_len = score
                    best = (ctx, tit)
        ctx, tit = best if best is not None else ("", "")
        tema_root[root] = _capitalizar(fallback_tema(sub, ctx, tit))

    for i in range(n):
        root = find(i)
        labels[i]["subtema"] = _capitalizar(canon.get(root, subs[i]))
        labels[i]["tema"] = tema_root.get(root, TEMA_POR_DEFECTO)
    return labels


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------

def _mejor_representante(idxs: Sequence[int], contextos: Sequence[str],
                         embeddings: Optional[Sequence[Any]] = None,
                         titulos: Optional[Sequence[Any]] = None) -> int:
    if not idxs:
        return 0
    con = [i for i in idxs if str(contextos[i] or "").strip()]
    pool = con or list(idxs)
    # The longest headline is the most complete naming of the fact
    # (prefix titles like '...compró Netflix' vs '...por US$587 millones').
    if titulos is not None:
        return max(
            pool,
            key=lambda i: (
                len(normalize_title_for_comparison(titulos[i])),
                len(str(contextos[i] or "")),
            ),
        )
    if embeddings is not None:
        vecs = [(i, embeddings[i]) for i in pool if embeddings[i] is not None]
        if len(vecs) >= 2:
            M = np.array([np.asarray(v, dtype=float).ravel() for _, v in vecs])
            centro = M.mean(axis=0)
            norms = np.linalg.norm(M, axis=1) * (np.linalg.norm(centro) or 1.0)
            sims = (M @ centro) / np.where(norms == 0, 1.0, norms)
            return vecs[int(np.argmax(sims))][0]
    return max(pool, key=lambda i: len(str(contextos[i] or "")))


def process_pipeline(
    rows: Sequence[Dict[str, Any]],
    marca: str,
    aliases=None,
    km: Optional[Dict[str, str]] = None,
    embed_fn: Optional[Callable[[List[str]], List[Any]]] = None,
    chat_fn: Optional[Callable] = None,
    pkl_tono_fn: Optional[Callable] = None,
    pkl_tema_fn: Optional[Callable] = None,
    progress: Optional[ProgressTracker] = None,
    on_stage: Optional[Callable[[Dict[str, Any]], None]] = None,
) -> Tuple[List[Dict[str, Any]], ProgressTracker]:
    """
    Canonical API_JC path:
    Limpieza → Duplicados → Contexto → Embedding único → Agrupación
    → Tono / Subtema (batched, one call per ~30 groups) → Subtemas
    canónicos (misma etiqueta para noticias iguales/similares) → Tema de
    taxonomía fija (≤20 categorías, derivado del subtema) → Excel caller.
    Duplicates are decided before the LLM (misma URL + misma mención para
    gráficas/Internet; misma mención + medio + hora para Radio/TV) and keep
    Tono IA = Duplicada.
    """
    km = km or KEYMAP
    progress = progress or ProgressTracker(on_stage=on_stage)
    if on_stage is not None and progress.on_stage is None:
        progress.on_stage = on_stage
    out = [dict(r) for r in rows]
    progress.stage("Limpieza", f"{len(out)} filas")

    out = detectar_duplicados(out, km)
    n_dup = sum(1 for r in out if r.get("is_duplicate"))
    progress.stage("Duplicados", f"{n_dup} duplicadas")

    titulos, resumenes, cuerpos = [], [], []
    for r in out:
        tit = r.get("_titulo_original", r.get(km["titulo"], ""))
        r["_titulo_original"] = titulo_original(tit)
        r[km["titulo"]] = r["_titulo_original"]
        titulos.append(r["_titulo_original"])
        resumenes.append(r.get(km["resumen"], ""))
        cuerpos.append(r.get(km.get("cuerpo", "Cuerpo Completo"), ""))

    matcher = _compile_alias_matcher(marca, aliases)
    for i, r in enumerate(out):
        # A duplicate is the SAME publication: reuse the original's brand
        # context instead of re-scanning its (identical) body.
        if r.get("is_duplicate"):
            o = r.get("_dup_of_index")
            if isinstance(o, int) and 0 <= o < i:
                src = out[o]
                r["Contexto analizado"] = src.get("Contexto analizado", "")
                r["Coincidencia marca"] = src.get("Coincidencia marca", "")
                r["Origen coincidencia"] = src.get("Origen coincidencia", "")
                continue
        meta = extraer_contexto_analizado(
            titulos[i], resumenes[i], marca, aliases, cuerpos[i],
            _matcher=matcher,
        )
        r["Contexto analizado"] = meta["contexto"]
        r["Coincidencia marca"] = meta["coincidencia"]
        r["Origen coincidencia"] = meta["origen"]
    contextos = [r["Contexto analizado"] for r in out]
    progress.stage("Contexto")

    embeddings = None
    texts_emb = [
        (contextos[i] or f"{titulos[i]}. {resumenes[i]}")[:1800]
        for i in range(len(out))
    ]
    if embed_fn is not None:
        # Embed each UNIQUE text once: duplicate rows and same-context
        # republications share a vector, so big dossiers send far fewer
        # embedding tokens to the API.
        unique_idx: Dict[str, int] = {}
        unique_texts: List[str] = []
        for t in texts_emb:
            if t not in unique_idx:
                unique_idx[t] = len(unique_texts)
                unique_texts.append(t)
        emb_unique = embed_fn(unique_texts) if unique_texts else []
        emb_by_text = dict(zip(unique_texts, emb_unique))
        embeddings = [emb_by_text.get(t) for t in texts_emb]
        progress.calls.embed += 1
        progress.calls.embed_items += len(unique_texts)
    progress.stage("Embedding único", f"{len(unique_texts) if embeddings else len(texts_emb)} textos únicos")

    grupos = agrupar_noticias_bloqueado(
        titulos, resumenes, contextos, embeddings,
        urls=[url_de_celda_link(r.get(km["link_nota"])) or url_de_celda_link(r.get(km["link_streaming"])) for r in out],
        fechas=[r.get(km["fecha"]) for r in out],
        horas=[r.get(km["hora"]) for r in out],
        counter=progress.comparisons,
    )
    # A duplicate always belongs to its original's group, and rows sharing
    # the same streaming link / Link Nota are the same publication.
    dsu2 = DSU(len(out))
    for idxs in grupos.values():
        for j in idxs[1:]:
            dsu2.union(idxs[0], j)
    for i, r in enumerate(out):
        o = r.get("_dup_of_index")
        if isinstance(o, int) and 0 <= o < len(out):
            dsu2.union(i, o)
    grupos = dsu2.grupos(len(out))
    gids = ids_grupo(grupos)
    for i, r in enumerate(out):
        r["Grupo noticia"] = gids[i]
    progress.stage("Agrupación", f"{len(grupos)} grupos · pares={progress.comparisons.n}")

    reps = []
    gid_of_rep = []
    for gid, idxs in grupos.items():
        ri = _mejor_representante(idxs, contextos, embeddings, titulos)
        # Prefer a non-duplicate representative when the group has one.
        nondup = [j for j in idxs if not out[j].get("is_duplicate")]
        if nondup:
            ri = _mejor_representante(nondup, contextos, embeddings, titulos)
        reps.append({
            "id": ri,
            "contexto": contextos[ri],
            "titulo": titulos[ri],
        })
        gid_of_rep.append(gid)

    labels = clasificar_lotes(
        reps, marca, aliases, chat_fn=chat_fn,
        call_counter=progress.calls,
    )
    # Canonical subtemas (same label for same/similar news) + tema from the
    # ≤20-category taxonomy, derived per subtema with the rep's context.
    labels = consolidar_temas_lote(labels, reps)
    # User-supplied .pkl models override the taxonomy defaults when present.
    if pkl_tono_fn is not None:
        try:
            tonos = pkl_tono_fn([x["contexto"] for x in reps])
            for i, t in enumerate(tonos or []):
                val = t.get("tono", t) if isinstance(t, dict) else t
                if val:
                    labels[i]["tono"] = str(val)
        except Exception:
            pass
    if pkl_tema_fn is not None:
        try:
            temas = pkl_tema_fn([x["contexto"] for x in reps])
            for i, t in enumerate(temas or []):
                if t:
                    labels[i]["tema"] = str(t)
        except Exception:
            pass

    by_gid = {gid: labels[k] for k, gid in enumerate(gid_of_rep)}
    progress.stage("Tono", f"{progress.calls.chat} llamadas chat")
    progress.stage("Tema/Subtema", f"{len(grupos)} grupos cacheados")

    for gid, idxs in grupos.items():
        lab = by_gid.get(gid) or {"tono": "Neutro", "tema": TEMA_POR_DEFECTO,
                                  "subtema": "Cobertura de información relevante"}
        for i in idxs:
            if out[i].get("is_duplicate"):
                out[i]["Tono IA"] = "Duplicada"
                out[i]["Tema"] = "-"
                out[i]["Subtema"] = "-"
                continue
            out[i]["Tono IA"] = lab["tono"]
            out[i]["Tema"] = lab["tema"]
            out[i]["Subtema"] = lab["subtema"]
    return out, progress


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generate_output_excel(rows: Sequence[Dict[str, Any]],
                          km: Optional[Dict[str, str]] = None) -> bytes:
    km = km or KEYMAP
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado"
    ws.append(list(OUTPUT_COLUMNS))
    font_header = Font(bold=True)
    font_hyperlink = Font(color="000000", underline=None)
    align_left = Alignment(horizontal="left")
    for i, _ in enumerate(OUTPUT_COLUMNS, start=1):
        ws.cell(row=1, column=i).font = font_header

    for row in rows:
        out_vals = []
        links = {}
        for ci, h in enumerate(OUTPUT_COLUMNS, start=1):
            if h == "Título":
                val = row.get("_titulo_original", row.get(h))
                val = titulo_original(val)
            else:
                val = row.get(h)
            if h == "Fecha" and val is not None and hasattr(val, "to_pydatetime"):
                val = val.to_pydatetime()
            if isinstance(val, dict) and "url" in val:
                cv = val.get("value", "Link")
                if val.get("url"):
                    links[ci] = val["url"]
                out_vals.append(cv)
            elif isinstance(val, str) and val.startswith("http"):
                out_vals.append("Link")
                links[ci] = val
            else:
                out_vals.append(val)
        ws.append(out_vals)
        current = ws.max_row
        for ci, url in links.items():
            cell = ws.cell(row=current, column=ci)
            cell.hyperlink = url
            cell.font = font_hyperlink
            cell.alignment = align_left
        date_cell = ws.cell(row=current, column=OUTPUT_COLUMNS.index("Fecha") + 1)
        if isinstance(date_cell.value, (datetime.datetime, datetime.date)):
            date_cell.number_format = "DD/MM/YYYY"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_titulo_and_links_from_xlsx(data: bytes) -> List[Dict[str, Any]]:
    wb = load_workbook(io.BytesIO(data), data_only=False)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for excel_row in ws.iter_rows(min_row=2):
        item = {}
        for h, cell in zip(headers, excel_row):
            if h == "Título":
                item["Título"] = cell.value
            if h in ("Link Nota", "Link (Streaming - Imagen)"):
                item[h] = valor_con_hipervinculo(cell)
            else:
                item.setdefault(h, cell.value)
        rows.append(item)
    return rows
