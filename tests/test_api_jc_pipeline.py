"""Regression tests for the API_JC repair (titles, BUSCARV, dups, groups, labels)."""
from __future__ import annotations

import io
import json
import time
import unittest
from copy import deepcopy

from openpyxl import Workbook, load_workbook

import api_jc_core as core


VIDEO_TITLE = (
    "Video | Topos Azteca destacan la solidaridad de los caleños "
    "durante las labores de rescate"
)

BUSCARV_FIXTURE = [
    ("el pais", "Nacional"),
    ("caracol radio", "Bogotá"),
    ("rcn radio", "Bogotá"),
    ("semana", "Nacional"),
    ("el tiempo", "Nacional"),
]
BUSCARV_INTERNET = [
    ("elpais.com.co", "El País"),
    ("semana.com", "Semana"),
    ("eltiempo.com", "El Tiempo"),
]


def _row(**kw):
    base = {
        "ID Noticia": kw.pop("id", "1"),
        "Fecha": kw.pop("fecha", "2026-01-15"),
        "Hora": kw.pop("hora", "08:00"),
        "Medio": kw.pop("medio", "El País"),
        "Tipo de Medio": kw.pop("tipo", "Internet"),
        "Sección - Programa": "Nacional",
        "Región": "Nacional",
        "Título": kw.pop("titulo", VIDEO_TITLE),
        "_titulo_original": kw.pop("titulo_original", None),
        "Tono IA": "",
        "Tema": "",
        "Subtema": "",
        "Link Nota": kw.pop("link", {"value": "Link", "url": "https://elpais.com.co/a"}),
        "Resumen - Aclaracion": kw.pop("resumen", "Resumen de la nota."),
        "Link (Streaming - Imagen)": None,
        "Menciones - Empresa": kw.pop("mencion", "Topos Azteca"),
        "ID duplicada": "",
        "Cuerpo Completo": kw.pop("cuerpo", ""),
        "Tono": "Neutro",
        "is_duplicate": False,
    }
    if base["_titulo_original"] is None:
        base["_titulo_original"] = base["Título"]
    base.update(kw)
    return base


class TestTitleUnchanged(unittest.TestCase):
    def test_01_video_pipe_title_byte_for_byte(self):
        rows = [_row(id="100", titulo=VIDEO_TITLE)]
        out, _ = core.process_pipeline(rows, "Topos Azteca", ["Topos"])
        self.assertEqual(out[0]["Título"], VIDEO_TITLE)
        xlsx = core.generate_output_excel(out)
        wb = load_workbook(io.BytesIO(xlsx))
        headers = [c.value for c in wb.active[1]]
        col = headers.index("Título") + 1
        self.assertEqual(wb.active.cell(2, col).value, VIDEO_TITLE)
        self.assertNotIn("|", core.normalize_title_for_comparison(VIDEO_TITLE) or "|")
        # Internal normalize may drop punctuation, output must not.
        self.assertTrue(VIDEO_TITLE.startswith("Video |"))


class TestBuscarv(unittest.TestCase):
    def test_02_buscarv_fixture_identical_mapping(self):
        region_map = core.construir_mapa_buscarv(BUSCARV_FIXTURE)
        internet_map = core.construir_mapa_buscarv(BUSCARV_INTERNET)
        self.assertEqual(core.aplicar_buscarv("El Pais", region_map), "Nacional")
        self.assertEqual(core.aplicar_buscarv("CARACOL RADIO", region_map), "Bogotá")
        self.assertEqual(core.aplicar_buscarv("medio-fantasma", region_map), "N/A")
        regiones, medios = core.aplicar_buscarv_dossier(
            ["El Pais", "elpais.com.co", "Caracol Radio"],
            region_map, internet_map,
            tipos=["Internet", "Internet", "Radio"],
        )
        self.assertEqual(regiones, ["Nacional", "N/A", "Bogotá"])
        self.assertEqual(medios[0], "El Pais")  # no Internet alias for this key
        self.assertEqual(medios[1], "El País")  # BUSCARV Internet
        self.assertEqual(medios[2], "Caracol Radio")  # radio unchanged
        # Control/branch name remains BUSCARV
        self.assertTrue(hasattr(core, "aplicar_buscarv"))
        self.assertTrue(hasattr(core, "construir_mapa_buscarv"))


class TestDuplicatesVsGrupo(unittest.TestCase):
    def test_03_internet_same_title_different_urls_not_duplicate(self):
        rows = [
            _row(id="10", titulo=VIDEO_TITLE, medio="El País",
                 link={"value": "Link", "url": "https://elpais.com.co/solidaridad-cali"}),
            _row(id="11", titulo=VIDEO_TITLE, medio="Semana",
                 link={"value": "Link", "url": "https://semana.com/solidaridad-cali"}),
        ]
        out, _ = core.process_pipeline(rows, "Topos Azteca", ["Topos"])
        self.assertFalse(out[0].get("is_duplicate"))
        self.assertFalse(out[1].get("is_duplicate"))
        self.assertEqual(out[1].get("ID duplicada") or "", "")
        self.assertEqual(out[0]["Grupo noticia"], out[1]["Grupo noticia"])
        self.assertEqual(out[0]["Tono IA"], out[1]["Tono IA"])
        self.assertEqual(out[0]["Tema"], out[1]["Tema"])
        self.assertEqual(out[0]["Subtema"], out[1]["Subtema"])
        self.assertNotEqual(out[0]["Tono IA"], "Duplicada")

    def test_04_internet_same_embedded_url_is_duplicate(self):
        url = "https://WWW.ElPais.com.co/nota/abc/?utm_source=tw&utm_campaign=x#frag"
        url2 = "https://elpais.com.co/nota/abc/"
        rows = [
            _row(id="20", titulo="Primera publicación",
                 link={"value": "Link", "url": url}),
            _row(id="21", titulo="Republicación distinta",
                 link={"value": "Link", "url": url2}),
        ]
        out = core.detectar_duplicados(rows)
        self.assertFalse(out[0].get("is_duplicate"))
        self.assertTrue(out[1].get("is_duplicate"))
        self.assertEqual(str(out[1]["ID duplicada"]), "20")
        self.assertEqual(out[1]["Tono IA"], "Duplicada")
        xlsx = core.generate_output_excel(out)
        wb = load_workbook(io.BytesIO(xlsx))
        headers = [c.value for c in wb.active[1]]
        link_col = headers.index("Link Nota") + 1
        cell = wb.active.cell(2, link_col)
        self.assertTrue(cell.hyperlink)
        self.assertIn("elpais.com.co/nota/abc", cell.hyperlink.target.lower())

    def test_05_radio_tv_same_mencion_medio_hora_duplicate(self):
        title_a = "Alcalde presenta balance de seguridad en Cali"
        title_b = "Titular completamente distinto sobre otro asunto"
        rows = [
            _row(id="30", tipo="Radio", medio="Caracol Radio", fecha="2026-03-01",
                 hora="07:00", titulo=title_a, link=None),
            # same mención + same medio + same hora, different title → duplicate
            _row(id="31", tipo="Radio", medio="Caracol Radio", fecha="2026-03-01",
                 hora="07:00", titulo=title_b, link=None),
            # same title, same hora, different medio → NOT duplicate (Grupo only)
            _row(id="32", tipo="Televisión", medio="Telepacífico", fecha="2026-03-01",
                 hora="07:00", titulo=title_a, link=None),
            # same medio, different hora → NOT duplicate
            _row(id="33", tipo="Radio", medio="Caracol Radio", fecha="2026-03-01",
                 hora="09:30", titulo=title_a, link=None),
            # same medio + hora but different mención → NOT duplicate
            _row(id="34", tipo="Radio", medio="Caracol Radio", fecha="2026-03-01",
                 hora="07:00", titulo=title_a, link=None, mencion="Otra Empresa"),
        ]
        out = core.detectar_duplicados(rows)
        self.assertFalse(out[0]["is_duplicate"])
        self.assertTrue(out[1]["is_duplicate"], "same mención+medio+hora is duplicate even with another title")
        self.assertEqual(str(out[1]["ID duplicada"]), "30")
        self.assertFalse(out[2]["is_duplicate"], "different Medio is not duplicate")
        self.assertFalse(out[3]["is_duplicate"], "different Hora is not duplicate")
        self.assertFalse(out[4]["is_duplicate"], "different mención is not duplicate")
        # Same title across media still shares a Grupo noticia and labels.
        out2, _ = core.process_pipeline([rows[0], rows[2]], "Alcaldía de Cali")
        self.assertEqual(out2[0]["Grupo noticia"], out2[1]["Grupo noticia"])
        self.assertEqual(out2[0]["Subtema"], out2[1]["Subtema"])

    def test_05b_streaming_link_same_mencion_duplicate_any_title(self):
        rows = [
            _row(id="35", titulo="Primer titular sobre la marca", link=None,
                 **{"Link (Streaming - Imagen)": {"value": "Link", "url": "https://stream.tv/clip/77"}}),
            _row(id="36", titulo="Titular totalmente diferente", link=None,
                 **{"Link (Streaming - Imagen)": {"value": "Link", "url": "https://stream.tv/clip/77?utm_source=x"}}),
            _row(id="37", titulo="Primer titular sobre la marca", link=None, mencion="Otra Empresa",
                 **{"Link (Streaming - Imagen)": {"value": "Link", "url": "https://stream.tv/clip/77"}}),
        ]
        out = core.detectar_duplicados(rows)
        self.assertFalse(out[0]["is_duplicate"])
        self.assertTrue(out[1]["is_duplicate"], "same streaming link + same mención → duplicate")
        self.assertEqual(str(out[1]["ID duplicada"]), "35")
        self.assertFalse(out[2]["is_duplicate"], "different mención keeps the row")


class TestGroupingAndLabels(unittest.TestCase):
    def test_06_similar_resumen_same_grupo_and_labels(self):
        r1 = (
            "Las autoridades confirmaron el aumento de heridos tras el "
            "terremoto que afectó el suroccidente del país."
        )
        r2 = (
            "Las autoridades confirmaron el aumento de heridos tras el "
            "terremoto que sacudió el suroccidente colombiano."
        )
        rows = [
            _row(id="40", titulo="Balance oficial tras el sismo", resumen=r1,
                 link={"value": "Link", "url": "https://a.com/1"}),
            _row(id="41", titulo="Nuevas cifras del movimiento telúrico", resumen=r2,
                 link={"value": "Link", "url": "https://b.com/2"}),
        ]
        out, _ = core.process_pipeline(rows, "Cruz Roja", ["Cruz Roja Colombiana"])
        self.assertEqual(out[0]["Grupo noticia"], out[1]["Grupo noticia"])
        self.assertEqual(out[0]["Tono IA"], out[1]["Tono IA"])
        self.assertEqual(out[0]["Tema"], out[1]["Tema"])
        self.assertEqual(out[0]["Subtema"], out[1]["Subtema"])
        self.assertFalse(out[0]["is_duplicate"])
        self.assertFalse(out[1]["is_duplicate"])

    def test_06b_title_variants_same_group_same_labels(self):
        base = "Topos Azteca destacan la solidaridad de los caleños durante las labores de rescate"
        variants = [
            f"Video | {base}",
            f"{base} | El País",
            "Topos Azteca destacan solidaridad de caleños en labores de rescate",
            "Los Topos Azteca destacaron la solidaridad de los caleños en el rescate",
        ]
        rows = [
            _row(id=str(60 + k), titulo=t, medio=f"Medio{k}",
                 resumen=f"Texto propio número {k} sobre la brigada en Cali.",
                 link={"value": "Link", "url": f"https://m{k}.com/nota-{k}"})
            for k, t in enumerate(variants)
        ]
        rows.append(_row(id="70", titulo="Gobernador anuncia inversión en vías terciarias del Cauca",
                         resumen="La Gobernación del Cauca invertirá en vías.",
                         link={"value": "Link", "url": "https://m9.com/vias"}))
        out, _ = core.process_pipeline(rows, "Topos Azteca", ["Topos"])
        gids = {r["Grupo noticia"] for r in out[:4]}
        self.assertEqual(len(gids), 1, f"title variants must share one group: {[r['Grupo noticia'] for r in out]}")
        self.assertNotEqual(out[4]["Grupo noticia"], out[0]["Grupo noticia"])
        self.assertEqual(len({r["Tono IA"] for r in out[:4]}), 1)
        self.assertEqual(len({r["Tema"] for r in out[:4]}), 1)
        self.assertEqual(len({r["Subtema"] for r in out[:4]}), 1)
        for r in out[:4]:
            self.assertFalse(r["is_duplicate"], "different URLs are not duplicates")
        # Output titles are still the originals.
        self.assertEqual(out[0]["Título"], variants[0])
        self.assertEqual(out[1]["Título"], variants[1])

    def test_07_bad_labels_rejected_grammatical_replacements(self):
        bad = [
            "Terremoto en colombia ascienden",
            "Sismo en cali cómo saber si",
            "Alcalde anunció obras en Cali",
            "Diporto resultados jornada fecha liga",
        ]
        for et in bad:
            self.assertFalse(core.validar_subtema(et), et)
        for tema in ["Agenda tendrá", "Balance seguridad", "Sismo registra", "Anunció inversión"]:
            self.assertFalse(core.validar_tema(tema), tema)
        for tema in ["Seguridad y orden público", "Infraestructura y movilidad", "Salud pública",
                     "Gestión pública local", "Emergencias y desastres naturales"]:
            self.assertTrue(core.validar_tema(tema), tema)
        self.assertEqual(core.fallback_tema("Agenda del concejo", "La agenda del concejo tendrá debates.", "Agenda tendrá"),
                         "Política y gobierno")
        self.assertEqual(core.fallback_tema("Inversión en vías terciarias del Cauca", "", ""),
                         "Infraestructura y movilidad")
        # Long headline object is trimmed at a phrase boundary, never mid-complement.
        s = core.fallback_subtema(
            "Los Topos Azteca llegaron.",
            "Video | Topos Azteca destacan la solidaridad de los caleños durante las labores de rescate humanitario en Cali",
        )
        self.assertTrue(core.validar_subtema(s), s)
        self.assertTrue(s.endswith("rescate humanitario") or s.endswith("caleños") or s.endswith("de rescate"), s)
        self.assertNotIn("labores", s.split()[-1], "must not end on a dangling noun of a split complement")
        ctx1 = "Tras el terremoto en Colombia ascienden las víctimas y los heridos."
        ctx2 = "Sismo en Cali: cómo saber si la vivienda quedó bien y qué recomendaciones seguir."
        s1 = core.fallback_subtema(ctx1, "Terremoto en colombia ascienden")
        s2 = core.fallback_subtema(ctx2, "Sismo en cali cómo saber si")
        self.assertTrue(core.validar_subtema(s1), s1)
        self.assertTrue(core.validar_subtema(s2), s2)
        self.assertNotIn("ascienden", s1.lower())
        self.assertNotIn("cómo saber", s2.lower())
        self.assertIn("víctimas", s1.lower())
        self.assertTrue("sismo" in s2.lower() or "cali" in s2.lower())

    def test_08_diporto_no_keyword_collage(self):
        collage = "Diporto resultados jornada fecha liga"
        self.assertFalse(core.validar_subtema(collage))
        sub = core.fallback_subtema(
            "Diporto entregó los resultados de la jornada y la fecha de liga.",
            collage,
        )
        self.assertTrue(core.validar_subtema(sub), sub)
        self.assertNotEqual(core._norm_text(sub), core._norm_text(collage))
        self.assertTrue(any(n in sub.lower() for n in ("de", "del", "en", "sobre")))


class TestOutputSchema(unittest.TestCase):
    def test_09_exact_output_columns_no_extras(self):
        rows = [_row(id="50", titulo=VIDEO_TITLE)]
        out, _ = core.process_pipeline(rows, "Topos Azteca")
        xlsx = core.generate_output_excel(out)
        wb = load_workbook(io.BytesIO(xlsx))
        headers = [c.value for c in wb.active[1]]
        self.assertEqual(headers, core.OUTPUT_COLUMNS)
        self.assertNotIn("Autor - Conductor", headers)
        self.assertNotIn("CPE", headers)
        self.assertNotIn("Nro. Pagina", headers)


class TestScaleAndBatching(unittest.TestCase):
    def test_10_400_rows_blocked_not_n2_bounded_chat(self):
        n = 400
        rows = []
        import random
        rnd = random.Random(7)
        vocab = ("terremoto cali colombia sismo victimas heridos autoridades alcaldia "
                 "rescate solidaridad balance emergencia ciudad barrio vivienda familias "
                 "damnificados bomberos gobierno informe cifras reporte").split()
        for i in range(n):
            # Realistic ~1.5k-char resúmenes sharing vocabulary: this is what
            # froze the UI with char-level SequenceMatcher and per-pair cosine.
            long_res = " ".join(rnd.choice(vocab) for _ in range(250)) + f" ZX{i:04d}."
            rows.append(_row(
                id=str(1000 + i),
                titulo=f"Sismo en Cali: {rnd.choice(vocab)} {rnd.choice(vocab)} {i}",
                resumen=long_res,
                fecha="2026-04-01",
                hora=f"{(i % 20) + 1:02d}:00",
                medio=f"Medio{i}",
                link={"value": "Link", "url": f"https://noticias.example/{i}/nota-{i}"},
                mencion="MarcaX",
            ))
        # Two extra rows that SHOULD group (similar resumen) and one URL dup.
        rows[10]["Resumen - Aclaracion"] = rows[11]["Resumen - Aclaracion"] = (
            "Las autoridades confirmaron el aumento de heridos tras el terremoto andino."
        )
        rows[10]["Título"] = "Balance oficial del terremoto andino"
        rows[11]["Título"] = "Cifras actualizadas del terremoto andino"
        rows[20]["Link Nota"] = deepcopy(rows[21]["Link Nota"])

        chat_calls = []

        def chat_fn(prompt):
            chat_calls.append(prompt)
            # Count items in this batch.
            try:
                start = prompt.index("ÍTEMS:")
                payload = json.loads(prompt[start + 6:])
            except Exception:
                payload = []
            items = []
            for it in payload:
                items.append({
                    "id": it["id"],
                    "tono": "Neutro",
                    "tema": "Emergencias y desastres naturales",
                    "subtema": "Balance de daños del sismo en Cali",
                })
            return {"items": items}

        embed_calls = []

        def embed_fn(textos):
            embed_calls.append(len(textos))
            rng = __import__("numpy").random.RandomState(0)
            return [rng.randn(1536) for _ in textos]

        t0 = time.time()
        out, prog = core.process_pipeline(
            rows, "MarcaX", embed_fn=embed_fn, chat_fn=chat_fn,
        )
        cpu = time.time() - t0
        self.assertLess(cpu, 8.0, f"grouping/pipeline CPU too slow: {cpu:.2f}s")
        self.assertEqual(len(embed_calls), 1, "must be a single embedding pass")
        self.assertEqual(embed_calls[0], n)
        max_pairs = n * (n - 1) // 2
        self.assertLess(prog.comparisons.n, n * 40)
        self.assertLess(prog.comparisons.n, max_pairs // 4)
        agr = next(e for e in prog.events if e["stage"] == "Agrupación")
        ctx = next(e for e in prog.events if e["stage"] == "Contexto")
        self.assertLess(agr["stage_s"], 1.5, f"Agrupación itself too slow: {agr['stage_s']:.2f}s")
        self.assertLess(ctx["stage_s"], 2.0, f"Contexto itself too slow: {ctx['stage_s']:.2f}s")
        self.assertIn("total", agr["label"])
        # Valid model output ⇒ one batched pass, no repair round.
        self.assertLessEqual(len(chat_calls), (n + core.CHAT_BATCH_SIZE - 1) // core.CHAT_BATCH_SIZE + 1)
        self.assertGreaterEqual(len(chat_calls), 1)
        self.assertTrue(all(r["Subtema"] in ("Balance de daños del sismo en Cali", "-") for r in out))
        for prompt in chat_calls:
            # Each request is a batch, not one news item.
            self.assertIn("ÍTEMS:", prompt)
        dups = [r for r in out if r.get("is_duplicate")]
        self.assertTrue(any(str(r.get("ID duplicada")) == str(rows[20]["ID Noticia"]) or
                            str(r.get("ID duplicada")) == str(rows[21]["ID Noticia"])
                            for r in dups))
        self.assertEqual(out[10]["Grupo noticia"], out[11]["Grupo noticia"])
        stages = [e["stage"] for e in prog.events]
        for required in ("Limpieza", "Duplicados", "Contexto", "Embedding único",
                         "Agrupación", "Tono", "Tema/Subtema"):
            self.assertIn(required, stages)


class TestGroupingSpeedRealistic(unittest.TestCase):
    def test_11_agrupacion_410_clustered_under_one_second(self):
        """Reproduce the 410-row Streamlit dossier: long resúmenes, clustered
        embeddings, shared dates. Agrupación must stay well under a second.
        """
        n = 410
        rnd = __import__("random").Random(3)
        rng = __import__("numpy").random.RandomState(3)
        vocab = ("terremoto cali colombia sismo victimas heridos autoridades alcaldia "
                 "rescate solidaridad balance emergencia ciudad barrio vivienda familias "
                 "damnificados bomberos gobierno informe cifras reporte topos azteca").split()
        titulos, resumenes, contextos, urls, fechas, horas = [], [], [], [], [], []
        clusters = [rng.randn(32).astype("float32") for _ in range(40)]
        embeddings = []
        for i in range(n):
            titulos.append(f"Sismo en Cali: {rnd.choice(vocab)} {rnd.choice(vocab)} {i}")
            resumenes.append(" ".join(rnd.choice(vocab) for _ in range(220)) + f" ZX{i:04d}.")
            contextos.append(resumenes[-1][:400])
            urls.append(f"https://noticias.example/{i}/nota-{i}")
            fechas.append("2026-04-01")
            horas.append(f"{(i % 18) + 6:02d}:00")
            embeddings.append(clusters[i % 40] + 0.08 * rng.randn(32).astype("float32"))
        # Known same-fact pair (paraphrased titles + similar resumen).
        titulos[10] = "Topos Azteca destacan la solidaridad de los caleños durante las labores de rescate"
        titulos[11] = "Los Topos Azteca destacaron la solidaridad de los caleños en el rescate"
        resumenes[10] = resumenes[11] = (
            "Las autoridades confirmaron el aumento de heridos tras el terremoto andino."
        )
        embeddings[11] = embeddings[10]

        t0 = time.time()
        grupos = core.agrupar_noticias_bloqueado(
            titulos, resumenes, contextos, embeddings, urls, fechas, horas,
            counter=core.ComparisonCounter(),
        )
        dt = time.time() - t0
        self.assertLess(dt, 1.0, f"agrupar_noticias_bloqueado took {dt:.2f}s")
        by_row = {}
        for g, idxs in grupos.items():
            for i in idxs:
                by_row[i] = g
        self.assertEqual(by_row[10], by_row[11])
        self.assertGreaterEqual(len(grupos), 40, "must not collapse the whole dossier into few groups")

    def test_12_contexto_skips_cuerpo_when_title_hits(self):
        cuerpo = ("palabra " * 4000) + "Relleno sin marca en el cuerpo largo."
        t0 = time.time()
        for _ in range(80):
            meta = core.extraer_contexto_analizado(
                "Topos Azteca destacan la solidaridad",
                "Resumen corto de la nota.",
                "Topos Azteca",
                ["Topos"],
                cuerpo,
            )
        dt = time.time() - t0
        self.assertLess(dt, 1.2, f"contexto with long cuerpo too slow: {dt:.2f}s")
        self.assertIn("Topos", meta["contexto"])
        self.assertEqual(meta["origen"], "Título")


class TestUrlAndHyperlinkHelpers(unittest.TestCase):
    def test_url_normalization_keeps_distinct_paths(self):
        a = core.normalize_url("https://WWW.Site.com/a/b/?utm_source=x#z")
        b = core.normalize_url("http://site.com/a/b")
        c = core.normalize_url("https://site.com/a/c/")
        self.assertEqual(a, b)
        self.assertNotEqual(a, c)

    def test_hyperlink_roundtrip_from_workbook(self):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Link Nota"
        ws["B1"] = "Título"
        cell = ws["A2"]
        cell.value = "Link"
        cell.hyperlink = "https://ejemplo.com/nota-1"
        ws["B2"] = VIDEO_TITLE
        buf = io.BytesIO()
        wb.save(buf)
        parsed = core.read_titulo_and_links_from_xlsx(buf.getvalue())
        self.assertEqual(parsed[0]["Título"], VIDEO_TITLE)
        self.assertEqual(core.url_de_celda_link(parsed[0]["Link Nota"]),
                         "https://ejemplo.com/nota-1")


if __name__ == "__main__":
    unittest.main()
